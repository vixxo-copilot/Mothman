"""Build KS101108 SR + invoice status Excel report via Gateway + VixxoLink MCP."""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
VETTING_SCRIPTS = SCRIPT_DIR.parent / ".agents" / "skills" / "sp-inbound-vetting" / "scripts"
sys.path.insert(0, str(VETTING_SCRIPTS))

from mcp_http import mcp_call, mcp_result_text  # noqa: E402

GATEWAY_URL = "https://vixxonow.com/mcp/gateway"
VIXXOLINK_URL = "https://vixxonow.com/mcp/vixxolink"

SP_NUMBER = "KS101108"
SP_NAME = "KS - Montgomery Lock & Key Inc."
OUT = SCRIPT_DIR / "KS101108_SR_Invoice_Report.xlsx"

HEADERS = [
    "SR #",
    "SR Status",
    "Sub-Status",
    "Customer #",
    "Customer Name",
    "Site City",
    "Site State",
    "Line of Service",
    "Short Description",
    "Created Date",
    "WO #",
    "Invoice # (VIID)",
    "Invoice Status",
    "Consolidated Status",
    "SP Invoice #",
    "SP Amount",
    "Customer Amount",
    "Invoice Created",
    "Invoice Last Updated",
    "Invoice Count",
    "SP Name",
    "SP #",
]

STATUS_FILLS = {
    "Rejected": PatternFill("solid", fgColor="FFC7CE"),
    "Review": PatternFill("solid", fgColor="FFEB9C"),
    "In Progress": PatternFill("solid", fgColor="DDEBF7"),
    "Paid": PatternFill("solid", fgColor="C6EFCE"),
}


def parse_json_blob(text: str):
    text = (text or "").strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}|\[.*\]", text, re.S)
        if m:
            try:
                return json.loads(m.group(0))
            except json.JSONDecodeError:
                return None
    return None


def unwrap_data(payload) -> dict:
    if isinstance(payload, dict) and isinstance(payload.get("data"), dict):
        return payload["data"]
    return payload if isinstance(payload, dict) else {}


def sr_rows(data) -> list[dict]:
    if not isinstance(data, dict):
        return []
    for key in ("results", "serviceRequestList", "serviceRequests", "items"):
        val = data.get(key)
        if isinstance(val, list):
            return val
    return []


def invoice_rows(data) -> list[dict]:
    if not isinstance(data, dict):
        return []
    if isinstance(data.get("invoiceList"), list):
        return data["invoiceList"]
    nested = data.get("data")
    if isinstance(nested, dict) and isinstance(nested.get("invoiceList"), list):
        return nested["invoiceList"]
    return []


def nested_get(row: dict, *keys: str):
    for key in keys:
        if row.get(key) not in (None, ""):
            return row.get(key)
    return ""


def sp_number_from_row(row: dict) -> str:
    sp = row.get("serviceProvider") or {}
    if isinstance(sp, dict):
        return str(sp.get("number") or "")
    return str(
        row.get("serviceProviderNumber")
        or row.get("siebelServiceProviderNum")
        or row.get("serviceProvider.number")
        or ""
    )


def fetch_srs() -> list[dict]:
    resp = mcp_call(
        VIXXOLINK_URL,
        "vixxolink_search_service_requests",
        {"searchString": SP_NUMBER, "pageSize": 100, "pageNumber": 1, "summary": True},
    )
    text = mcp_result_text(resp)
    if text.startswith("HTTP ") or text.startswith("MCP stdio"):
        raise RuntimeError(f"VixxoLink SR search failed: {text[:300]}")
    data = unwrap_data(parse_json_blob(text))
    rows = sr_rows(data)
    filtered = [r for r in rows if sp_number_from_row(r) == SP_NUMBER]
    if not filtered and rows:
        # searchString already scoped to SP; accept rows with matching SP name
        filtered = [
            r for r in rows
            if SP_NAME.lower() in str(nested_get(r, "serviceProvider.name", "serviceProviderName", "serviceProvider.name")).lower()
            or "Montgomery Lock" in str(nested_get(r, "serviceProvider.name", "serviceProviderName")).lower()
        ]
    return filtered or rows


def fetch_invoices() -> list[dict]:
    resp = mcp_call(GATEWAY_URL, "gateway_search_invoices", {"serviceProviderNumber": SP_NUMBER})
    text = mcp_result_text(resp)
    if text.startswith("HTTP ") or text.startswith("MCP stdio"):
        raise RuntimeError(f"Gateway invoice search failed: {text[:300]}")
    return invoice_rows(parse_json_blob(text))


def parse_dt(raw) -> datetime | None:
    if not raw:
        return None
    raw = str(raw).replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


def customer_name_from_sr(sr: dict) -> str:
    cust = sr.get("customer") or {}
    if isinstance(cust, dict):
        return str(cust.get("name") or cust.get("number") or "")
    return str(sr.get("customer.name") or sr.get("customerNumber") or "")


def customer_number_from_sr(sr: dict) -> str:
    cust = sr.get("customer") or {}
    if isinstance(cust, dict):
        return str(cust.get("number") or "")
    return str(sr.get("customer.number") or sr.get("customerNumber") or "")


def site_field(sr: dict, field: str) -> str:
    site = sr.get("site") or {}
    if isinstance(site, dict) and site.get(field):
        return str(site.get(field))
    dotted = f"site.{field}"
    return str(sr.get(dotted) or "")


def pick_latest_invoice(invoices: list[dict]) -> dict | None:
    if not invoices:
        return None
    return max(
        invoices,
        key=lambda inv: parse_dt(inv.get("lastUpdatedDate")) or parse_dt(inv.get("createdDate")) or datetime.min,
    )


def status_fill(status: str) -> PatternFill | None:
    status = status or ""
    for key, fill in STATUS_FILLS.items():
        if key.lower() in status.lower():
            return fill
    return None


def build_rows(srs: list[dict], invoices: list[dict]) -> list[dict]:
    by_sr: dict[str, list[dict]] = defaultdict(list)
    for inv in invoices:
        sr_num = str(inv.get("serviceRequestNumber") or "")
        if sr_num:
            by_sr[sr_num].append(inv)

    report_rows: list[dict] = []
    for sr in sorted(srs, key=lambda r: str(nested_get(r, "localCreatedDate", "createdDate", "srCreatedDate")), reverse=True):
        sr_num = str(nested_get(sr, "number", "serviceRequestNumber"))
        invs = by_sr.get(sr_num, [])
        latest = pick_latest_invoice(invs)
        report_rows.append(
            {
                "SR #": sr_num,
                "SR Status": nested_get(sr, "status", "statusName"),
                "Sub-Status": nested_get(sr, "subStatus", "subStatusName"),
                "Customer #": customer_number_from_sr(sr),
                "Customer Name": customer_name_from_sr(sr),
                "Site City": site_field(sr, "city"),
                "Site State": site_field(sr, "state"),
                "Line of Service": nested_get(sr, "lineOfService", "lineOfServiceName"),
                "Short Description": nested_get(sr, "shortDescription", "description"),
                "Created Date": nested_get(sr, "localCreatedDate", "createdDate", "srCreatedDate"),
                "WO #": nested_get(sr, "workOrderNumber"),
                "Invoice # (VIID)": latest.get("viid") if latest else "",
                "Invoice Status": latest.get("status") if latest else "No Invoice",
                "Consolidated Status": latest.get("consolidatedStatus") if latest else "",
                "SP Invoice #": latest.get("serviceProviderInvoiceNumber") if latest else "",
                "SP Amount": latest.get("serviceProviderTotalAmount") if latest else "",
                "Customer Amount": latest.get("customerTotalAmount") if latest else "",
                "Invoice Created": latest.get("createdDate") if latest else "",
                "Invoice Last Updated": latest.get("lastUpdatedDate") if latest else "",
                "Invoice Count": len(invs),
                "SP Name": SP_NAME,
                "SP #": SP_NUMBER,
            }
        )
    return report_rows


def write_workbook(report_rows: list[dict], all_invoices: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SR Invoice Report"

    ws["A1"] = f"{SP_NAME} ({SP_NUMBER}) — SR & Invoice Status Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A2"] = f"Generated: {date.today().isoformat()}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    header_row = 4
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, row in enumerate(report_rows, start=header_row + 1):
        for col, header in enumerate(HEADERS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=idx, column=col, value=value)
            if header == "Invoice Status":
                fill = status_fill(str(value))
                if fill:
                    cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A5"
    if report_rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADERS))}{header_row + len(report_rows)}"

    widths = [14, 10, 22, 10, 28, 12, 8, 14, 24, 22, 14, 14, 22, 18, 14, 10, 12, 22, 22, 12, 34, 10]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary.append(["Service Provider", f"{SP_NAME} ({SP_NUMBER})"])
    summary.append(["Total SRs", len(report_rows)])
    summary.append(["Total invoices (Gateway)", len(all_invoices)])
    summary.append(["SRs with invoices", sum(1 for r in report_rows if r.get("Invoice Count", 0) > 0)])
    summary.append(["SRs without invoices", sum(1 for r in report_rows if r.get("Invoice Count", 0) == 0)])
    summary.append([])

    sr_status = Counter(str(r.get("SR Status") or "") for r in report_rows)
    summary.append(["SR Status", "Count"])
    for status, count in sorted(sr_status.items()):
        summary.append([status, count])
    summary.append([])

    inv_status = Counter(str(r.get("Invoice Status") or "") for r in report_rows)
    summary.append(["Current Invoice Status (latest per SR)", "Count"])
    for status, count in sorted(inv_status.items()):
        summary.append([status, count])

    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 16

    inv_sheet = wb.create_sheet("All Invoices")
    inv_headers = [
        "VIID", "SR #", "Status", "Consolidated Status", "Customer", "LOS",
        "Short Description", "SP Amount", "Customer Amount", "SP Invoice #",
        "Created", "Last Updated",
    ]
    for col, header in enumerate(inv_headers, start=1):
        cell = inv_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for row_idx, inv in enumerate(
        sorted(all_invoices, key=lambda i: parse_dt(i.get("lastUpdatedDate")) or datetime.min, reverse=True),
        start=2,
    ):
        inv_sheet.cell(row=row_idx, column=1, value=inv.get("viid"))
        inv_sheet.cell(row=row_idx, column=2, value=inv.get("serviceRequestNumber"))
        inv_sheet.cell(row=row_idx, column=3, value=inv.get("status"))
        inv_sheet.cell(row=row_idx, column=4, value=inv.get("consolidatedStatus"))
        inv_sheet.cell(row=row_idx, column=5, value=inv.get("customerName"))
        inv_sheet.cell(row=row_idx, column=6, value=inv.get("lineOfService"))
        inv_sheet.cell(row=row_idx, column=7, value=inv.get("shortDescription"))
        inv_sheet.cell(row=row_idx, column=8, value=inv.get("serviceProviderTotalAmount"))
        inv_sheet.cell(row=row_idx, column=9, value=inv.get("customerTotalAmount"))
        inv_sheet.cell(row=row_idx, column=10, value=inv.get("serviceProviderInvoiceNumber"))
        inv_sheet.cell(row=row_idx, column=11, value=inv.get("createdDate"))
        inv_sheet.cell(row=row_idx, column=12, value=inv.get("lastUpdatedDate"))
    for col, width in enumerate([12, 14, 22, 18, 28, 12, 24, 10, 12, 14, 22, 22], start=1):
        inv_sheet.column_dimensions[get_column_letter(col)].width = width

    wb.save(OUT)


def main() -> int:
    srs = fetch_srs()
    invoices = fetch_invoices()
    report_rows = build_rows(srs, invoices)
    write_workbook(report_rows, invoices)
    print(json.dumps({
        "output": str(OUT),
        "sr_count": len(report_rows),
        "invoice_count": len(invoices),
        "columns": HEADERS,
        "sr_numbers": [r["SR #"] for r in report_rows],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
