#!/usr/bin/env python3
"""Build KS101108 Excel from cached VixxoLink/Gateway snapshots."""

from __future__ import annotations

import json
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
SRS_PATH = ROOT / "ks101108_srs.json"
INVOICES_PATH = ROOT / "ks101108_invoices.json"
OUT = ROOT / "KS101108_SR_Invoice_Report.xlsx"

SP_NUMBER = "KS101108"
SP_NAME = "KS - Montgomery Lock & Key Inc."

# siteId from vixxolink_get_service_request_details
SITE_BY_SR = {
    "1-6566190622": ("9555", "T-Mobile - Montgomery/AL - 9555"),
    "1-6566041692": ("9555", "T-Mobile - Montgomery/AL - 9555"),
    "1-6537476132": ("9555", "T-Mobile - Montgomery/AL - 9555"),
    "1-6527240392": ("9555", "T-Mobile - Montgomery/AL - 9555"),
    "1-6539929401": ("TAYLOR", "Taylor Crosg Animal Hosp - Montgomery/AL - TAYLOR"),
    "1-6534589222": ("TAYLOR", "Taylor Crosg Animal Hosp - Montgomery/AL - TAYLOR"),
    "1-6515883974": ("TAYLOR", "Taylor Crosg Animal Hosp - Montgomery/AL - TAYLOR"),
}

HEADERS = [
    "SR Number",
    "Customer #",
    "Site #",
    "Site Name",
    "Service Contractor Name",
    "Service Contractor",
    "SC Status",
    "Status",
    "Sub-Status",
    "Date Opened",
    "Short Description",
    "WO #",
    "Line of Service",
    "Gateway Invoice # (VIID)",
    "Gateway Invoice Status",
    "Gateway Consolidated Status",
    "SP Invoice #",
    "SP Amount",
    "Customer Amount",
]

SUB_STATUS_FILLS = {
    "Invoice Review": "FFEB9C",
    "Delinquent SC Invoice": "FFC7CE",
    "Invoice Required for SR": "FFE699",
    "Cancelled": "D9D9D9",
}


def parse_dt(raw) -> datetime | None:
    if not raw:
        return None
    try:
        return datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
    except ValueError:
        return None


def format_opened(raw) -> str:
    dt = parse_dt(raw)
    if not dt:
        return str(raw or "")
    if sys.platform == "win32":
        return dt.strftime("%#m/%#d/%Y %I:%M %p")
    return dt.strftime("%-m/%-d/%Y %I:%M %p")


def pick_latest(invoices: list[dict]) -> dict | None:
    if not invoices:
        return None
    return max(invoices, key=lambda i: parse_dt(i.get("lastUpdatedDate")) or datetime.min)


def main() -> None:
    srs = json.loads(SRS_PATH.read_text(encoding="utf-8"))
    invoices = json.loads(INVOICES_PATH.read_text(encoding="utf-8"))

    by_sr: dict[str, list[dict]] = {}
    for inv in invoices:
        by_sr.setdefault(str(inv.get("serviceRequestNumber")), []).append(inv)

    report_rows = []
    for sr in srs:
        sr_num = sr["number"]
        site_num, site_name = SITE_BY_SR.get(sr_num, ("", ""))
        invs = by_sr.get(sr_num, [])
        latest = pick_latest(invs)
        report_rows.append(
            {
                "SR Number": sr_num,
                "Customer #": sr.get("customer.number"),
                "Site #": site_num,
                "Site Name": site_name,
                "Service Contractor Name": sr.get("serviceProvider.name") or SP_NAME,
                "Service Contractor": sr.get("serviceProviderNumber") or SP_NUMBER,
                "SC Status": "Assigned",
                "Status": sr.get("status"),
                "Sub-Status": sr.get("subStatus"),
                "Date Opened": format_opened(sr.get("localCreatedDate")),
                "Short Description": sr.get("shortDescription"),
                "WO #": sr.get("workOrderNumber"),
                "Line of Service": sr.get("lineOfService"),
                "Gateway Invoice # (VIID)": latest.get("viid") if latest else "",
                "Gateway Invoice Status": latest.get("status") if latest else "",
                "Gateway Consolidated Status": latest.get("consolidatedStatus") if latest else "",
                "SP Invoice #": latest.get("serviceProviderInvoiceNumber") if latest else "",
                "SP Amount": latest.get("serviceProviderTotalAmount") if latest else "",
                "Customer Amount": latest.get("customerTotalAmount") if latest else "",
                "_sort": parse_dt(sr.get("localCreatedDate")) or datetime.min,
            }
        )

    report_rows.sort(key=lambda r: r["_sort"], reverse=True)
    for row in report_rows:
        row.pop("_sort", None)

    wb = Workbook()
    ws = wb.active
    ws.title = "SR Invoice Report"
    ws["A1"] = f"{SP_NAME} ({SP_NUMBER}) — Assigned SRs & Invoice Status"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A2"] = (
        f"Generated: {date.today().isoformat()} | "
        "Sub-Status = current SR invoice status (matches VixxoLink SP view)"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    header_row = 4
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, row in enumerate(report_rows, start=header_row + 1):
        for col, header in enumerate(HEADERS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=idx, column=col, value=value)
            if header == "Sub-Status":
                color = SUB_STATUS_FILLS.get(str(value))
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(HEADERS))}{header_row + len(report_rows)}"
    )
    widths = [14, 10, 8, 34, 34, 14, 10, 8, 24, 20, 22, 12, 14, 16, 22, 18, 12, 10, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary.append(["Service Provider", f"{SP_NAME} ({SP_NUMBER})"])
    summary.append(["Total SRs", len(report_rows)])
    summary.append(["Total Gateway invoices", len(invoices)])
    summary.append([])
    summary.append(["Sub-Status (invoice status)", "Count"])
    for status, count in sorted(Counter(r["Sub-Status"] for r in report_rows).items()):
        summary.append([status, count])

    inv_sheet = wb.create_sheet("All Gateway Invoices")
    inv_headers = [
        "VIID", "SR #", "Status", "Consolidated Status", "Customer", "Site #",
        "Site Name", "LOS", "Short Description", "SP Amount", "Customer Amount",
        "SP Invoice #", "Created", "Last Updated",
    ]
    for col, h in enumerate(inv_headers, start=1):
        c = inv_sheet.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
    for row_idx, inv in enumerate(
        sorted(invoices, key=lambda i: parse_dt(i.get("lastUpdatedDate")) or datetime.min, reverse=True),
        start=2,
    ):
        site_num, site_name = SITE_BY_SR.get(str(inv.get("serviceRequestNumber")), ("", ""))
        vals = [
            inv.get("viid"),
            inv.get("serviceRequestNumber"),
            inv.get("status"),
            inv.get("consolidatedStatus"),
            inv.get("customerName"),
            site_num,
            site_name,
            inv.get("lineOfService"),
            inv.get("shortDescription"),
            inv.get("serviceProviderTotalAmount"),
            inv.get("customerTotalAmount"),
            inv.get("serviceProviderInvoiceNumber"),
            inv.get("createdDate"),
            inv.get("lastUpdatedDate"),
        ]
        for col, val in enumerate(vals, start=1):
            inv_sheet.cell(row=row_idx, column=col, value=val)

    wb.save(OUT)
    print(json.dumps({"output": str(OUT), "sr_count": len(report_rows), "columns": HEADERS}, indent=2))


if __name__ == "__main__":
    main()
