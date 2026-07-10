from openpyxl import load_workbook
from collections import Counter

path = r"C:\Users\CGagner\Downloads\KS101344 - KS - Quality Maintenance Solutions LLC - SP Coverage.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb["Customer SC Relationships"]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
data = [dict(zip(headers, r)) for r in rows[1:] if any(r)]

print("Total rows:", len(data))
print("\nBy LOS:")
for los, cnt in Counter(d["LOS"] for d in data).most_common():
    print(f"  {los}: {cnt}")
print("\nBy State:")
for st, cnt in Counter(d["Site State"] for d in data).most_common():
    print(f"  {st}: {cnt}")
print("\nBy Customer:")
for c, cnt in Counter(d["Customer Number"] for d in data).most_common():
    print(f"  {c}: {cnt}")

ELECTRICAL_LOS = {
    "Electrical",
    "Electrical - Interior",
    "Electrical - Exterior",
    "Lighting",
    "Lighting - Exterior",
    "Lighting - Interior",
}

print("\n=== FULL COVERAGE LIST ===")
for d in data:
    los = d["LOS"]
    trade = "ELECTRICAL" if los in ELECTRICAL_LOS or "Electrical" in str(los) or "Lighting" in str(los) else "OTHER"
    print(
        f"{trade}\t{d['Customer Number']}\t{d['Site City']}, {d['Site State']}\t{los}\t{d['Short Description']}\t{d['Site Number']}\t{d['Site ID']}"
    )

wb.close()
