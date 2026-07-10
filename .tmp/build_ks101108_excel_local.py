"""Build KS101108 Excel report from cached Gateway/VixxoLink JSON snapshots."""
from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
SRS_PATH = ROOT / "ks101108_srs.json"
INVOICES_PATH = ROOT / "ks101108_invoices.json"
OUT = ROOT / "KS101108_SR_Invoice_Report.xlsx"

SP_NUMBER = "KS101108"
SP_NAME = "KS - Montgomery Lock & Key Inc."

HEADERS = [
    "SR #", "SR Status", "Sub-Status", "Customer #", "Customer Name",
    "Site City", "Site State", "Line of Service", "Short Description",
    "Created Date", "WO #", "Invoice # (VIID)", "Invoice Status",
    "Consolidated Status", "SP Invoice #", "SP Amount", "Customer Amount",
    "Invoice Created", "Invoice Last Updated", "Invoice Count",
    "SP Name", "SP #",
]

STATUS_FILLS = {
    "Rejected": PatternFill("solid", fgColor="FFC7CE"),
    "Review": PatternFill("solid", fgColor="FFEB9C"),
    "In Progress": PatternFill("solid", fgColor="DDEBF7"),
}


def parse_dt(raw) -> datetime | None:
    if not raw:
        return None
    try:
        return datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
    except ValueError:
        return None


def pick_latest(invoices: list[dict]) -> dict | None:
    if not invoices:
        return None
    return max(invoices, key=lambda i: parse_dt(i.get("lastUpdatedDate")) or datetime.min)


def status_fill(status: str) -> PatternFill | None:
    for key, fill in STATUS_FILLS.items():
        if key.lower() in (status or "").lower():
            return fill
    return None


def main() -> None:
    srs = json.loads(SRS_PATH.read_text(encoding="utf-8"))
    invoices = json.loads(INVOICES_PATH.read_text(encoding="utf-8"))

    by_sr: dict[str, list[dict]] = defaultdict(list)
    for inv in invoices:
        by_sr[str(inv.get("serviceRequestNumber"))].append(inv)

    report_rows = []
    for sr in sorted(srs, key=lambda r: r.get("localCreatedDate", ""), reverse=True):
        sr_num = sr["number"]
        invs = by_sr.get(sr_num, [])
        latest = pick_latest(invs)
        report_rows.append({
            "SR #": sr_num,
            "SR Status": sr.get("status"),
            "Sub-Status": sr.get("subStatus"),
            "Customer #": sr.get("customer.number"),
            "Customer Name": sr.get("customerName"),
            "Site City": sr.get("site.city"),
            "Site State": sr.get("site.state"),
            "Line of Service": sr.get("lineOfService"),
            "Short Description": sr.get("shortDescription"),
            "Created Date": sr.get("localCreatedDate"),
            "WO #": sr.get("workOrderNumber"),
            "Invoice # (VIID)": latest.get("viid") if latest else "",
            "Invoice Status": latest.get("status") if latest else "No Invoice",
            "Consolidated Status": latest.get("consolidatedStatus") if latest else "",
            "SP Invoice #": latest.get("serviceProviderInvoiceNumber") if latest else "",
            "SP Amount": latest.get("serviceProviderTotalAmount") if latest else "",
            "Customer Amount": latest.get("customerTotalAmount") if latest else "",
            "Invoice Created": latest.get("createdDate") if latest else "",
            "Invoice Last Updated": latest.get("lastUpdatedDate") if latest else "",
            "Invoice Count": len(invs),
            "SP Name": SP_NAME,
            "SP #": SP_NUMBER,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "SR Invoice Report"
    ws["A1"] = f"{SP_NAME} ({SP_NUMBER}) — SR & Invoice Status Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A2"] = f"Generated: {date.today().isoformat()}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    header_row = 4
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, row in enumerate(report_rows, start=header_row + 1):
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=idx, column=col, value=row.get(header, ""))
            if header == "Invoice Status":
                fill = status_fill(str(row.get(header, "")))
                if fill:
                    cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADERS))}{header_row + len(report_rows)}"
    for col, width in enumerate([14, 10, 22, 10, 28, 12, 8, 14, 24, 22, 14, 14, 22, 18, 14, 10, 12, 22, 22, 12, 34, 10], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary.append(["Service Provider", f"{SP_NAME} ({SP_NUMBER})"])
    summary.append(["Total SRs", len(report_rows)])
    summary.append(["Total invoices (Gateway)", len(invoices)])
    summary.append(["SRs with invoices", sum(1 for r in report_rows if r["Invoice Count"] > 0)])
    summary.append(["SRs without invoices", sum(1 for r in report_rows if r["Invoice Count"] == 0)])
    summary.append([])
    summary.append(["SR Status", "Count"])
    for status, count in sorted(Counter(r["SR Status"] for r in report_rows).items()):
        summary.append([status, count])
    summary.append([])
    summary.append(["Current Invoice Status (latest per SR)", "Count"])
    for status, count in sorted(Counter(r["Invoice Status"] for r in report_rows).items()):
        summary.append([status, count])
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 16

    inv_sheet = wb.create_sheet("All Invoices")
    inv_headers = ["VIID", "SR #", "Status", "Consolidated Status", "Customer", "LOS", "Short Description", "SP Amount", "Customer Amount", "SP Invoice #", "Created", "Last Updated"]
    for col, h in enumerate(inv_headers, start=1):
        c = inv_sheet.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
    for row_idx, inv in enumerate(sorted(invoices, key=lambda i: parse_dt(i.get("lastUpdatedDate")) or datetime.min, reverse=True), start=2):
        vals = [inv.get("viid"), inv.get("serviceRequestNumber"), inv.get("status"), inv.get("consolidatedStatus"), inv.get("customerName"), inv.get("lineOfService"), inv.get("shortDescription"), inv.get("serviceProviderTotalAmount"), inv.get("customerTotalAmount"), inv.get("serviceProviderInvoiceNumber"), inv.get("createdDate"), inv.get("lastUpdatedDate")]
        for col, val in enumerate(vals, start=1):
            inv_sheet.cell(row=row_idx, column=col, value=val)
    for col, width in enumerate([12, 14, 22, 18, 28, 12, 24, 10, 12, 14, 22, 22], start=1):
        inv_sheet.column_dimensions[get_column_letter(col)].width = width

    wb.save(OUT)
    print(json.dumps({"output": str(OUT), "sr_count": len(report_rows), "invoice_count": len(invoices), "columns": HEADERS}, indent=2))


if __name__ == "__main__":
    main()
