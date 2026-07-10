"""Build KS101108 invoice audit + notes export for Downloads folder."""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
VETTING_SCRIPTS = SCRIPT_DIR.parent / ".agents" / "skills" / "sp-inbound-vetting" / "scripts"
sys.path.insert(0, str(VETTING_SCRIPTS))

from mcp_http import mcp_call, mcp_result_text  # noqa: E402

GATEWAY_URL = "https://vixxonow.com/mcp/gateway"
SP_NUMBER = "KS101108"
SP_NAME = "KS - Montgomery Lock & Key Inc."
TARGET_SRS = [
    "1-6566190622",
    "1-6566041692",
    "1-6539929401",
    "1-6537476132",
    "1-6534589222",
    "1-6527240392",
    "1-6515883974",
]
OUT = Path.home() / "Downloads" / f"KS101108_SR_Invoice_Audit_Export_{date.today().isoformat()}.xlsx"

KEY_NOTE_TYPES = {
    "Rejected Note",
    "Tracking Hold",
    "Visual Audit",
    "Review By AP",
    "RESOLUTION",
    "Invoice",
}


def parse_json_blob(text: str):
    text = (text or "").strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}|\[.*\]", text, re.S)
        if m:
            try:
                return json.loads(m.group(0))
            except json.JSONDecodeError:
                return None
    return None


def unwrap_data(payload) -> dict:
    if isinstance(payload, dict) and isinstance(payload.get("data"), dict):
        return payload["data"]
    return payload if isinstance(payload, dict) else {}


def gateway_call(tool: str, args: dict) -> dict:
    resp = mcp_call(GATEWAY_URL, tool, args)
    text = mcp_result_text(resp)
    if text.startswith("HTTP ") or text.startswith("MCP stdio"):
        raise RuntimeError(f"{tool} failed: {text[:400]}")
    blob = parse_json_blob(text)
    if isinstance(blob, dict) and blob.get("data"):
        return blob["data"]
    return blob if isinstance(blob, dict) else {}


def invoice_rows(data: dict) -> list[dict]:
    if isinstance(data.get("invoiceList"), list):
        return data["invoiceList"]
    nested = data.get("data")
    if isinstance(nested, dict) and isinstance(nested.get("invoiceList"), list):
        return nested["invoiceList"]
    return []


def line_item_rows(data: dict) -> list[dict]:
    return data.get("lineItemList") or []


def note_rows(data: dict) -> list[dict]:
    return data.get("noteList") or []


def extract_rejection_reason(notes: list[dict], status: str) -> str:
    reasons: list[str] = []
    for note in notes:
        ntype = str(note.get("type") or "")
        text = str(note.get("text") or "").strip()
        if ntype == "Rejected Note":
            m = re.search(r"Rejection Reason Code:\s*(.+?)(?:\n\n|$)", text, re.S)
            if m:
                reasons.append(m.group(1).strip())
            else:
                reasons.append(text)
        elif ntype == "Tracking Hold" and text:
            reasons.append(text)
        elif "Reject" in text and "Invoice status changed" in text and "to 'Rejected" in text:
            m = re.search(r"selected action '([^']+)'", text)
            if m:
                reasons.append(f"Action: {m.group(1)}")
    if not reasons and str(status).startswith("Rejected"):
        for note in reversed(notes):
            if note.get("type") == "Visual Audit":
                m = re.search(r"Note:\s*\r?\n(.+)", str(note.get("text") or ""), re.S)
                if m and m.group(1).strip():
                    reasons.append(m.group(1).strip())
                    break
    deduped: list[str] = []
    for r in reasons:
        if r and r not in deduped:
            deduped.append(r)
    return " | ".join(deduped)


def extract_audit_notes(notes: list[dict]) -> str:
    parts: list[str] = []
    for note in notes:
        ntype = str(note.get("type") or "")
        text = str(note.get("text") or "").strip()
        if ntype in KEY_NOTE_TYPES and ntype not in {"Invoice", "RESOLUTION"}:
            if ntype == "Visual Audit" or ntype == "Review By AP":
                m = re.search(r"Selected Action:\s*([^\r\n]+)(?:\r?\n\r?\nNote:\s*\r?\n(.+))?", text, re.S)
                if m:
                    action = m.group(1).strip()
                    detail = (m.group(2) or "").strip()
                    entry = f"{action}: {detail}" if detail else action
                    if entry and entry not in parts:
                        parts.append(entry)
            elif text and text not in parts:
                parts.append(f"{ntype}: {text[:500]}")
    return " | ".join(parts[:6])


def failed_audit_rows(data: dict) -> list[dict]:
    return data.get("failedAudits") or []


def pricing(line: dict, side: str) -> dict:
    key = "serviceProviderPricingDetail" if side == "sp" else "customerPricingDetail"
    return line.get(key) or {}


def write_sheet_headers(ws, headers: list[str], row: int = 1) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths: list[int]) -> None:
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def main() -> int:
    inv_data = gateway_call("gateway_search_invoices", {"serviceProviderNumber": SP_NUMBER})
    all_invoices = invoice_rows(inv_data)
    target_invoices = [
        inv for inv in all_invoices
        if str(inv.get("serviceRequestNumber")) in TARGET_SRS
    ]
    target_invoices.sort(key=lambda i: (str(i.get("serviceRequestNumber")), str(i.get("createdDate"))))

    invoice_summary_rows: list[dict] = []
    line_rows: list[dict] = []
    audit_fail_rows: list[dict] = []
    note_rows_out: list[dict] = []

    for inv in target_invoices:
        viid = inv.get("viid")
        sr = str(inv.get("serviceRequestNumber") or "")
        li_data = gateway_call("gateway_get_invoice_lineitems", {"viid": viid})
        notes_data = gateway_call("gateway_list_invoice_notes", {"viid": viid})
        audit_data = gateway_call("gateway_run_invoice_audit", {"viid": viid})
        notes = note_rows(notes_data)
        rejection = extract_rejection_reason(notes, str(inv.get("status") or ""))
        audit_notes = extract_audit_notes(notes)
        failed = failed_audit_rows(audit_data)
        failed_summary = " | ".join(
            f"Rule {a.get('ruleId')} ({a.get('ruleDescription')}): {a.get('message') or a.get('result')}"
            for a in failed
        )

        invoice_summary_rows.append({
            "SR #": sr,
            "VIID": viid,
            "Status": inv.get("status"),
            "Consolidated Status": inv.get("consolidatedStatus"),
            "SP Invoice #": inv.get("serviceProviderInvoiceNumber"),
            "SP Amount": inv.get("serviceProviderTotalAmount"),
            "Customer Amount": inv.get("customerTotalAmount"),
            "Customer": inv.get("customerName"),
            "LOS": inv.get("lineOfService"),
            "Created": inv.get("createdDate"),
            "Last Updated": inv.get("lastUpdatedDate"),
            "Rejection Reason(s)": rejection,
            "Audit / Review Notes": audit_notes,
            "Failed Audit Rules": failed_summary,
        })

        for line in line_item_rows(li_data):
            sp = pricing(line, "sp")
            cust = pricing(line, "cust")
            if str(line.get("typeName")) in {"Processing Fee"} and sp.get("total") in (0, 0.0, None) and cust.get("total") in (0, 0.0, None):
                continue
            line_rows.append({
                "SR #": sr,
                "VIID": viid,
                "Line Type": line.get("typeName"),
                "Description": line.get("description"),
                "Qty (SP)": sp.get("quantity"),
                "SP Unit": sp.get("unitPrice"),
                "SP Total": sp.get("total"),
                "Qty (Cust)": cust.get("quantity"),
                "Cust Unit": cust.get("unitPrice"),
                "Cust Total": cust.get("total"),
            })

        for audit in failed:
            audit_fail_rows.append({
                "SR #": sr,
                "VIID": viid,
                "Rule ID": audit.get("ruleId"),
                "Rule": audit.get("ruleDescription"),
                "Result": audit.get("result"),
                "Rule Type": audit.get("ruleType"),
                "Message": audit.get("message"),
            })

        for note in notes:
            note_rows_out.append({
                "SR #": sr,
                "VIID": viid,
                "Note ID": note.get("noteId"),
                "Type": note.get("type"),
                "Created": note.get("createdDate"),
                "Created By": note.get("createdUserId"),
                "SP Visible": note.get("visibleToServiceProvider"),
                "Text": str(note.get("text") or "").replace("\r\n", "\n"),
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Summary"
    ws["A1"] = f"{SP_NAME} ({SP_NUMBER}) — Invoice Audit Export"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | SRs: {', '.join(TARGET_SRS)}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)

    inv_headers = [
        "SR #", "VIID", "Status", "Consolidated Status", "SP Invoice #", "SP Amount",
        "Customer Amount", "Customer", "LOS", "Created", "Last Updated",
        "Rejection Reason(s)", "Audit / Review Notes", "Failed Audit Rules",
    ]
    write_sheet_headers(ws, inv_headers, row=4)
    for ridx, row in enumerate(invoice_summary_rows, start=5):
        for cidx, key in enumerate(inv_headers, start=1):
            cell = ws.cell(row=ridx, column=cidx, value=row.get(key, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A5"
    autosize(ws, [14, 12, 22, 16, 12, 10, 12, 24, 12, 20, 20, 40, 40, 50])

    sr_sheet = wb.create_sheet("SR Overview")
    sr_headers = ["SR #", "Invoice Count", "Latest VIID", "Latest Status", "Rejection / Hold Summary"]
    write_sheet_headers(sr_sheet, sr_headers)
    for sr in TARGET_SRS:
        invs = [i for i in invoice_summary_rows if i["SR #"] == sr]
        if not invs:
            sr_sheet.append([sr, 0, "", "No Invoice", "Invoice Required — no invoice on file"])
            continue
        latest = max(invs, key=lambda i: str(i.get("Last Updated") or ""))
        rejections = " | ".join(
            r for r in dict.fromkeys(i.get("Rejection Reason(s)") or "" for i in invs if i.get("Rejection Reason(s)"))
        )
        sr_sheet.append([sr, len(invs), latest.get("VIID"), latest.get("Status"), rejections or latest.get("Audit / Review Notes")])
    autosize(sr_sheet, [14, 12, 12, 22, 60])

    li_sheet = wb.create_sheet("Line Items")
    li_headers = ["SR #", "VIID", "Line Type", "Description", "Qty (SP)", "SP Unit", "SP Total", "Qty (Cust)", "Cust Unit", "Cust Total"]
    write_sheet_headers(li_sheet, li_headers)
    for row in line_rows:
        li_sheet.append([row.get(h) for h in li_headers])
    autosize(li_sheet, [14, 12, 14, 28, 8, 10, 10, 8, 10, 10])

    audit_sheet = wb.create_sheet("Failed Audits")
    audit_headers = ["SR #", "VIID", "Rule ID", "Rule", "Result", "Rule Type", "Message"]
    write_sheet_headers(audit_sheet, audit_headers)
    for row in audit_fail_rows:
        audit_sheet.append([row.get(h) for h in audit_headers])
    autosize(audit_sheet, [14, 12, 8, 28, 10, 10, 60])

    notes_sheet = wb.create_sheet("All Notes")
    notes_headers = ["SR #", "VIID", "Note ID", "Type", "Created", "Created By", "SP Visible", "Text"]
    write_sheet_headers(notes_sheet, notes_headers)
    for row in note_rows_out:
        notes_sheet.append([row.get(h) for h in notes_headers])
        notes_sheet.cell(row=notes_sheet.max_row, column=8).alignment = Alignment(wrap_text=True, vertical="top")
    autosize(notes_sheet, [14, 12, 10, 18, 20, 16, 10, 80])

    wb.save(OUT)
    print(json.dumps({
        "output": str(OUT),
        "invoice_count": len(invoice_summary_rows),
        "line_item_count": len(line_rows),
        "failed_audit_count": len(audit_fail_rows),
        "note_count": len(note_rows_out),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
