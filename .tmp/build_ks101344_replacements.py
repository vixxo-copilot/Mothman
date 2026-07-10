from openpyxl import load_workbook, Workbook

SRC = r"C:\Users\CGagner\Downloads\KS101344 - KS - Quality Maintenance Solutions LLC - SP Coverage.xlsx"
OUT = r"C:\Users\CGagner\source\mothman\.tmp\KS101344_coverage_replacement_proposal_v3.xlsx"

WSD = "KS69211 : WSD Solutions LLC (Active)"
MECKLEY = "F5013521 : Meckley Services, Inc. (Active)"
OPEN = "TBD — local SP search required"

ELECTRICAL_LOS = {"Electrical", "Electrical - Interior", "Electrical - Exterior", "Lighting", "Lighting - Exterior", "Lighting - Interior"}

LOCKSMITHS = {
    "Harrisonburg, VA": {
        "site_context": "OSI01 O304724 — Locks (261 University Blvd); Harrisonburg CBSA",
        "primary": "KS101552 : KS - Ace Lock and Security (Active)",
        "primary_notes": "Already Rank 2 on Locks row; Earlysville; Active in VixxoLink; 434-293-9006",
        "backup": "KS69615 : KS - Pop-A-Lock of Northern Virginia (Active)",
        "backup_notes": "Fredericksburg; Active in VixxoLink; 571-384-0175; viable Harrisonburg CBSA option per ops",
        "alternate": "13147 : Norvac Lock Technology Inc (Winchester)",
        "alternate_notes": "Shenandoah Valley / I-81 corridor; 540-662-5641; confirm Siebel active status",
    },
    "Charlottesville, VA": {
        "site_context": "OSI01 O304717 / OMF01 46160501 / CMX01 07290 — no Locks row today; local lock SP if coverage added",
        "primary": "KS101552 : KS - Ace Lock and Security (Active)",
        "primary_notes": "Same SP as Harrisonburg; Earlysville is in Charlottesville CBSA; best active incumbent in market",
        "backup": "12817 : Albermarie Safe & Lock Co. Inc. (Charlottesville)",
        "backup_notes": "Local Charlottesville shop; 540-946-0012; inactive in VixxoLink — reactivation review if needed",
        "alternate": "12378 : Action Lock, Safe & Security (Charlottesville)",
        "alternate_notes": "Local Charlottesville; inactive in VixxoLink",
    },
}


def replacement_for(row):
    los = row["LOS"]
    state = row["Site State"]

    if los in ELECTRICAL_LOS or "Electrical" in str(los):
        return MECKLEY, "Meckley — incumbent; licensed for Electrical"

    if los == "Handyman" and state == "VA":
        return WSD, "WSD — licensed Handyman; not licensed Electrical"

    if los == "Plumbing" and state == "VA":
        return WSD, "WSD — strong VA plumbing history; confirm license scope"

    if los == "Doors" and state == "VA":
        return WSD, "WSD — VA doors work on record; confirm license scope"

    if los == "Locks" and state == "VA":
        rec = LOCKSMITHS["Harrisonburg, VA"]
        return rec["primary"], f"Promote Ace Lock to Rank 1; backup Pop-A-Lock KS69615 or Norvac 13147"

    if los == "Heating Ventilation AC":
        return OPEN, f"HVAC — site is {state}; outside VA incumbent cluster"

    if los == "Landscaping":
        return OPEN, f"Landscaping — site is {state}; needs local landscaping SP"

    if los == "Redirect":
        return OPEN, "Redirect / fee-management row — confirm with ops before replacing"

    if state == "VA":
        return OPEN, f"Review manually — {los} in VA"

    return OPEN, f"Review manually — {los} in {state}"


wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Customer SC Relationships"]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
data = [dict(zip(headers, r)) for r in rows[1:] if any(r)]
wb.close()

out_wb = Workbook()
summary = out_wb.active
summary.title = "Summary"
summary.append(["Replacement logic"])
summary.append(["Electrical / Lighting", MECKLEY])
summary.append(["Handyman (VA)", WSD])
summary.append(["Plumbing / Doors (VA)", WSD, "confirm license scope"])
summary.append(["Locks — Harrisonburg", LOCKSMITHS["Harrisonburg, VA"]["primary"], "Backup: KS69615 Pop-A-Lock"])
summary.append(["Lock SP — Charlottesville metro", LOCKSMITHS["Charlottesville, VA"]["primary"]])
summary.append(["Non-VA HVAC / Landscaping / Redirect", "TBD local SP"])
summary.append([])
summary.append(["Counts"])

assign = out_wb.create_sheet("Proposed Replacements")
assign_headers = list(headers) + ["Proposed Rank 1", "Assignment Rationale", "Action"]
assign.append(assign_headers)

counts = {"Meckley": 0, "WSD": 0, "Locksmith": 0, "TBD": 0}
for row in data:
    proposed, rationale = replacement_for(row)
    if proposed == MECKLEY:
        action = "Replace Rank 1 with Meckley"
        counts["Meckley"] += 1
    elif proposed == WSD:
        action = "Replace Rank 1 with WSD"
        counts["WSD"] += 1
    elif "Ace Lock" in proposed:
        action = "Promote Ace Lock to Rank 1 on Locks row; remove QMS"
        counts["Locksmith"] += 1
    else:
        action = "Hold — identify alternate SP"
        counts["TBD"] += 1
    assign.append([row.get(h) for h in headers] + [proposed, rationale, action])

summary.append(["Total QMS rows", len(data)])
summary.append(["Assign to Meckley", counts["Meckley"]])
summary.append(["Assign to WSD", counts["WSD"]])
summary.append(["Assign to Ace Lock", counts["Locksmith"]])
summary.append(["Needs alternate SP", counts["TBD"]])

review = out_wb.create_sheet("Load Ready")
review.append(
    [
        "Customer Number",
        "Site Number",
        "Site ID",
        "Site City",
        "Site State",
        "LOS",
        "Short Description",
        "Current Rank 1",
        "Current Rank 2",
        "Proposed Rank 1",
        "Rationale",
    ]
)
for row in data:
    proposed, rationale = replacement_for(row)
    review.append(
        [
            row["Customer Number"],
            row["Site Number"],
            row["Site ID"],
            row["Site City"],
            row["Site State"],
            row["LOS"],
            row["Short Description"],
            row["Rank 1"],
            row["Rank 2"],
            proposed,
            rationale,
        ]
    )

lock_sheet = out_wb.create_sheet("Locksmith Suggestions")
lock_sheet.append(
    [
        "Location",
        "Site / Customer Context",
        "Recommended Rank 1",
        "Notes",
        "Backup Rank 2",
        "Backup Notes",
        "Alternate",
        "Alternate Notes",
    ]
)
for location, rec in LOCKSMITHS.items():
    lock_sheet.append(
        [
            location,
            rec["site_context"],
            rec["primary"],
            rec["primary_notes"],
            rec["backup"],
            rec["backup_notes"],
            rec["alternate"],
            rec["alternate_notes"],
        ]
    )

lock_sheet.append([])
lock_sheet.append(["Lynchburg note", "CMX01 06056 Doors row", "13088 : Hawkins Lock & Key Co. Inc. (Lynchburg)", "Local Lynchburg locksmith; 434-846-0793; inactive in VixxoLink — optional if door hardware needs lock vendor", "", "", "", ""])

out_wb.save(OUT)
print("Saved:", OUT)
