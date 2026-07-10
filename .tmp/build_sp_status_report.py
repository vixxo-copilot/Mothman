"""Build Excel SP status report for requested companies."""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "SP_VixxoLink_Gateway_Status_Report.xlsx"

ROWS = [
    ("Anderson Custom Window Coverings, Inc.", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Atlantic Window Coverings", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Bigelow Refrigeration", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("DesignTeam Inc", "KS102207", "Active in VixxoLink"),
    ("Just Rite Acoustics", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Ryan Fireprotection, Inc", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Service Pro Locksmith", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Ambius Nashville", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Crespo's Wildlife Solutions LLC", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Glass Doctor of Finger Lakes", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Jiffy Rooter, LLC", "60523", "Inactive in VixxoLink (Jiffy Rooter, Inc.)"),
    ("Luxeshine Cleaning NY 8 LLC", "", "SF Account exists; no SP # assigned"),
    ("Northmen Glass LLC", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Summit Fire & Security LLC", "68127", "Active in VixxoLink"),
    ("Summit Fire National Accounts", "", "Closed SF Lead only — not an SP record"),
    ("The Flying Locksmiths - Sarasota", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("The Village Locksmith", "KS101988", "Active in VixxoLink"),
    ("Epic Septic & Service LLC", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("TS Backflow", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Precision Locksmith & Door Maintenance", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("C&K Paving Contractors, Inc", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("G&H WorX", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Richards Board Up Service", "", "Not found — no SP # in Salesforce or VixxoLink"),
    ("Franks Appliance Repair", "KS101961", "Active in VixxoLink"),
    ("Johnson Equipment Company", "KS101962", "Active in VixxoLink (legacy SP 7086 also active)"),
]

HEADERS = ["Company Name", "SP #", "Status"]

STATUS_FILLS = {
    "Active": PatternFill("solid", fgColor="C6EFCE"),
    "Inactive": PatternFill("solid", fgColor="FFEB9C"),
    "Not found": PatternFill("solid", fgColor="FFC7CE"),
    "SF Account": PatternFill("solid", fgColor="FFE699"),
    "Lead only": PatternFill("solid", fgColor="DDEBF7"),
}


def status_fill(status: str) -> PatternFill | None:
    for key, fill in STATUS_FILLS.items():
        if key.lower() in status.lower():
            return fill
    return None


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SP Status"

    ws["A1"] = "Service Provider VixxoLink / Gateway Status Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Generated: {date.today().isoformat()}"
    ws.merge_cells("A2:C2")

    header_row = 4
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, (company, sp_number, status) in enumerate(ROWS, start=header_row + 1):
        ws.cell(row=idx, column=1, value=company)
        ws.cell(row=idx, column=2, value=sp_number)
        status_cell = ws.cell(row=idx, column=3, value=status)
        fill = status_fill(status)
        if fill:
            status_cell.fill = fill

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:C{header_row + len(ROWS)}"

    widths = [42, 14, 58]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + len(ROWS), min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = wb.create_sheet("Summary")
    summary.append(["Status category", "Count"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    counts = {
        "Active in VixxoLink": 0,
        "Inactive in VixxoLink": 0,
        "No SP # assigned / not found": 0,
        "Other (SF account or Lead only)": 0,
    }
    for _, sp, status in ROWS:
        if "Active in VixxoLink" in status and "legacy" not in status.lower():
            counts["Active in VixxoLink"] += 1
        elif "Inactive" in status:
            counts["Inactive in VixxoLink"] += 1
        elif "Not found" in status:
            counts["No SP # assigned / not found"] += 1
        else:
            counts["Other (SF account or Lead only)"] += 1
    for label, count in counts.items():
        summary.append([label, count])
    summary.append([])
    summary.append(["Total companies reviewed", len(ROWS)])
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 12

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
