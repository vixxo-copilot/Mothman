#!/usr/bin/env python3
"""Build or refresh exports/vixxo-sr-status-reference.xlsx.

Live refresh order for Gateway invoice statuses:
1. gateway_list_invoice_statuses (authoritative picklist; requires Gateway OAuth bearer)
2. gateway_search_invoices sampling (requires Gateway OAuth bearer)
3. Workspace-documented fallback values
"""

from __future__ import annotations

import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "exports" / "vixxo-sr-status-reference.xlsx"
SCRIPTS = ROOT / ".agents" / "skills" / "sp-inbound-vetting" / "scripts"
GATEWAY_URL = "https://vixxonow.com/mcp/gateway"
TODAY = date.today().isoformat()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
META_FILL = PatternFill("solid", fgColor="FFF2CC")
LIVE_FILL = PatternFill("solid", fgColor="E2EFDA")
DOC_FILL = PatternFill("solid", fgColor="FCE4D6")
PICKLIST_FILL = PatternFill("solid", fgColor="C6E0B4")


def style_header(ws, row: int) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 60) -> None:
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), max_width)


def _mcp_modules():
    sys.path.insert(0, str(SCRIPTS))
    from gateway_vetting import gateway_search_invoices, parse_json_blob
    from mcp_http import _load_env, _token, mcp_call, mcp_result_text

    return gateway_search_invoices, parse_json_blob, _load_env, _token, mcp_call, mcp_result_text


def _parse_status_blob(data: Any) -> list[str]:
    """Normalize gateway_list_invoice_statuses payloads to status strings."""
    if data is None:
        return []
    if isinstance(data, list):
        out: list[str] = []
        for item in data:
            if isinstance(item, str) and item.strip():
                out.append(item.strip())
            elif isinstance(item, dict):
                for key in ("status", "name", "value", "code", "invoiceStatus"):
                    val = item.get(key)
                    if val:
                        out.append(str(val).strip())
                        break
        return out
    if isinstance(data, dict):
        for key in (
            "statuses",
            "invoiceStatuses",
            "invoice_statuses",
            "items",
            "data",
            "results",
        ):
            if key in data:
                return _parse_status_blob(data[key])
        for key in ("status", "name", "value"):
            if data.get(key):
                return [str(data[key]).strip()]
    return []


def gateway_picklist_invoice_statuses() -> tuple[list[str], str]:
    """Fetch authoritative Gateway invoice.status picklist when OAuth bearer is present."""
    try:
        _, parse_json_blob, _load_env, _token, mcp_call, mcp_result_text = _mcp_modules()
    except ImportError:
        return [], "mcp modules unavailable"

    _load_env()
    if not _token():
        return [], "no Gateway bearer token"

    resp = mcp_call(GATEWAY_URL, "gateway_list_invoice_statuses", {})
    text = mcp_result_text(resp)
    if not text or "invalid_token" in text or text.startswith("HTTP 401"):
        return [], "gateway_list_invoice_statuses unauthorized (Gateway OAuth bearer required)"

    data = parse_json_blob(text)
    if data is None and text and not text.startswith("{"):
        # Occasionally returned as plain newline-separated values.
        statuses = [line.strip() for line in text.splitlines() if line.strip()]
        if statuses:
            return statuses, "gateway_list_invoice_statuses"
        return [], f"unparsed picklist response: {text[:120]}"

    statuses = _parse_status_blob(data)
    if statuses:
        return sorted({s for s in statuses if s}), "gateway_list_invoice_statuses"
    return [], f"empty picklist response: {text[:120]}"


def gateway_sampled_invoice_statuses() -> tuple[list[dict], str]:
    """Sample invoice.status values from gateway_search_invoices."""
    try:
        gateway_search_invoices, _, _load_env, _token, _, _ = _mcp_modules()
    except ImportError:
        return [], "mcp modules unavailable"

    _load_env()
    if not _token():
        return [], "no Gateway bearer token"

    probes = [
        {"serviceProviderNumber": "KS69315", "pageSize": 50},
        {"serviceProviderNumber": "KS101031", "pageSize": 50},
        {"serviceProviderNumber": "KS101347", "pageSize": 50},
        {"customerNumber": "811", "pageSize": 50},
        {"searchString": "KS69315", "pageSize": 50},
        {"status": "Paid", "pageSize": 25},
        {"status": "Billed", "pageSize": 25},
        {"status": "Accepted", "pageSize": 25},
    ]

    seen: dict[str, dict] = {}
    auth_failed = False
    for kwargs in probes:
        try:
            rows = gateway_search_invoices(**kwargs)
        except Exception:
            continue
        if not rows:
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            status = (
                row.get("status")
                or row.get("invoiceStatus")
                or row.get("invoice_status")
            )
            if not status:
                continue
            key = str(status).strip()
            if not key or key in seen:
                continue
            seen[key] = {
                "status": key,
                "sample_sr": row.get("serviceRequestNumber"),
                "sample_sp": row.get("serviceProviderNumber") or row.get("serviceProviderName"),
                "probe": json.dumps(kwargs, sort_keys=True),
            }

    if seen:
        return [seen[k] for k in sorted(seen)], "gateway_search_invoices"
    return [], "gateway_search_invoices returned no invoice rows"


def gateway_live_invoice_statuses() -> tuple[list[dict], dict[str, str]]:
    """Combine picklist + sampled live Gateway invoice statuses."""
    diagnostics: dict[str, str] = {}
    live_rows: list[dict] = []
    live_statuses: set[str] = set()

    picklist, pick_msg = gateway_picklist_invoice_statuses()
    diagnostics["picklist"] = pick_msg
    for status in picklist:
        live_statuses.add(status)
        live_rows.append(
            {
                "status": status,
                "category": "Gateway picklist",
                "origin": "Gateway MCP picklist",
                "source_ref": "gateway_list_invoice_statuses",
                "sample_sr": "",
                "sample_sp": "",
            }
        )

    samples, sample_msg = gateway_sampled_invoice_statuses()
    diagnostics["sample"] = sample_msg
    for row in samples:
        status = row["status"]
        if status in live_statuses:
            continue
        live_statuses.add(status)
        live_rows.append(
            {
                "status": status,
                "category": "Gateway live sample",
                "origin": "Gateway MCP sample",
                "source_ref": f"gateway_search_invoices ({row['probe']})",
                "sample_sr": row.get("sample_sr") or "",
                "sample_sp": row.get("sample_sp") or "",
            }
        )

    return live_rows, diagnostics


def documented_gateway_invoice_statuses() -> list[tuple]:
    return [
        ("Accepted", "Payment lifecycle", "Invoice accepted in billing workflow", "vixxo-freshdesk-invoice-review/SKILL.md"),
        ("Billed", "Payment lifecycle", "Invoice billed to customer/AP", "vixxo-freshdesk-invoice-review/SKILL.md"),
        ("Paid", "Payment lifecycle", "Invoice paid", "vixxo-freshdesk-invoice-review/SKILL.md"),
        ("UnderARInvestigation", "AP review / hold", "AR investigation hold", "vixxo-freshdesk-invoice-review/SKILL.md; vixxo-spm-invoice-concerns/SKILL.md"),
        ("ReviewByAP", "AP review / hold", "Accounts payable review required", "vixxo-freshdesk-invoice-review/SKILL.md; vixxo-spm-invoice-concerns/SKILL.md"),
        ("VisualAudit", "AP review / hold", "Visual audit review", "vixxo-freshdesk-invoice-review/SKILL.md; vixxo-spm-invoice-concerns/SKILL.md"),
        ("Invoice Review", "AP review / hold", "Invoice in AP line-item review (also appears as SR Substatus)", "vixxo-freshdesk-invoice-review/SKILL.md; sub-status-decoder.md"),
        ("PaidOnlyPending", "AP review / hold", "Paid-only pending state", "vixxo-freshdesk-invoice-review/SKILL.md; vixxo-spm-invoice-concerns/SKILL.md"),
        ("Invoice Creation Pending", "AP review / hold", "Invoice creation pending in billing system", "vixxo-freshdesk-invoice-review/SKILL.md; vixxo-spm-invoice-concerns/SKILL.md"),
        ("RejectedDoNotProcess", "Rejection", "Rejected; do not process payment", "vixxo-freshdesk-invoice-review/SKILL.md; starbucks-completed-sr-report/SKILL.md"),
        ("Delinquent SC Invoice", "Delinquent / vendor billing", "SP missed Service Channel invoice SLA (also SR Substatus)", "vixxo-freshdesk-invoice-review/SKILL.md; sub-status-decoder.md"),
        ("Invoice Required for SR", "SR billing overlap", "Work complete; SP invoice not yet uploaded (SR Substatus in VixxoLink)", "sub-status-decoder.md; vixxo-spm-invoice-concerns/SKILL.md"),
        ("AR Successful", "SR billing overlap", "AR booked invoice successfully (SR Substatus)", "sub-status-decoder.md"),
        ("No Invoice Pending", "SR billing overlap", "No invoice expected on SR", "sub-status-decoder.md"),
    ]


def build_invoice_tab_rows(live_rows: list[dict]) -> list[list]:
    rows: list[list] = []
    live_statuses = {r["status"] for r in live_rows}

    for live in live_rows:
        rows.append([
            live["status"],
            "invoice.status",
            live.get("category") or "Gateway live",
            "",
            live.get("source_ref") or "",
            live.get("sample_sr") or "",
            live.get("sample_sp") or "",
            live.get("origin") or "Gateway MCP",
        ])

    for value, category, meaning, source in documented_gateway_invoice_statuses():
        if value in live_statuses:
            continue
        rows.append([
            value,
            "invoice.status",
            category,
            meaning,
            source,
            "",
            "",
            "Workspace documented (not returned by live Gateway query)",
        ])

    rows.sort(key=lambda r: (r[2], r[0]))
    return rows


def _fill_row(ws, row_num: int, fill: PatternFill, cols: int = 8) -> None:
    for col in range(1, cols + 1):
        ws.cell(row=row_num, column=col).fill = fill


def ensure_invoice_tab(wb: Workbook, live_rows: list[dict], diagnostics: dict[str, str]) -> None:
    title = "SP Invoice Status (Gateway)"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)

    ws["A1"] = f"Gateway SP invoice status reference — compiled {TODAY}"
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=12)

    picklist_count = sum(1 for r in live_rows if r.get("category") == "Gateway picklist")
    sample_count = sum(1 for r in live_rows if r.get("category") == "Gateway live sample")

    if picklist_count:
        note = (
            f"Source: live Gateway MCP picklist via gateway_list_invoice_statuses "
            f"({picklist_count} values)"
        )
        if sample_count:
            note += f"; plus {sample_count} additional sampled values from gateway_search_invoices."
        note += " Green rows are live; orange rows are workspace-documented gaps."
    elif live_rows:
        note = (
            "Source: live Gateway MCP sampling via gateway_search_invoices only "
            f"({sample_count} values). Picklist call did not return data "
            f"({diagnostics.get('picklist', 'n/a')})."
        )
    else:
        note = (
            "Source: workspace-documented Gateway invoice.status values only. "
            f"Live Gateway MCP calls failed — picklist: {diagnostics.get('picklist', 'n/a')}; "
            f"sample: {diagnostics.get('sample', 'n/a')}. "
            "Authenticate Gateway MCP in Cursor (OAuth) or set GATEWAY_API_TOKEN / "
            "~/.vixxo/gateway_api_token, then re-run this script."
        )
    ws["A2"] = note
    ws.merge_cells("A2:H2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A2"].fill = META_FILL
    ws.row_dimensions[2].height = 55

    headers = [
        "Invoice Status",
        "Gateway Field",
        "Category",
        "Meaning / AP Notes",
        "Source Reference",
        "Sample SR",
        "Sample SP",
        "Data Origin",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws, ws.max_row)

    for row in build_invoice_tab_rows(live_rows):
        ws.append(row)
        origin = str(ws.cell(row=ws.max_row, column=8).value or "")
        category = str(ws.cell(row=ws.max_row, column=3).value or "")
        if "picklist" in category.lower():
            _fill_row(ws, ws.max_row, PICKLIST_FILL)
        elif "live" in origin.lower() or "live" in category.lower():
            _fill_row(ws, ws.max_row, LIVE_FILL)
        elif "documented" in origin.lower():
            _fill_row(ws, ws.max_row, DOC_FILL)

    autosize(ws)


def touch_sheet_dates(wb: Workbook) -> None:
  date_re = re.compile(r"compiled \d{4}-\d{2}-\d{2}")
  for name in ("SR Status", "SR Sub-Status"):
      if name not in wb.sheetnames:
          continue
      cell = wb[name]["A1"]
      if cell.value and date_re.search(str(cell.value)):
          cell.value = date_re.sub(f"compiled {TODAY}", str(cell.value))


def main() -> int:
    live_rows, diagnostics = gateway_live_invoice_statuses()
    if OUT.exists():
        wb = load_workbook(OUT)
    else:
        wb = Workbook()
    touch_sheet_dates(wb)
    ensure_invoice_tab(wb, live_rows, diagnostics)
    wb.save(OUT)
    print(f"Updated {OUT}")
    print(f"Live Gateway invoice statuses: {len(live_rows)}")
    print(f"Diagnostics: {json.dumps(diagnostics)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
