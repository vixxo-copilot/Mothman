#!/usr/bin/env python3
"""Extract candidate rate mismatches from the weekly RBAT to ROR workbook."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RATE_RE = re.compile(
    r"\$?\s*(?P<amount>\d+(?:\.\d{1,2})?)\s*(?P<kind>labor|travel|trip|helper|ot|overtime|rate)?",
    re.IGNORECASE,
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def one_line(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def find_header_row(rows: list[tuple[Any, ...]]) -> int:
    required = {"sr number", "service provider name", "comments"}
    for index, row in enumerate(rows):
        normalized = {clean(cell).lower() for cell in row}
        if required.issubset(normalized):
            return index
    raise ValueError("Could not find detail header row with SR Number, Service Provider Name, and comments.")


def extract_rates(comment: str) -> list[dict[str, str]]:
    rates: list[dict[str, str]] = []
    for match in RATE_RE.finditer(comment):
        kind = (match.group("kind") or "unknown").lower()
        rates.append({"amount": match.group("amount"), "kind": kind})
    return rates


def row_value(row: dict[str, str], *names: str) -> str:
    lowered = {key.lower(): value for key, value in row.items()}
    for name in names:
        value = lowered.get(name.lower())
        if value:
            return value
    return ""


def unique_headers(raw_headers: tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for idx, cell in enumerate(raw_headers):
        base = clean(cell) or f"Column {idx + 1}"
        key = base.lower()
        seen[key] = seen.get(key, 0) + 1
        headers.append(base if seen[key] == 1 else f"{base} {seen[key]}")
    return headers


def load_detail_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    best_rows: list[dict[str, Any]] = []

    for sheet in workbook.worksheets:
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            continue
        try:
            header_index = find_header_row(raw_rows[:25])
        except ValueError:
            continue

        headers = unique_headers(raw_rows[header_index])
        rows: list[dict[str, Any]] = []
        for raw in raw_rows[header_index + 1 :]:
            values = [clean(cell) for cell in raw]
            if not any(values):
                continue
            row = dict(zip(headers, values))
            comment = row_value(row, "comments")
            categorized = row_value(row, "Categorized comments")
            rates = extract_rates(comment)
            row["_sheet"] = sheet.title
            row["_rates_found"] = rates
            row["_rate_amounts"] = ", ".join(rate["amount"] for rate in rates)
            row["_rate_kinds"] = ", ".join(sorted({rate["kind"] for rate in rates if rate["kind"]}))
            row["_review_key"] = " | ".join(
                value
                for value in [
                    row_value(row, "SR Number"),
                    row_value(row, "Billing Reference Number", "VIID"),
                    row_value(row, "Service Provider Name"),
                    row_value(row, "Line of Service"),
                    row_value(row, "Short Description"),
                ]
                if value
            )
            row["_is_rate_issue"] = "rate" in comment.lower() or "rate" in categorized.lower()
            rows.append(row)

        if len(rows) > len(best_rows):
            best_rows = rows

    if not best_rows:
        raise ValueError("No detail sheet found in workbook.")
    return best_rows


def write_outputs(rows: list[dict[str, Any]], out_dir: Path, workbook_path: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    rate_rows = [row for row in rows if row.get("_is_rate_issue")]
    csv_path = out_dir / "rate-issues-normalized.csv"
    json_path = out_dir / "rate-issues-normalized.json"
    summary_path = out_dir / "rate-issues-summary.md"

    fieldnames: list[str] = []
    for row in rate_rows:
        for key in row:
            if key == "_rates_found":
                continue
            if key not in fieldnames:
                fieldnames.append(key)

    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rate_rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})

    with json_path.open("w", encoding="utf-8") as handle:
        json.dump(rate_rows, handle, indent=2, ensure_ascii=False)

    by_provider: dict[str, int] = {}
    by_customer: dict[str, int] = {}
    for row in rate_rows:
        provider = row_value(row, "Service Provider Name") or "Unknown provider"
        customer = row_value(row, "Customer Site-Customer Number") or "Unknown customer"
        by_provider[provider] = by_provider.get(provider, 0) + 1
        by_customer[customer] = by_customer.get(customer, 0) + 1

    provider_lines = "\n".join(f"- {name}: {count}" for name, count in sorted(by_provider.items(), key=lambda item: item[1], reverse=True)[:15])
    customer_lines = "\n".join(f"- {name}: {count}" for name, count in sorted(by_customer.items(), key=lambda item: item[1], reverse=True)[:15])
    sample_lines = "\n".join(
        f"- {row.get('_review_key', '')}: {one_line(row_value(row, 'comments'))[:180]}"
        for row in rate_rows[:20]
    )

    summary_path.write_text(
        "\n".join(
            [
                "# RBAT to ROR Rate Issue Extract",
                "",
                f"Workbook: `{workbook_path}`",
                f"Rows in detail sheet: {len(rows)}",
                f"Rate issue rows: {len(rate_rows)}",
                "",
                "## Top Providers",
                provider_lines or "- None",
                "",
                "## Top Customers",
                customer_lines or "- None",
                "",
                "## First Rows for Gateway Review",
                sample_lines or "- None",
                "",
                "## Outputs",
                f"- `{csv_path}`",
                f"- `{json_path}`",
            ]
        ),
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out", type=Path, default=Path(".tmp/rbat-ror-rate-review"))
    args = parser.parse_args()

    rows = load_detail_rows(args.workbook)
    write_outputs(rows, args.out, args.workbook)
    print(f"Extracted {sum(1 for row in rows if row.get('_is_rate_issue'))} rate issue rows to {args.out}")


if __name__ == "__main__":
    main()
