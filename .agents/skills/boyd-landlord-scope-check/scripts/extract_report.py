#!/usr/bin/env python3
"""Extract BOYD landlord/tenant responsibility workbook to CSV for skill lookup."""
from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: python3 -m pip install openpyxl") from exc

DEFAULT_WORKBOOK = Path.home() / "Downloads" / "Landlord-Tenant Responsibility Report.XLSX"
SKILL_DIR = Path(__file__).resolve().parents[1]
DEFAULT_CSV = SKILL_DIR / "data" / "responsibility_report.csv"
DEFAULT_METADATA = SKILL_DIR / "data" / "report_metadata.json"
EXPECTED_HEADERS = [
    "Client Lease ID",
    "Client Value 2",
    "Address 1",
    "City",
    "State",
    "Country",
    "Lease Status",
    "Primary Use",
    "Landlord Name",
    "Repairs - Item",
    "Repairs - Responsible Party",
    "Repairs - Reimbursable By",
    "Repairs - Comments",
]


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def extract(workbook: Path, csv_path: Path, metadata_path: Path) -> dict:
    if not workbook.exists():
        raise SystemExit(f"Workbook not found: {workbook}")
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # title row
    headers = [clean(c) for c in next(rows)]
    if headers != EXPECTED_HEADERS:
        raise SystemExit(f"Unexpected headers: {headers}")

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    site_numbers = set()
    repair_items = set()
    count = 0
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=EXPECTED_HEADERS)
        writer.writeheader()
        for row in rows:
            rec = {h: clean(v) for h, v in zip(headers, row)}
            if not any(rec.values()):
                continue
            writer.writerow(rec)
            count += 1
            if rec.get("Client Value 2"):
                site_numbers.add(rec["Client Value 2"].upper())
            if rec.get("Repairs - Item"):
                repair_items.add(rec["Repairs - Item"])
    wb.close()

    meta = {
        "source_workbook": str(workbook),
        "source_mtime": datetime.fromtimestamp(workbook.stat().st_mtime, timezone.utc).isoformat(),
        "extracted_at": datetime.now(timezone.utc).isoformat(),
        "csv_path": str(csv_path),
        "row_count": count,
        "site_count": len(site_numbers),
        "repair_items": sorted(repair_items),
    }
    metadata_path.write_text(json.dumps(meta, indent=2), encoding="utf-8")
    return meta


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--metadata", type=Path, default=DEFAULT_METADATA)
    args = parser.parse_args()
    meta = extract(args.workbook, args.csv, args.metadata)
    print(json.dumps(meta, indent=2))


if __name__ == "__main__":
    main()
