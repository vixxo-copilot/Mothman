#!/usr/bin/env python3
"""Check a BOYD/Gerber SR scope against the responsibility report."""
from __future__ import annotations

import argparse
import csv
import difflib
import html
import json
import re
from pathlib import Path

SKILL_DIR = Path(__file__).resolve().parents[1]
DEFAULT_CSV = SKILL_DIR / "data" / "responsibility_report.csv"
DEFAULT_METADATA = SKILL_DIR / "data" / "report_metadata.json"
DEFAULT_WORKBOOK = Path.home() / "Downloads" / "Landlord-Tenant Responsibility Report.XLSX"

TRIGGER_KEYWORDS = {
    "exterior signage": ["sign install", "sign removal", "sign replacement", "exterior sign", "signage"],
    "shared/common area": ["common area", "lobby", "hallway", "shared wall", "loading dock", "driveway"],
    "new electrical run": ["new electrical", "new line", "rewire", "rewiring", "panel upgrade"],
    "irrigation/shared landscaping": ["irrigation", "sprinkler", "landscaping"],
    "capital HVAC": ["coil", "compressor", "rtu", "rooftop", "unit replacement", "replace hvac", "full unit"],
    "exterior/below-grade plumbing": ["below grade", "underground", "main line", "sewer", "shared restroom", "outside tenant"],
    "structural": ["roof", "foundation", "structural", "exterior wall", "cutting into", "gutter"],
    "exterior painting": ["exterior paint", "paint refresh", "painting exterior", "paint exterior"],
    "parking/sidewalk": ["parking lot", "sidewalk", "repave", "striping", "pothole", "drainage"],
    "exterior lighting": ["wall pack", "wall packs", "pole mounted", "pole light", "parking lot light", "exterior lighting"],
    "capital lighting replacement": ["demo and install", "fixture replacement", "replace fixtures", "high bay fixtures", "led retrofit"],
}

ITEM_RULES = [
    ("Painting Exterior", ["exterior paint", "paint exterior", "painting exterior", "paint refresh"]),
    ("Painting Interior", ["interior paint", "paint interior", "painting interior"]),
    ("HVAC PM", ["hvac pm", "preventive maintenance", "planned maintenance"]),
    ("HVAC Replacement", ["hvac replacement", "replace hvac", "replace unit", "rooftop unit", "rtu", "compressor", "coil"]),
    ("HVAC Repair", ["hvac", "too hot", "too cold", "air conditioning", "heating", "cooling"]),
    ("Roof Replacement", ["roof replacement", "replace roof"]),
    ("Roof Repair / Gutters", ["roof", "gutter", "leak from roof"]),
    ("Parking Lot / Sidewalk Replacement", ["parking lot replacement", "sidewalk replacement", "repave"]),
    ("Parking Lot / Sidewalk Repairs", ["parking", "sidewalk", "pothole", "striping", "drainage"]),
    ("Plumbing underground", ["underground", "below grade", "main line", "sewer", "oil-water", "lift station"]),
    ("Plumbing interior", ["plumbing", "toilet", "sink", "drain", "faucet", "water leak"]),
    ("Signage", ["sign", "signage"]),
    ("Electrical Systems Panels", ["panel", "breaker", "electrical system"]),
    ("Electrical Lighting", ["lighting", "lights", "bulb"]),
    ("Doors Overhead", ["overhead door", "garage door", "bay door"]),
    ("Doors Manual", ["manual door", "entry door", "door"]),
    ("Locks", ["lock", "key"]),
    ("Windows / Glass", ["window", "glass"]),
    ("Fence / Gates", ["fence", "gate"]),
    ("Fire Life Safety System Inspections & Repairs", ["fire life", "fire alarm", "sprinkler system", "fls"]),
    ("Fire Extiguishers", ["extinguisher"]),
    ("Exterior Common Area Maintenance", ["exterior common", "common area", "driveway", "loading dock"]),
    ("Structural Walls Repairs/Foundations", ["foundation", "structural wall", "exterior wall"]),
]

PASS_THROUGH = {
    "HVAC": "Lennox KS68581",
    "FLS": "Academy KS68696",
    "ROOF": "Let's Roof KS68591",
    "PAVING": "Let's Pave LLC KS68595",
}

SEARCH_STOP_TERMS = {
    "avenue",
    "ave",
    "boulevard",
    "blvd",
    "circle",
    "cir",
    "court",
    "ct",
    "drive",
    "dr",
    "highway",
    "hwy",
    "lane",
    "ln",
    "parkway",
    "pkwy",
    "place",
    "pl",
    "road",
    "rd",
    "route",
    "rt",
    "street",
    "st",
    "suite",
    "ste",
    "tx",
    "texas",
    "united",
    "states",
    "usa",
}

ADDRESS_STOP_TERMS = SEARCH_STOP_TERMS | {
    "al",
    "ak",
    "az",
    "ar",
    "ca",
    "co",
    "ct",
    "de",
    "fl",
    "ga",
    "hi",
    "ia",
    "id",
    "il",
    "in",
    "ks",
    "ky",
    "la",
    "ma",
    "md",
    "me",
    "mi",
    "mn",
    "mo",
    "ms",
    "mt",
    "nc",
    "nd",
    "ne",
    "nh",
    "nj",
    "nm",
    "nv",
    "ny",
    "oh",
    "ok",
    "or",
    "pa",
    "ri",
    "sc",
    "sd",
    "tn",
    "ut",
    "va",
    "vt",
    "wa",
    "wi",
    "wv",
    "wy",
    "canada",
}


def norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (text or "").lower()).strip()


def address_tokens(text: str) -> list[str]:
    tokens = norm(text).split()
    return [token for token in tokens if token not in ADDRESS_STOP_TERMS and not re.fullmatch(r"\d{5}(?:\d{4})?", token)]


def words_fuzzy_overlap(candidate_words: set[str], supplied_words: set[str]) -> int:
    overlap = 0
    used_supplied: set[str] = set()
    for candidate_word in sorted(candidate_words, key=len, reverse=True):
        for supplied_word in sorted(supplied_words - used_supplied, key=len, reverse=True):
            if candidate_word == supplied_word:
                used_supplied.add(supplied_word)
                overlap += 1
                break
            if (
                min(len(candidate_word), len(supplied_word)) >= 4
                and (candidate_word in supplied_word or supplied_word in candidate_word)
            ):
                used_supplied.add(supplied_word)
                overlap += 1
                break
            if difflib.SequenceMatcher(None, candidate_word, supplied_word).ratio() >= 0.88:
                used_supplied.add(supplied_word)
                overlap += 1
                break
    return overlap


def address_matches(candidate: str, supplied: str) -> bool:
    candidate_tokens = address_tokens(candidate)
    supplied_tokens = address_tokens(supplied)
    if not candidate_tokens or not supplied_tokens:
        return False

    candidate_numbers = {token for token in candidate_tokens if token.isdigit()}
    supplied_numbers = {token for token in supplied_tokens if token.isdigit()}
    if candidate_numbers and supplied_numbers and candidate_numbers.isdisjoint(supplied_numbers):
        return False

    candidate_words = {token for token in candidate_tokens if not token.isdigit() and len(token) > 1}
    supplied_words = {token for token in supplied_tokens if not token.isdigit() and len(token) > 1}
    if not candidate_words or not supplied_words:
        return False

    overlap = words_fuzzy_overlap(candidate_words, supplied_words)
    required_overlap = min(len(candidate_words), 2)
    if overlap >= required_overlap:
        return True

    # Handle joined/split street-name variants like "Westshore" vs "West Shore"
    # after the street number has already matched.
    candidate_joined = "".join(sorted(candidate_words))
    supplied_joined = "".join(sorted(supplied_words))
    return difflib.SequenceMatcher(None, candidate_joined, supplied_joined).ratio() >= 0.88


def load_rows(csv_path: Path) -> list[dict]:
    if not csv_path.exists():
        raise SystemExit(f"Extracted CSV not found: {csv_path}. Run extract_report.py first.")
    with csv_path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def load_metadata(metadata_path: Path) -> dict:
    if not metadata_path.exists():
        return {}
    try:
        return json.loads(metadata_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"metadata_error": f"Could not parse metadata JSON at {metadata_path}"}


def search_values(*values: str | None) -> list[str]:
    terms: list[str] = []
    for value in values:
        if not value:
            continue
        raw = str(value).strip()
        if not raw:
            continue
        terms.append(raw)
        cc_match = re.search(r"\bCC(\d{4,})\b", raw, flags=re.IGNORECASE)
        if cc_match:
            terms.append(cc_match.group(1))
        for token in re.split(r"[^A-Za-z0-9]+", raw):
            token = token.strip()
            if len(token) >= 4 and token.lower() not in SEARCH_STOP_TERMS:
                terms.append(token)
    deduped: list[str] = []
    seen = set()
    for term in terms:
        key = term.lower()
        if key not in seen:
            seen.add(key)
            deduped.append(term)
    return deduped


def term_matches(haystack: str, term: str) -> bool:
    if not term:
        return False
    if re.fullmatch(r"[A-Za-z0-9]+", term):
        return re.search(rf"(?<![A-Za-z0-9]){re.escape(term)}(?![A-Za-z0-9])", haystack, flags=re.IGNORECASE) is not None
    return term.lower() in haystack.lower()


def csv_search_count(rows: list[dict], terms: list[str]) -> int:
    if not terms:
        return 0
    count = 0
    for row in rows:
        haystack = " | ".join(str(value or "") for value in row.values())
        if any(term_matches(haystack, term) for term in terms):
            count += 1
    return count


def workbook_search(workbook_path: Path, terms: list[str], limit: int = 20) -> dict:
    if not terms:
        return {"searched": False, "reason": "No search terms supplied.", "match_count_returned": 0, "matches": []}
    if not workbook_path.exists():
        return {
            "searched": False,
            "reason": f"Workbook not found: {workbook_path}",
            "workbook": str(workbook_path),
            "match_count_returned": 0,
            "matches": [],
        }
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "searched": False,
            "reason": "openpyxl is not installed.",
            "workbook": str(workbook_path),
            "match_count_returned": 0,
            "matches": [],
        }

    matches = []
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_text = " | ".join("" if value is None else str(value) for value in row)
                matched_terms = [term for term in terms if term_matches(row_text, term)]
                if matched_terms:
                    matches.append(
                        {
                            "sheet": ws.title,
                            "row": row_index,
                            "matched_terms": matched_terms,
                            "preview": row_text[:500],
                        }
                    )
                    if len(matches) >= limit:
                        return {
                            "searched": True,
                            "workbook": str(workbook_path),
                            "terms": terms,
                            "match_count_returned": len(matches),
                            "matches": matches,
                            "truncated": True,
                        }
    finally:
        wb.close()
    return {
        "searched": True,
        "workbook": str(workbook_path),
        "terms": terms,
        "match_count_returned": len(matches),
        "matches": matches,
        "truncated": False,
    }


def site_matches(rows: list[dict], site: str | None, address: str | None) -> list[dict]:
    if site:
        site_n = site.strip().upper()
        matches = [r for r in rows if (r.get("Client Value 2") or "").strip().upper() == site_n]
        if matches:
            return matches
    if address:
        matches = [
            r
            for r in rows
            if address_matches(" ".join([r.get("Address 1", ""), r.get("City", ""), r.get("State", ""), r.get("Country", "")]), address)
        ]
        if matches:
            return matches
    return []


def match_item(scope: str, items: list[str]) -> tuple[str | None, list[str]]:
    s = norm(scope)
    for item, phrases in ITEM_RULES:
        if any(norm(p) in s for p in phrases):
            return item, []
    close = difflib.get_close_matches(scope, items, n=5, cutoff=0.45)
    return (close[0] if close else None), close


def triggers(scope: str) -> list[str]:
    s = norm(scope)
    hits = []
    for name, phrases in TRIGGER_KEYWORDS.items():
        if any(norm(p) in s for p in phrases):
            hits.append(name)
    return hits


def pass_through_provider(line_of_service: str, repair_item: str, priority: str) -> str | None:
    if "P1" in (priority or "").upper():
        return None
    hay = norm(" ".join([line_of_service or "", repair_item or ""]))
    if "hvac" in hay:
        return PASS_THROUGH["HVAC"]
    if "fire" in hay or "fls" in hay:
        return PASS_THROUGH["FLS"]
    if "roof" in hay:
        return PASS_THROUGH["ROOF"]
    if "parking" in hay or "paving" in hay or "sidewalk" in hay:
        return PASS_THROUGH["PAVING"]
    return None


def classify(row: dict, trigger_hits: list[str], provider: str | None) -> tuple[str, str, str]:
    responsible = (row.get("Repairs - Responsible Party") or "").strip().lower()
    reimbursable = (row.get("Repairs - Reimbursable By") or "").strip().lower()
    if responsible == "landlord":
        return "landlord_responsible", "high", "Workbook assigns responsible party to Landlord."
    if responsible == "tenant" and reimbursable == "tenant" and not trigger_hits:
        return "pass_through_direct_dispatch" if provider else "tenant_responsible", "high", "Workbook assigns tenant responsibility."
    if responsible == "tenant" and reimbursable == "tenant" and trigger_hits:
        return "lease_review_required", "medium", "Workbook says tenant, but SOP trigger(s) require lease/approval review."
    if provider and responsible not in {"landlord"}:
        return "pass_through_direct_dispatch", "medium", "Pass-through trade and workbook does not assign landlord responsibility."
    return "lease_review_required", "medium", "Workbook is silent, mixed, blank, or not clean tenant/landlord responsibility."


def lease_expert_assessment(recommendation: str, row: dict, trigger_hits: list[str]) -> str:
    responsible = row.get("Repairs - Responsible Party") or ""
    reimbursable = row.get("Repairs - Reimbursable By") or ""
    repair_item = row.get("Repairs - Item") or "the matched repair item"
    if recommendation == "landlord_responsible":
        return f"Tenant-side lease review favors landlord handling because the workbook assigns {repair_item} to the landlord."
    if recommendation == "tenant_responsible":
        return f"Tenant-side lease review can support Boyd/Vixxo handling because the workbook assigns both responsibility and reimbursement for {repair_item} to the tenant."
    if recommendation == "pass_through_direct_dispatch":
        return f"Tenant-side lease review does not identify landlord responsibility for {repair_item}; route only under the approved pass-through process."
    if trigger_hits:
        return f"Tenant-side lease review should not treat this as clean tenant work because {repair_item} has approval/lease-review trigger(s): {', '.join(trigger_hits)}."
    return f"Tenant-side lease review should keep this in lease review because the workbook result is not clean tenant responsibility: responsible party = {responsible}, reimbursable by = {reimbursable}."


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--site", help="BOYD site number, e.g. CC123011")
    parser.add_argument("--address", help="Address fallback if site number is unknown")
    parser.add_argument("--scope", required=True, help="SR short description + description")
    parser.add_argument("--line-of-service", default="")
    parser.add_argument("--priority", default="")
    parser.add_argument("--sr", default="")
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--metadata", type=Path, default=DEFAULT_METADATA)
    parser.add_argument("--workbook", type=Path, help="Source workbook to verify site-not-found misses")
    parser.add_argument("--skip-workbook-search", action="store_true")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    rows = load_rows(args.csv)
    metadata = load_metadata(args.metadata)
    workbook = args.workbook or Path(metadata.get("source_workbook") or DEFAULT_WORKBOOK)
    all_scope = " ".join([args.line_of_service, args.scope])
    scope_triggers = triggers(all_scope)
    site_rows = site_matches(rows, args.site, args.address)
    if not site_rows:
        searched_terms = search_values(args.site, args.address)
        xlsx_search = (
            {"searched": False, "reason": "Skipped by --skip-workbook-search.", "match_count_returned": 0, "matches": []}
            if args.skip_workbook_search
            else workbook_search(workbook, searched_terms)
        )
        result = {
            "sr": args.sr,
            "site": args.site,
            "address": args.address,
            "recommendation": "lease_review_required",
            "diagnostic_code": "site_not_found_in_responsibility_report",
            "confidence": "low",
            "rationale": "No matching site found in extracted responsibility report.",
            "searched_site_identifiers": [args.site] if args.site else [],
            "searched_address_terms": search_values(args.address),
            "csv_path": str(args.csv),
            "csv_match_count": csv_search_count(rows, searched_terms),
            "source_workbook": str(workbook),
            "source_xlsx_search": xlsx_search,
            "report_metadata": metadata,
            "scope_review_triggers": scope_triggers,
            "tenant_side_lease_expert_assessment": (
                "Tenant-side lease review should hold this in lease review because the responsibility report does not contain "
                "a site-specific row or lease citation to support clean tenant responsibility."
            ),
            "next_action": (
                "Keep in landlord review / lease review. Verify BOYD site mapping in VixxoLink and CoStar or refresh/correct "
                "the monthly responsibility workbook before dispatching as tenant or landlord responsibility."
            ),
        }
    else:
        items = sorted({r.get("Repairs - Item", "") for r in site_rows if r.get("Repairs - Item")})
        item, alternatives = match_item(all_scope, items)
        row = next((r for r in site_rows if r.get("Repairs - Item") == item), None)
        if not row:
            result = {
                "sr": args.sr,
                "site": args.site,
                "recommendation": "repair_item_not_found",
                "confidence": "low",
                "rationale": "Site found, but scope could not be matched to a repair item.",
                "candidate_items": alternatives or items[:20],
                "report_metadata": metadata,
                "scope_review_triggers": scope_triggers,
                "next_action": "Manually choose closest repair item from workbook, then review lease if ambiguous.",
            }
        else:
            trigger_hits = scope_triggers
            provider = pass_through_provider(args.line_of_service, row.get("Repairs - Item", ""), args.priority)
            recommendation, confidence, rationale = classify(row, trigger_hits, provider)
            if recommendation == "landlord_responsible":
                next_action = "Reassign SR to BYD_LLD, set spend category Landlord, notify landlord with Boyd regional contact copied, and follow the SOP follow-up/billing process."
            elif recommendation == "tenant_responsible":
                next_action = "Move SR from BYD_LLDREV to BYD_TL01 for normal Boyd/Vixxo handling."
            elif recommendation == "pass_through_direct_dispatch":
                next_action = f"Dispatch via pass-through process to {provider}; do not use for P1 emergencies."
            else:
                next_action = "Hold in landlord review and verify lease/approval path in CoStar or with Boyd LL Request Admin before dispatch."
            lease_citation = html.unescape(row.get("Repairs - Comments") or "")
            result = {
                "sr": args.sr,
                "site": row.get("Client Value 2"),
                "client_lease_id": row.get("Client Lease ID"),
                "address": ", ".join(html.unescape(x) for x in [row.get("Address 1"), row.get("City"), row.get("State"), row.get("Country")] if x),
                "lease_status": row.get("Lease Status"),
                "landlord_name": html.unescape(row.get("Landlord Name") or ""),
                "repair_item": row.get("Repairs - Item"),
                "responsible_party": row.get("Repairs - Responsible Party"),
                "reimbursable_by": row.get("Repairs - Reimbursable By"),
                "recommendation": recommendation,
                "confidence": confidence,
                "rationale": rationale,
                "tenant_side_lease_expert_assessment": lease_expert_assessment(recommendation, row, trigger_hits),
                "lease_review_triggers": trigger_hits,
                "pass_through_provider": provider,
                "lease_citation": lease_citation,
                "comments": lease_citation,
                "report_metadata": metadata,
                "next_action": next_action,
            }

    if args.json:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(f"Recommendation: {result.get('recommendation')} ({result.get('confidence')})")
        for key in ["sr", "site", "client_lease_id", "address", "repair_item", "responsible_party", "reimbursable_by", "landlord_name", "diagnostic_code"]:
            if result.get(key):
                print(f"{key}: {result[key]}")
        if result.get("lease_review_triggers"):
            print("lease_review_triggers: " + ", ".join(result["lease_review_triggers"]))
        if result.get("scope_review_triggers"):
            print("scope_review_triggers: " + ", ".join(result["scope_review_triggers"]))
        if result.get("pass_through_provider"):
            print("pass_through_provider: " + result["pass_through_provider"])
        if result.get("csv_match_count") is not None:
            print(f"csv_match_count: {result['csv_match_count']}")
        if result.get("source_xlsx_search"):
            print(f"source_xlsx_match_count: {result['source_xlsx_search'].get('match_count_returned', 0)}")
        print("rationale: " + result.get("rationale", ""))
        if result.get("tenant_side_lease_expert_assessment"):
            print("tenant_side_lease_expert_assessment: " + result["tenant_side_lease_expert_assessment"])
        print("next_action: " + result.get("next_action", ""))
        if result.get("lease_citation"):
            print("lease_citation: " + result["lease_citation"][:1500].replace("\n", " | "))


if __name__ == "__main__":
    main()
