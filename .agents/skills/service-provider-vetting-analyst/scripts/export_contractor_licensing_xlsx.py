#!/usr/bin/env python3
"""Export contractor licensing reference to Excel (template-style + municipal + specialty)."""

from __future__ import annotations

import copy
import json
import re
import shutil
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

SKILL_ROOT = Path(__file__).resolve().parent.parent
REF = SKILL_ROOT / "reference"
TEMPLATE = Path(
    r"c:\Users\CGagner\OneDrive - Vixxo\Desktop\Vixxo - Vendor Forms\License Resources\Contractor's License By State.xlsx"
)
QUICK_MD = REF / "contractor-licensing-quick-lookup-table.md"
SPECIALTY_MD = REF / "contractor-licensing-specialty-trades.md"
OUT_REPO = REF / "contractor-licensing-by-state.xlsx"
OUT_ONEDRIVE = TEMPLATE.parent / "Contractor's License By State (Updated).xlsx"

REVIEW_DATE = "2026-06-30"

# Template state name -> quick-lookup jurisdiction code
STATE_TO_CODE: dict[str, str] = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshite": "NH",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington DC": "DC",
    "Washington": "WA",
    "West Virgina": "WV",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "Guam": "GU",
    "Puerto Rico": "PR",
    "US Virgin Islands": "VI",
    "Virgin Islands": "VI",
    "Northern Mariana Islands": "CNMI",
    "CNMI": "CNMI",
    "American Samoa": "AS",
}

CODE_TO_NAME: dict[str, str] = {
    "AL": "Alabama",
    "AK": "Alaska",
    "AZ": "Arizona",
    "AR": "Arkansas",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DE": "Delaware",
    "DC": "Washington DC",
    "FL": "Florida",
    "GA": "Georgia",
    "HI": "Hawaii",
    "ID": "Idaho",
    "IL": "Illinois",
    "IN": "Indiana",
    "IA": "Iowa",
    "KS": "Kansas",
    "KY": "Kentucky",
    "LA": "Louisiana",
    "ME": "Maine",
    "MD": "Maryland",
    "MA": "Massachusetts",
    "MI": "Michigan",
    "MN": "Minnesota",
    "MS": "Mississippi",
    "MO": "Missouri",
    "MT": "Montana",
    "NE": "Nebraska",
    "NV": "Nevada",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NY": "New York",
    "NC": "North Carolina",
    "ND": "North Dakota",
    "OH": "Ohio",
    "OK": "Oklahoma",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "RI": "Rhode Island",
    "SC": "South Carolina",
    "SD": "South Dakota",
    "TN": "Tennessee",
    "TX": "Texas",
    "UT": "Utah",
    "VT": "Vermont",
    "VA": "Virginia",
    "WA": "Washington",
    "WV": "West Virginia",
    "WI": "Wisconsin",
    "WY": "Wyoming",
    "GU": "Guam",
    "PR": "Puerto Rico",
    "VI": "US Virgin Islands",
    "CNMI": "Northern Mariana Islands",
    "AS": "American Samoa",
}

# Trade-specific verification URLs (supplement quick lookup)
TRADE_URLS: dict[str, dict[str, str]] = {
    "AL": {
        "electrical": "https://www.aecb.state.al.us/",
        "plumbing": "https://www.pgfb.state.al.us/",
        "hvac": "https://www.hvacboard.state.al.us/",
    },
    "AK": {
        "electrical": "https://www.commerce.alaska.gov/cbp/main/Search/Professional",
        "plumbing": "https://www.commerce.alaska.gov/cbp/main/Search/Professional",
        "hvac": "https://www.commerce.alaska.gov/cbp/main/Search/Professional",
    },
    "AZ": {
        "electrical": "https://azroc.my.site.com/AZRoc/s/contractor-search",
        "plumbing": "https://azroc.my.site.com/AZRoc/s/contractor-search",
        "hvac": "https://azroc.my.site.com/AZRoc/s/contractor-search",
    },
    "AR": {
        "electrical": "https://www.ark.org/labor/electrician/search.php",
        "plumbing": "https://www.healthy.arkansas.gov/programs-services/topics/plumbing-natural-gas",
        "hvac": "https://www.healthy.arkansas.gov/programs-services/topics/hvac-r",
    },
    "CA": {
        "electrical": "https://www.cslb.ca.gov/OnlineServices/CheckLicenseII/CheckLicense.aspx (C-10)",
        "plumbing": "https://www.cslb.ca.gov/OnlineServices/CheckLicenseII/CheckLicense.aspx (C-36)",
        "hvac": "https://www.cslb.ca.gov/OnlineServices/CheckLicenseII/CheckLicense.aspx (C-20/C-38)",
    },
    "CO": {
        "electrical": "https://dpo.colorado.gov/Lookup",
        "plumbing": "https://dpo.colorado.gov/Lookup",
        "hvac": "Local licensing — see Municipal Licensing sheet",
    },
    "CT": {
        "electrical": "https://www.elicense.ct.gov/Lookup/LicenseLookup.aspx",
        "plumbing": "https://www.elicense.ct.gov/Lookup/LicenseLookup.aspx",
        "hvac": "https://www.elicense.ct.gov/Lookup/LicenseLookup.aspx",
    },
    "FL": {
        "electrical": "https://www.myfloridalicense.com/wl11.asp (Electrical)",
        "plumbing": "https://www.myfloridalicense.com/wl11.asp (Construction/Plumbing)",
        "hvac": "https://www.myfloridalicense.com/wl11.asp (Construction)",
    },
    "IL": {
        "electrical": "Local — see Municipal Licensing sheet (Chicago DOB, etc.)",
        "plumbing": "https://online-dfpr.micropact.com/Lookup/LicenseLookup.aspx",
        "hvac": "Local — see Municipal Licensing sheet",
    },
    "IN": {
        "electrical": "Local — see Municipal Licensing sheet",
        "plumbing": "https://mylicense.in.gov/EVerification/Search.aspx",
        "hvac": "Local — see Municipal Licensing sheet",
    },
    "IA": {
        "electrical": "https://dial.iowa.gov/licenses/building/plumbing-mechanical/plumbing-licensure/contractor-license",
        "plumbing": "https://dial.iowa.gov/licenses/building/plumbing-mechanical/plumbing-licensure/contractor-license",
        "hvac": "https://dial.iowa.gov/licenses/building/plumbing-mechanical/plumbing-licensure/contractor-license",
    },
    "KY": {
        "electrical": "https://hbc.ky.gov/Pages/License-Search.aspx",
        "plumbing": "https://hbc.ky.gov/Pages/License-Search.aspx",
        "hvac": "https://hbc.ky.gov/Pages/License-Search.aspx",
    },
    "LA": {
        "electrical": "https://www.lslbc.louisiana.gov/contractor-search/",
        "plumbing": "https://www.lslbc.louisiana.gov/contractor-search/",
        "hvac": "https://www.lslbc.louisiana.gov/contractor-search/",
    },
    "MD": {
        "electrical": "https://www.dllr.state.md.us/license/mhic/mhic_search.shtml",
        "plumbing": "https://www.dllr.state.md.us/license/mhic/mhic_search.shtml",
        "hvac": "https://www.dllr.state.md.us/license/mhic/mhic_search.shtml",
    },
    "MI": {
        "electrical": "https://www.lara.michigan.gov/colaLicVerify/",
        "plumbing": "https://www.lara.michigan.gov/colaLicVerify/",
        "hvac": "https://www.lara.michigan.gov/colaLicVerify/",
    },
    "MN": {
        "electrical": "https://www.dli.mn.gov/business/contractors/license-and-registration-lookup",
        "plumbing": "https://www.dli.mn.gov/business/contractors/license-and-registration-lookup",
        "hvac": "https://www.dli.mn.gov/business/contractors/license-and-registration-lookup",
    },
    "MO": {
        "electrical": "https://pr.mo.gov/licensee-search.asp",
        "plumbing": "Local — see Municipal Licensing sheet",
        "hvac": "Local — see Municipal Licensing sheet",
    },
    "NE": {
        "electrical": "https://www.nebraska.gov/dol/contractor/search",
        "plumbing": "Local — Omaha/Lincoln; see Municipal Licensing sheet",
        "hvac": "Local — Omaha/Lincoln; see Municipal Licensing sheet",
    },
    "NJ": {
        "electrical": "https://newjersey.mylicense.com/verification/",
        "plumbing": "https://newjersey.mylicense.com/verification/",
        "hvac": "https://newjersey.mylicense.com/verification/",
    },
    "NY": {
        "electrical": "Local — see Municipal Licensing sheet (NYC DOB, etc.)",
        "plumbing": "Local — see Municipal Licensing sheet",
        "hvac": "Local — see Municipal Licensing sheet",
    },
    "NC": {
        "electrical": "https://nclicensing.org/",
        "plumbing": "https://nclicensing.org/",
        "hvac": "https://nclicensing.org/",
    },
    "OH": {
        "electrical": "https://elicense.ohio.gov/oh_verifylicense",
        "plumbing": "https://elicense.ohio.gov/oh_verifylicense",
        "hvac": "https://elicense.ohio.gov/oh_verifylicense",
    },
    "OK": {
        "electrical": "http://cibverify.ok.gov/",
        "plumbing": "http://cibverify.ok.gov/",
        "hvac": "http://cibverify.ok.gov/",
    },
    "OR": {
        "electrical": "https://www.oregon.gov/bcd/licensing/pages/search.aspx (BCD) + https://search.ccb.state.or.us/search/ (CCB)",
        "plumbing": "https://www.oregon.gov/bcd/licensing/pages/search.aspx (BCD) + https://search.ccb.state.or.us/search/ (CCB)",
        "hvac": "https://www.oregon.gov/bcd/licensing/pages/search.aspx (BCD) + https://search.ccb.state.or.us/search/ (CCB)",
    },
    "PA": {
        "electrical": "Local — Philadelphia L&I and municipality; see Municipal Licensing sheet",
        "plumbing": "Local — see Municipal Licensing sheet",
        "hvac": "Local — see Municipal Licensing sheet",
    },
    "TX": {
        "electrical": "https://www.tdlr.texas.gov/LicenseSearch/",
        "plumbing": "https://www.tsbpe.texas.gov/",
        "hvac": "https://www.tdlr.texas.gov/LicenseSearch/",
    },
    "WA": {
        "electrical": "https://secure.lni.wa.gov/verify/",
        "plumbing": "https://secure.lni.wa.gov/verify/",
        "hvac": "https://secure.lni.wa.gov/verify/",
    },
    "WY": {
        "electrical": "https://wsfm.wyo.gov/electrical-licensing",
        "plumbing": "Local — see Municipal Licensing sheet",
        "hvac": "Local — see Municipal Licensing sheet",
    },
}


def strip_md(text: str) -> str:
    text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    return text.strip()


def parse_md_table_lines(lines: list[str]) -> list[list[str]]:
    rows: list[list[str]] = []
    for line in lines:
        if not line.startswith("|"):
            break
        if re.match(r"^\|\s*-+", line):
            continue
        cells = [strip_md(c.strip()) for c in line.strip("|").split("|")]
        rows.append(cells)
    return rows


def parse_jurisdiction_lookup(md: str) -> dict[str, dict[str, str]]:
    start = md.find("| Jurisdiction | Regime |")
    if start < 0:
        return {}
    chunk = md[start:].splitlines()
    rows = parse_md_table_lines(chunk)
    out: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        if len(row) < 6:
            continue
        code = row[0].split()[0] if row[0] else row[0]
        out[code] = {
            "regime": row[1],
            "url": row[2],
            "gc": row[3],
            "trades": row[4],
            "muni": row[5],
        }
    return out


def parse_municipal_sections(md: str) -> list[dict[str, str]]:
    """Parse municipal tables only — stops before Specialty trades section."""
    records: list[dict[str, str]] = []

    def parse_chunk(chunk: str, default_category: str) -> None:
        current_state = ""
        current_category = default_category
        for line in chunk.splitlines():
            m = re.match(r"^### (.+?) \(([A-Z]{2,5})\)", line)
            if m:
                current_state = f"{m.group(1)} ({m.group(2)})"
                continue
            m_state = re.match(r"^## (.+?) \(([A-Z]{2,5})\)", line)
            if m_state and not line.startswith("###"):
                current_state = f"{m_state.group(1)} ({m_state.group(2)})"
                continue
            m2 = re.match(r"^## (.+)$", line)
            if m2 and not line.startswith("###"):
                title = m2.group(1)
                if title.startswith("Municipal"):
                    current_category = title
                continue
            if line.startswith("| Scope |") or line.startswith("| State |") or line.startswith("| Layer |"):
                continue
            if line.startswith("| ") and "---" not in line:
                cells = [strip_md(c.strip()) for c in line.strip("|").split("|")]
                if len(cells) >= 4 and cells[0] not in ("Scope", "State", "Layer"):
                    records.append(
                        {
                            "state": current_state or cells[0],
                            "category": current_category,
                            "scope": cells[0],
                            "authority": cells[1] if len(cells) > 1 else "",
                            "url": cells[2] if len(cells) > 2 else "",
                            "notes": cells[3] if len(cells) > 3 else "",
                        }
                    )

    or_start = md.find("## Oregon (OR)")
    muni_start = md.find("## Municipal trade")
    specialty_start = md.find("## Specialty trades")
    if or_start >= 0 and muni_start > or_start:
        parse_chunk(md[or_start:muni_start], "Oregon — dual state verification")
    if muni_start >= 0:
        end = specialty_start if specialty_start > muni_start else len(md)
        parse_chunk(md[muni_start:end], "Municipal licensing")

    return records


def split_locality_from_scope(scope: str) -> tuple[str, str]:
    """Return (locality label, scope text with locality context)."""
    scope = scope.strip()
    for sep in (" — ", " - ", " – "):
        if sep in scope:
            left, right = scope.split(sep, 1)
            left = left.strip()
            right = right.strip()
            if left.lower().startswith(("state", "other", "municipal", "town", "every")):
                return "", scope
            if left.lower().endswith("co.") or len(left.split()) <= 4:
                return left, f"{left} — {right}"
    return "", scope


def classify_muni_row(scope: str, authority: str, notes: str, category: str) -> list[str]:
    blob = f"{scope} {authority} {notes} {category}".lower()
    tags: list[str] = []
    if any(k in blob for k in ("fire alarm", "fire protection", "sprinkler", "fdny", "sfmo", "fems", "fire-line")):
        tags.append("Fire protection / alarm")
    if any(k in blob for k in ("backflow", "cross-connection", "bpat", "bat cert")):
        tags.append("Backflow")
    if any(k in blob for k in ("pest", "applicator", "structural pest", "wdo")):
        tags.append("Pest control")
    if any(k in blob for k in ("locksmith", "low voltage", "cctv", "access control", "burglar", "alarm")):
        tags.append("Locksmith / low voltage / alarm")
    if "electrical" in blob or "elec" in blob.split():
        tags.append("Electrical")
    if "plumb" in blob:
        tags.append("Plumbing")
    if any(k in blob for k in ("hvac", "mechanical", "refrigeration", "heating")):
        tags.append("HVAC / mechanical")
    if any(k in blob for k in ("general contractor", " gc", "gc ", "home improvement", "hic", "contractor reg", "contractor registration", "residential builder")):
        tags.append("General contractor / registration")
    if "roof" in blob:
        tags.append("Roofing")
    if "registration" in blob and "General contractor" not in " ".join(tags):
        if "reg" in blob or "registration" in blob:
            tags.append("Registration (administrative — not trade license)")
    if "business license" in blob:
        tags.append("Business license (not trade competency)")
    if "permit" in blob:
        tags.append("Permit / local AHJ")
    return tags


def enrich_muni_notes(record: dict[str, str]) -> str:
    scope = record.get("scope", "")
    notes = (record.get("notes") or "").strip()
    authority = record.get("authority", "")
    category = record.get("category", "")
    tags = classify_muni_row(scope, authority, notes, category)

    cat_hint = ""
    state_label = record.get("state", "")
    if "Texas" in state_label:
        if any(c in scope.lower() for c in ("austin", "dallas", "san antonio", "houston")):
            cat_hint = (
                "Municipal contractor registration (TX) — TDLR/TSBPE state trade licenses "
                "do not replace local GC registration. "
            )
        elif "state" in scope.lower():
            cat_hint = "Texas state trade licensing — GC registration is municipal. "
    elif "follow-up" in category.lower() or (
        "registration" in category.lower() and "Texas" not in state_label
    ):
        cat_hint = "State registration follow-up — confirm municipal permit eligibility. "
    elif "IL, IN, NY, MO, PA" in category or "trade & GC" in category.lower():
        cat_hint = "Municipal GC/trade licensing (no statewide GC). "
    elif "Oregon" in category or "Oregon" in state_label:
        cat_hint = "Oregon dual verification — CCB registration plus BCD trade license. "

    tag_text = f"Pertains to: {'; '.join(tags)}. " if tags else "Pertains to: municipal / local licensing lookup. "

    if not notes:
        body = f"Use verification URL to confirm active local license or registration for: {scope}."
    elif re.match(r"^https?://\S+$", notes):
        body = f"Registry lookup portal for {scope}. Confirm license type, status, and expiration at board site."
    elif notes.lower().startswith("http") and len(notes.split()) == 1:
        body = f"Registry lookup for {scope}. Confirm credential matches work scope."
    else:
        body = notes

    return f"{cat_hint}{tag_text}{body}"


def build_muni_scope(record: dict[str, str]) -> str:
    locality, scoped = split_locality_from_scope(record.get("scope", ""))
    if locality and locality not in scoped:
        return scoped
    return record.get("scope", scoped)


def write_municipal_sheet(wb: openpyxl.Workbook, records: list[dict[str, str]]) -> None:
    title = "Municipal Licensing"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    headers = ["State / Territory", "Scope", "Authority", "Verification URL", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header_row(ws, 1, len(headers))

    row = 2
    idx = 0
    while idx < len(records):
        state = records[idx]["state"]
        group: list[dict[str, str]] = []
        while idx < len(records) and records[idx]["state"] == state:
            group.append(records[idx])
            idx += 1
        start_row = row
        ws.cell(row, 1, state)
        for rec in group:
            ws.cell(row, 2, build_muni_scope(rec))
            ws.cell(row, 3, rec.get("authority", ""))
            ws.cell(row, 4, rec.get("url", ""))
            ws.cell(row, 5, enrich_muni_notes(rec))
            for col in range(1, 6):
                ws.cell(row, col).alignment = WRAP
                ws.cell(row, col).border = BORDER
            row += 1
        if row - 1 > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
            ws.cell(start_row, 1).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 55
    if row > 2:
        ws.auto_filter.ref = f"A1:E{row - 1}"


def apply_sheet_filter(ws, header_row: int, max_col: int) -> None:
    max_row = ws.max_row
    if max_row > header_row:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(max_col)}{max_row}"
        )



def norm_header(text: str) -> str:
    return re.sub(r"\s+", " ", strip_md(text).lower())


def format_state_territory(code: str) -> str:
    code = strip_md(code).strip()
    if not code:
        return code
    name = CODE_TO_NAME.get(code)
    if name:
        return f"{name} ({code})"
    return code


def extract_urls(text: str) -> list[str]:
    return re.findall(r"https?://[^\s|·]+", text or "")


def enrich_specialty_row_note(specialty: str, row: list[str]) -> str:
    base = SPECIALTY_NOTE_TEMPLATES.get(specialty, f"Pertains to: {specialty}.")
    detail = " ".join(cell for cell in row[1:] if cell).strip()
    if not detail:
        return base
    if re.match(r"^https?://", detail) and len(detail.split()) == 1:
        return f"{base} Registry lookup — confirm active status, scope, and expiration."
    if "contact" in detail.lower() and "http" not in detail.lower():
        return f"{base} No statewide online lookup — contact agency or local water purveyor/AHJ."
    return f"{base} {detail[:240]}{'...' if len(detail) > 240 else ''}"


def parse_specialty_table(md: str, heading: str, headers: tuple[str, ...]) -> list[list[str]]:
    idx = md.find(heading)
    if idx < 0:
        return []
    chunk = md[idx:].splitlines()[1:]
    table_start = next((i for i, line in enumerate(chunk) if line.startswith("|")), None)
    if table_start is None:
        return []
    rows = parse_md_table_lines(chunk[table_start:])
    if not rows:
        return []
    if norm_header(rows[0][0]) == norm_header(headers[0]):
        rows = rows[1:]
    return [row for row in rows if row and row[0] and norm_header(row[0]) != norm_header(headers[0])]


def write_specialty_sheet(
    wb: openpyxl.Workbook,
    title: str,
    headers: list[str],
    rows: list[list[str]],
    col_widths: list[float],
    url_col_indexes: list[int],
) -> int:
    """Write specialty trade sheet with explicit visible formatting and Excel table."""
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 100

    all_headers = headers + ["Notes"]
    for col, header in enumerate(all_headers, 1):
        ws.cell(1, col, header)
    style_header_row(ws, 1, len(all_headers))

    if not rows:
        ws.cell(2, 1, "NO DATA — regenerate from specialty reference markdown")
        ws.cell(2, 1).font = Font(bold=True, color="FF0000", size=12)
        return 0

    for r_idx, raw_row in enumerate(rows, 2):
        row = list(raw_row)
        while len(row) < len(headers):
            row.append("")
        row[0] = format_state_territory(row[0])
        notes = enrich_specialty_row_note(title, row)
        ws.row_dimensions[r_idx].height = 48
        for c_idx in range(1, len(all_headers) + 1):
            if c_idx == len(all_headers):
                value = notes
            elif c_idx <= len(row):
                value = row[c_idx - 1]
            else:
                value = ""
            cell = ws.cell(r_idx, c_idx, value)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.font = DATA_FONT
            if c_idx in url_col_indexes:
                urls = extract_urls(str(value))
                if urls:
                    cell.hyperlink = urls[0]
                    cell.font = LINK_FONT
                    if str(value).strip().startswith("http"):
                        cell.value = urls[0]

    last_row = len(rows) + 1
    last_col = get_column_letter(len(all_headers))
    table_name = re.sub(r"[^A-Za-z0-9]", "", title)[:20]
    if table_name and table_name[0].isdigit():
        table_name = "T" + table_name
    tab = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    for i, width in enumerate(col_widths + [55], 1):
        ws.column_dimensions[get_column_letter(i)].width = min(width, 90)
    return len(rows)


def regime_note(regime: str) -> str:
    notes = {
        "A": "Statewide GC license regime.",
        "B": "Trade-only at state level; GC often local.",
        "C": "No statewide GC license — municipal licensing required.",
        "A+B": "Dual verification required (OR: CCB + BCD).",
    }
    return notes.get(regime.split()[0], regime)


def build_gc_cell(info: dict[str, str], legacy: str | None) -> str:
    url = info.get("url", "")
    gc = info.get("gc", "")
    regime = info.get("regime", "")
    parts = [f"Regime {regime}: {regime_note(regime)}", gc]
    if url and not url.startswith("—"):
        parts.append(f"Verify: {url}")
    if info.get("muni", "").lower().startswith("yes") or regime in ("B", "C", "A+B"):
        parts.append("Municipal follow-up may be required — see Municipal Licensing sheet.")
    if regime == "A+B":
        parts.append("Oregon: verify BOTH CCB (https://search.ccb.state.or.us/search/) AND BCD (https://www.oregon.gov/bcd/licensing/pages/search.aspx).")
    text = "\n".join(p for p in parts if p)
    if legacy and len(legacy) > 80 and regime not in ("C",):
        return f"{text}\n\nPrior reference: {legacy[:400]}{'...' if len(legacy) > 400 else ''}"
    return text


def build_trade_cell(code: str, trade: str, info: dict[str, str], legacy: str | None) -> str:
    trade_urls = TRADE_URLS.get(code, {})
    url = trade_urls.get(trade, "")
    regime = info.get("regime", "")
    trades = info.get("trades", "")
    if url:
        return f"State license required.\nVerify: {url}\n({trades})"
    if regime == "C" or "local" in trades.lower():
        return f"Licensed locally in most jurisdictions.\nSee Municipal Licensing sheet.\n({trades})"
    primary = info.get("url", "")
    if primary and not primary.startswith("—"):
        return f"Verify at state portal: {primary}\n({trades})"
    if legacy:
        return legacy
    return f"Verify state trade board.\n({trades})"


def build_local_cell(code: str, info: dict[str, str], municipal: list[dict[str, str]], legacy: str | None) -> str:
    muni_flag = info.get("muni", "")
    regime = info.get("regime", "")
    lines: list[str] = []
    if muni_flag.lower().startswith("yes") or regime in ("C", "B"):
        lines.append("Municipal license/registration required for project city/county.")
    state_records = [r for r in municipal if f"({code})" in r.get("state", "")]
    for rec in state_records[:8]:
        scope = rec.get("scope", "")
        url = rec.get("url", "")
        if scope and url:
            lines.append(f"{scope}: {url}")
        elif scope:
            lines.append(f"{scope}: {rec.get('authority', '')}")
    if not lines and legacy:
        return legacy
    if not lines:
        if regime == "A" and not muni_flag.lower().startswith("yes"):
            return "No major municipal GC requirement documented statewide; confirm local permits."
        return "See Municipal Licensing sheet for city/county portals."
    if legacy and len(legacy) > 100:
        lines.append("")
        lines.append(f"Additional local notes: {legacy[:500]}{'...' if len(legacy) > 500 else ''}")
    return "\n".join(lines)


HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True, color="000000", size=11)
DATA_FONT = Font(name="Calibri", size=11, color="000000")
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SPECIALTY_NOTE_TEMPLATES = {
    "Backflow": (
        "Pertains to: Backflow prevention — BPAT/BAT tester certification and/or "
        "installer authorization. Plumber/GC license does not substitute for tester cert."
    ),
    "Fire Protection": (
        "Pertains to: Fire protection — sprinkler/suppression and fire alarm are often "
        "separate licenses. Verify the credential matching the work scope."
    ),
    "Pest Control": (
        "Pertains to: Structural pest control / commercial pesticide applicator — "
        "company license plus certified applicator/technician credentials."
    ),
    "Locksmith and Low Voltage": (
        "Pertains to: Locksmith, security alarm, CCTV, or low-voltage/access control — "
        "locksmith alone usually does not cover electronic access/CCTV."
    ),
}


def style_header_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_data_sheet(
    wb: openpyxl.Workbook,
    title: str,
    headers: list[str],
    rows: list[list[str]],
    col_widths: list[float] | None = None,
    enable_filter: bool = True,
) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header_row(ws, 1, len(headers))
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.alignment = WRAP
            cell.border = BORDER
    ws.freeze_panes = "A2"
    widths = col_widths or [18] + [40] * (len(headers) - 1)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(width, 80)
    if enable_filter and rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def build_summary_sheet(wb: openpyxl.Workbook) -> None:
    title = "Summary"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, 0)
    ws["A1"] = "US Contractor Licensing Reference"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Last reviewed: {REVIEW_DATE}"
    ws["A3"] = "Vixxo service provider vetting — official verification URLs"
    rows = [
        ("Sheet", "Contents"),
        ("States by State", "General contractor, electrical, plumbing, HVAC, and local requirements (template layout)"),
        ("Municipal Licensing", "City/county GC and trade licensing — IL, IN, NY, MO, PA, and registration-state follow-up"),
        ("Backflow", "Backflow prevention tester/installer verification by state"),
        ("Fire Protection", "Fire sprinkler, alarm, and suppression contractor verification"),
        ("Pest Control", "Structural pest and commercial applicator licensing"),
        ("Locksmith and Low Voltage", "Locksmith, alarm, CCTV, and low-voltage credentials"),
    ]
    for i, (a, b) in enumerate(rows, 5):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
        ws.cell(i, 1).font = HEADER_FONT if i == 5 else Font(bold=True if i == 5 else False)
    style_header_row(ws, 5, 2)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90


def update_states_sheet(
    ws,
    lookup: dict[str, dict[str, str]],
    municipal: list[dict[str, str]],
) -> None:
    ws.title = "States by State"
    # Template may contain merged cells in data rows — unmerge before writing
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 4:
            ws.unmerge_cells(str(merged))
    ws["C1"] = "State Requirements"
    ws["F1"] = "Local Requirements"
    for col, label in enumerate(
        ["State", "General Contractor", "Electrical", "Plumbing", "HVAC", "Cities / Counties"],
        1,
    ):
        ws.cell(2, col, label)
        ws.cell(3, col, label)
    style_header_row(ws, 2, 6)
    style_header_row(ws, 3, 6)

    existing: dict[str, int] = {}
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if name:
            existing[str(name).strip()] = row

    # Update existing template rows
    for name, row in existing.items():
        code = STATE_TO_CODE.get(name, "")
        info = lookup.get(code, {})
        legacy = {
            "gc": ws.cell(row, 2).value,
            "elec": ws.cell(row, 3).value,
            "plumb": ws.cell(row, 4).value,
            "hvac": ws.cell(row, 5).value,
            "local": ws.cell(row, 6).value,
        }
        if info:
            ws.cell(row, 2, build_gc_cell(info, str(legacy["gc"] or "")))
            ws.cell(row, 3, build_trade_cell(code, "electrical", info, str(legacy["elec"] or "")))
            ws.cell(row, 4, build_trade_cell(code, "plumbing", info, str(legacy["plumb"] or "")))
            ws.cell(row, 5, build_trade_cell(code, "hvac", info, str(legacy["hvac"] or "")))
            ws.cell(row, 6, build_local_cell(code, info, municipal, str(legacy["local"] or "")))
        # Fix known typos in state names
        if name == "New Hampshite":
            ws.cell(row, 1, "New Hampshire")
        if name == "West Virgina":
            ws.cell(row, 1, "West Virginia")
        for col in range(1, 7):
            ws.cell(row, col).alignment = WRAP
            ws.cell(row, col).border = BORDER

    # Append territories not in original template
    next_row = ws.max_row + 1
    template_codes = {STATE_TO_CODE.get(str(ws.cell(r, 1).value), "") for r in existing.values()}
    for code in ["GU", "PR", "VI", "CNMI", "AS"]:
        if code in template_codes:
            continue
        info = lookup.get(code, {})
        if not info:
            continue
        name = CODE_TO_NAME[code]
        ws.cell(next_row, 1, name)
        ws.cell(next_row, 2, build_gc_cell(info, None))
        ws.cell(next_row, 3, build_trade_cell(code, "electrical", info, None))
        ws.cell(next_row, 4, build_trade_cell(code, "plumbing", info, None))
        ws.cell(next_row, 5, build_trade_cell(code, "hvac", info, None))
        ws.cell(next_row, 6, build_local_cell(code, info, municipal, None))
        for col in range(1, 7):
            ws.cell(next_row, col).alignment = WRAP
            ws.cell(next_row, col).border = BORDER
        next_row += 1

    apply_sheet_filter(ws, header_row=3, max_col=6)


def export_workbook() -> dict[str, str]:
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE}")

    quick_md = QUICK_MD.read_text(encoding="utf-8")
    specialty_md = SPECIALTY_MD.read_text(encoding="utf-8")
    lookup = parse_jurisdiction_lookup(quick_md)
    municipal = parse_municipal_sections(quick_md)

    wb = openpyxl.load_workbook(TEMPLATE)
    states_ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb.active
    update_states_sheet(states_ws, lookup, municipal)

    write_municipal_sheet(wb, municipal)

    backflow = parse_specialty_table(
        specialty_md,
        "### Backflow — state verification routes",
        ("State/Terr.", "Primary authority", "Verification URL / method"),
    )
    fire = parse_specialty_table(
        specialty_md,
        "### Fire — state verification routes",
        ("State/Terr.", "Sprinkler / suppression", "Fire alarm / detection", "Verification URL"),
    )
    pest = parse_specialty_table(
        specialty_md,
        "### Pest control — state verification routes",
        ("State/Terr.", "Authority", "Verification URL"),
    )
    locksmith = parse_specialty_table(
        specialty_md,
        "### Locksmith / low voltage / alarm — state verification routes",
        ("State/Terr.", "Locksmith", "Alarm / security / CCTV / access", "Verification URL"),
    )

    specialty_counts = {
        "Backflow": write_specialty_sheet(
            wb,
            "Backflow",
            ["State / Territory", "Primary Authority", "Verification URL / Method"],
            backflow,
            [24, 32, 52],
            url_col_indexes=[3],
        ),
        "Fire Protection": write_specialty_sheet(
            wb,
            "Fire Protection",
            [
                "State / Territory",
                "Sprinkler / Suppression",
                "Fire Alarm / Detection",
                "Verification URL",
            ],
            fire,
            [24, 28, 28, 40],
            url_col_indexes=[4],
        ),
        "Pest Control": write_specialty_sheet(
            wb,
            "Pest Control",
            ["State / Territory", "Authority", "Verification URL"],
            pest,
            [24, 34, 48],
            url_col_indexes=[3],
        ),
        "Locksmith and Low Voltage": write_specialty_sheet(
            wb,
            "Locksmith and Low Voltage",
            [
                "State / Territory",
                "Locksmith",
                "Alarm / Security / CCTV / Access",
                "Verification URL",
            ],
            locksmith,
            [24, 22, 30, 40],
            url_col_indexes=[4],
        ),
    }

    for sheet_name, count in specialty_counts.items():
        if count < 40:
            raise RuntimeError(
                f"{sheet_name} sheet has only {count} rows — specialty markdown parse failed"
            )

    build_summary_sheet(wb)

    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_REPO)
    wb.save(OUT_ONEDRIVE)

    return {
        "repo_xlsx": str(OUT_REPO),
        "onedrive_xlsx": str(OUT_ONEDRIVE),
        "template": str(TEMPLATE),
        "jurisdictions": str(len(lookup)),
        "municipal_rows": str(len(municipal)),
        "specialty_rows": specialty_counts,
    }


def main() -> None:
    result = export_workbook()
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
