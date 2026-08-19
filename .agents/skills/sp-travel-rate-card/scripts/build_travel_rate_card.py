"""Build a Vixxo-branded SP travel rate-card workbook from a JSON config.

Usage:
  python .agents/skills/sp-travel-rate-card/scripts/build_travel_rate_card.py \\
    --config .agents/skills/sp-travel-rate-card/examples/ks69082-gray.json

  python .agents/skills/sp-travel-rate-card/scripts/build_travel_rate_card.py --self-test
"""
from __future__ import annotations

import argparse
import json
import sys
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path
from shutil import copyfile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pricing import (
    range_band,
    resolve_anchor_travel,
    suggested_travel,
    territory,
)

HERE = Path(__file__).resolve().parent
SKILL = HERE.parent
REPO = HERE.parents[3]
TEMPLATE = REPO / "assets" / "branded-documents" / "Excel Template.xlsx"
CACHE_PATH = REPO / ".tmp" / "sp-travel-rate-card" / "geocode-cache.json"
UA = "VixxoTravelRateCard/1.0 (work; crystal.gagner@vixxo.com)"

GREEN = "1B5E3B"
ROW_ALT = "EAF4EE"
WARN = "FFF4D6"
WHITE = "FFFFFF"
THIN = Border(
    left=Side(style="thin", color="C5D5CB"),
    right=Side(style="thin", color="C5D5CB"),
    top=Side(style="thin", color="C5D5CB"),
    bottom=Side(style="thin", color="C5D5CB"),
)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=GREEN)
SUB_FONT = Font(name="Calibri", size=11, color="333333")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
HEAD_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=10, color="222222")
BOLD_FONT = Font(name="Calibri", size=10, bold=True, color="222222")
HEAD_FILL = PatternFill("solid", fgColor=GREEN)
ALT_FILL = PatternFill("solid", fgColor=ROW_ALT)
WARN_FILL = PatternFill("solid", fgColor=WARN)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
USD = '"$"#,##0'
MILES_FMT = "0.0"
INT_FMT = "#,##0"
RATE_FMT = '"$"#,##0.00'


def load_cache() -> dict:
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    return {"geocode": {}, "routes": {}}


def save_cache(cache: dict) -> None:
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(cache, indent=2), encoding="utf-8")


def http_get(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=45) as resp:
        return json.loads(resp.read().decode("utf-8"))


def census_geocode(address: str) -> dict | None:
    params = urllib.parse.urlencode(
        {"address": address, "benchmark": "Public_AR_Current", "format": "json"}
    )
    url = f"https://geocoding.geo.census.gov/geocoder/locations/onelineaddress?{params}"
    data = http_get(url)
    matches = (data.get("result") or {}).get("addressMatches") or []
    if not matches:
        return None
    hit = matches[0]
    coords = hit["coordinates"]
    return {
        "lat": float(coords["y"]),
        "lon": float(coords["x"]),
        "display": hit.get("matchedAddress"),
        "source": "census",
    }


def nominatim_geocode(query: str) -> dict | None:
    params = urllib.parse.urlencode(
        {"q": query, "format": "json", "limit": 1, "countrycodes": "us"}
    )
    url = f"https://nominatim.openstreetmap.org/search?{params}"
    data = http_get(url)
    time.sleep(1.1)
    if not data:
        return None
    hit = data[0]
    return {
        "lat": float(hit["lat"]),
        "lon": float(hit["lon"]),
        "display": hit.get("display_name"),
        "source": "nominatim",
    }


def geocode(query: str, kind: str, cache: dict) -> dict:
    key = f"{kind}|{query}"
    if key in cache["geocode"]:
        return cache["geocode"][key]
    hit = census_geocode(query) if kind == "street" else None
    if not hit:
        hit = nominatim_geocode(query)
    if not hit:
        raise RuntimeError(f"Could not geocode {kind} query: {query}")
    hit["query"] = query
    cache["geocode"][key] = hit
    save_cache(cache)
    return hit


def osrm_route(origin: dict, dest: dict, cache: dict) -> dict:
    key = (
        f"{origin['lon']:.6f},{origin['lat']:.6f}|"
        f"{dest['lon']:.6f},{dest['lat']:.6f}"
    )
    if key in cache["routes"]:
        return cache["routes"][key]
    url = (
        "https://router.project-osrm.org/route/v1/driving/"
        f"{origin['lon']},{origin['lat']};{dest['lon']},{dest['lat']}"
        "?overview=false"
    )
    data = http_get(url)
    time.sleep(0.25)
    if data.get("code") != "Ok":
        raise RuntimeError(f"OSRM failed: {data.get('code')} {data.get('message')}")
    r = data["routes"][0]
    route = {
        "miles": round(r["distance"] / 1609.344, 1),
        "minutes": int(round(r["duration"] / 60)),
    }
    cache["routes"][key] = route
    save_cache(cache)
    return route


def apply_alias(name: str, aliases: dict[str, str]) -> str:
    stripped = name.strip()
    for src, dest in aliases.items():
        if stripped.lower() == src.lower():
            return dest
    return stripped


def extract_sites(sr_path: Path) -> list[dict]:
    from openpyxl import load_workbook as lw

    wb = lw(sr_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c or "").strip() for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    needed = ["Site Name", "Address", "City", "State", "Zip", "Site #", "Customer #"]
    missing = [c for c in needed if c not in idx]
    if missing:
        raise RuntimeError(f"SR export missing columns: {missing}")
    sites: dict[tuple, dict] = {}
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        key = (
            r[idx["Customer #"]],
            r[idx["Site #"]],
            r[idx["Address"]],
            r[idx["City"]],
            r[idx["State"]],
            r[idx["Zip"]],
        )
        if key not in sites:
            city = r[idx["City"]]
            state = r[idx["State"]]
            zipc = r[idx["Zip"]]
            addr = r[idx["Address"]]
            street = f"{addr}, {city}, {state} {zipc}"
            sites[key] = {
                "customer": r[idx["Customer #"]],
                "site_no": r[idx["Site #"]],
                "site_name": r[idx["Site Name"]],
                "city": city,
                "state": state,
                "address": street,
                "srs": 0,
            }
        sites[key]["srs"] += 1
    return list(sites.values())


def style_header_row(ws: Worksheet, row: int, cols: int) -> None:
    ws.row_dimensions[row].height = 28
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def write_table(
    ws: Worksheet,
    start_row: int,
    headers: list[str],
    rows: list[list[object]],
    money_cols: set[int] | None = None,
    miles_cols: set[int] | None = None,
    int_cols: set[int] | None = None,
    rate_cols: set[int] | None = None,
    center_cols: set[int] | None = None,
) -> int:
    money_cols = money_cols or set()
    miles_cols = miles_cols or set()
    int_cols = int_cols or set()
    rate_cols = rate_cols or set()
    center_cols = center_cols or set()
    for i, h in enumerate(headers, 1):
        ws.cell(start_row, i, h)
        ws.cell(start_row, i).font = HEAD_FONT
        ws.cell(start_row, i).fill = HEAD_FILL
        ws.cell(start_row, i).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.cell(start_row, i).border = THIN
    ws.row_dimensions[start_row].height = 28
    r = start_row + 1
    for idx, row in enumerate(rows):
        terr = None
        for h, v in zip(headers, row):
            if h == "Territory":
                terr = v
        fill = WARN_FILL if terr == "Extended" else (ALT_FILL if idx % 2 else None)
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val if val is not None else None)
            cell.font = BODY_FONT
            cell.border = THIN
            if c in center_cols:
                cell.alignment = CENTER
            elif c in money_cols | miles_cols | int_cols | rate_cols:
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
            if fill:
                cell.fill = fill
            if isinstance(val, (int, float)):
                if c in money_cols:
                    cell.number_format = USD
                elif c in rate_cols:
                    cell.number_format = RATE_FMT
                elif c in miles_cols:
                    cell.number_format = MILES_FMT
                elif c in int_cols:
                    cell.number_format = INT_FMT
        r += 1
    last = max(r - 1, start_row)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = f"A{start_row + 1}"
    return r


def autosize(ws: Worksheet, min_w: int = 11, max_w: int = 52) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def title_block(ws: Worksheet, title: str, subtitle: str, note: str) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT
    ws["A3"] = note
    ws["A3"].font = NOTE_FONT
    ws.merge_cells("A1:N1")
    ws.merge_cells("A2:N2")
    ws.merge_cells("A3:N3")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 36


def enrich_row(
    miles: float,
    minutes: int,
    cfg: dict,
    anchor_miles: float,
    anchor_travel: int,
    *,
    hold_anchor: bool = False,
) -> dict:
    local = cfg["local"]
    labor = int(cfg["extended"].get("labor_rate") or 0)
    terr = territory(miles, local["max_miles"])
    travel = suggested_travel(
        miles,
        local_max_miles=local["max_miles"],
        local_travel=int(local["travel"]),
        anchor_miles=anchor_miles,
        anchor_travel=anchor_travel,
        round_to=int(cfg["extended"].get("round_to", 5)),
        hold_anchor=hold_anchor,
    )
    return {
        "miles": miles,
        "round_trip": round(miles * 2, 1),
        "minutes": minutes,
        "band": range_band(miles, local.get("band_low_miles", 30), local["max_miles"]),
        "territory": terr,
        "travel": travel,
        "trip_min": travel + labor,
        "vs_anchor": travel - anchor_travel,
        "scaled_raw": round(anchor_travel * (miles / anchor_miles), 1) if anchor_miles else None,
        "implied_rate": round(travel / miles, 2) if miles else None,
    }


def is_anchor_name(name: str, anchor_city: str) -> bool:
    left = name.split(",")[0].strip().lower()
    right = anchor_city.split(",")[0].strip().lower()
    return left == right


def build_workbook(cfg: dict, origin: dict, city_rows_data: list[dict], site_rows_data: list[dict], out: Path) -> None:
    if not TEMPLATE.exists():
        raise RuntimeError(f"Branded template missing: {TEMPLATE}")
    copyfile(TEMPLATE, out)
    wb = load_workbook(out)
    ws = wb.active
    ws.title = "Assumptions"
    ext = cfg["extended"]
    local = cfg["local"]
    anchor_city = ext["anchor_city"]
    anchor_travel = cfg["_anchor_travel"]
    anchor_how = cfg["_anchor_how"]
    anchor_miles = cfg["_anchor_miles"]
    labor = ext.get("labor_rate")
    title = f"{cfg.get('sp_name') or 'Service provider'} ({cfg.get('sp_number') or 'SP'}) — travel rate card"
    subtitle = f"Origin: {cfg['origin']['address']}"
    note = (
        f"Local <= {local['max_miles']} miles = ${local['travel']}  |  "
        f"Extended scales from {anchor_city} at ${anchor_travel} for {anchor_miles} miles  |  "
        f"One-way OSRM driving miles. Suggested only — not an approved rate card."
    )
    title_block(ws, title, subtitle, note)
    ws["A5"] = "Item"
    ws["B5"] = "Value"
    style_header_row(ws, 5, 2)
    assumptions = [
        ["SP #", cfg.get("sp_number") or ""],
        ["SP name", cfg.get("sp_name") or ""],
        ["Origin", cfg["origin"]["address"]],
        ["Origin geocode", origin.get("display")],
        ["Local max miles", local["max_miles"]],
        ["Local travel", f"${local['travel']}"],
        ["Local band low (label only)", local.get("band_low_miles", 30)],
        ["Anchor city", anchor_city],
        ["Anchor travel", f"${anchor_travel} ({anchor_how})"],
        ["Anchor miles (one-way)", anchor_miles],
        ["Labor rate (trip min add-on)", f"${labor}" if labor is not None else ""],
        ["Round extended to", ext.get("round_to", 5)],
        ["Hold anchor exact", bool(ext.get("hold_anchor_exact", True))],
        ["Prepared", date.today().isoformat()],
    ]
    for i, (k, v) in enumerate(assumptions):
        ws.cell(6 + i, 1, k).font = BOLD_FONT
        ws.cell(6 + i, 1).border = THIN
        cell = ws.cell(6 + i, 2, v)
        cell.font = BODY_FONT
        cell.border = THIN
        if i % 2:
            ws.cell(6 + i, 1).fill = ALT_FILL
            cell.fill = ALT_FILL
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 88

    headers = [
        "City",
        "Parent CBSA",
        "Source",
        "One-way miles",
        "Round-trip miles",
        "Drive minutes",
        "Range band",
        "Territory",
        "Suggested flat travel",
        "Suggested trip min (travel + 1 hr labor)",
        f"Vs {anchor_city.split(',')[0]} travel",
        "Scaled raw (before round)",
        "Implied $/one-way mile",
        "Geocode",
    ]
    city_rows = []
    for c in city_rows_data:
        e = c["enrich"]
        city_rows.append(
            [
                c["name"],
                c.get("parent_cbsa") or "",
                c.get("source") or "config",
                e["miles"],
                e["round_trip"],
                e["minutes"],
                e["band"],
                e["territory"],
                e["travel"],
                e["trip_min"],
                e["vs_anchor"],
                e["scaled_raw"],
                e["implied_rate"],
                c.get("display") or "",
            ]
        )
    city_rows.sort(key=lambda r: (str(r[1]), r[3]))
    ws_c = wb.create_sheet("Cities")
    title_block(ws_c, title, "City / CBSA centroids from the coverage breakdown.", note)
    write_table(
        ws_c, 5, headers, city_rows,
        money_cols={9, 10, 11, 12}, miles_cols={4, 5}, int_cols={6},
        rate_cols={13}, center_cols={3, 7, 8},
    )
    autosize(ws_c)

    site_headers = [
        "Customer #", "Site #", "Site name", "City", "Site address", "SRs in export",
        "One-way miles", "Round-trip miles", "Drive minutes", "Range band", "Territory",
        "Suggested flat travel", "Suggested trip min", f"Vs {anchor_city.split(',')[0]} travel",
        "Geocode match", "Note",
    ]
    site_rows = []
    for s in site_rows_data:
        e = s["enrich"]
        site_rows.append(
            [
                s.get("customer"), s.get("site_no"), s.get("site_name"), s.get("city"),
                s.get("address"), s.get("srs"),
                e["miles"], e["round_trip"], e["minutes"], e["band"], e["territory"],
                e["travel"], e["trip_min"], e["vs_anchor"],
                s.get("display") or "", s.get("note") or "",
            ]
        )
    site_rows.sort(key=lambda r: (r[6] or 0, str(r[2])))
    ws_s = wb.create_sheet("Sites")
    title_block(ws_s, title, "Unique sites from the SR export (if supplied).", note)
    write_table(
        ws_s, 5, site_headers, site_rows,
        money_cols={12, 13, 14}, miles_cols={7, 8}, int_cols={6, 9},
        center_cols={10, 11},
    )
    autosize(ws_s)

    ext_headers = [
        "Location", "Type", "One-way miles", "Drive minutes", "Suggested flat travel",
        "Suggested trip min", "% of anchor miles", "Vs anchor travel", "Why this rate",
    ]
    ext_rows = []
    for c in city_rows_data:
        e = c["enrich"]
        if e["territory"] != "Extended":
            continue
        pct = e["miles"] / anchor_miles if anchor_miles else 0
        why = (
            f"{e['miles']} miles is {pct:.0%} of {anchor_city} ({anchor_miles} miles), "
            f"so travel scales from ${anchor_travel}."
        )
        if is_anchor_name(c["name"], anchor_city):
            why = f"Anchor. Travel held at ${anchor_travel} ({anchor_how})."
        ext_rows.append(
            [c["name"], "City", e["miles"], e["minutes"], e["travel"], e["trip_min"], pct, e["vs_anchor"], why]
        )
    for s in site_rows_data:
        e = s["enrich"]
        if e["territory"] != "Extended":
            continue
        pct = e["miles"] / anchor_miles if anchor_miles else 0
        loc = f"{s.get('site_name')} ({s.get('city')})"
        why = f"{e['miles']} miles is {pct:.0%} of {anchor_city}; travel scaled from ${anchor_travel}."
        if is_anchor_name(str(s.get("city") or ""), anchor_city) and ext.get("hold_anchor_exact", True):
            why = f"Anchor-city site. Travel held at ${anchor_travel}."
        ext_rows.append(
            [loc, "Site", e["miles"], e["minutes"], e["travel"], e["trip_min"], pct, e["vs_anchor"], why]
        )
    ext_rows.sort(key=lambda r: r[2])
    ws_e = wb.create_sheet("Extended analysis")
    title_block(ws_e, title, "Extended markets only (over local max miles).", note)
    write_table(
        ws_e, 5, ext_headers, ext_rows,
        money_cols={5, 6, 8}, miles_cols={3}, int_cols={4}, center_cols={2},
    )
    for row in ws_e.iter_rows(min_row=6, max_row=5 + len(ext_rows), min_col=7, max_col=7):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0%"
    rec_row = 7 + len(ext_rows)
    ws_e.cell(rec_row, 1, "Suggested city travel card (copy-ready)").font = BOLD_FONT
    rec = []
    for c in city_rows_data:
        e = c["enrich"]
        note_t = "Local truck roll"
        if e["territory"] == "Extended":
            note_t = f"Scaled from {anchor_city} ${anchor_travel}"
        if is_anchor_name(c["name"], anchor_city):
            note_t = f"Anchor — {anchor_how}"
        rec.append([c["name"], e["territory"], e["travel"], note_t])
    write_table(
        ws_e, rec_row + 1, ["City", "Territory", "Suggested flat travel", "Notes"], rec,
        money_cols={3}, center_cols={2},
    )
    autosize(ws_e)
    wb.save(out)


def run_self_test() -> int:
    travel, how = resolve_anchor_travel(
        {"trip_minimum": 350, "labor_rate": 75, "labor_hours_in_minimum": 1}
    )
    assert travel == 275, travel
    assert "trip_minimum" in how
    local = suggested_travel(
        34.5, local_max_miles=50, local_travel=75,
        anchor_miles=180.3, anchor_travel=275, round_to=5,
    )
    assert local == 75, local
    albany = suggested_travel(
        121.8, local_max_miles=50, local_travel=75,
        anchor_miles=180.3, anchor_travel=275, round_to=5,
    )
    assert albany == 185, albany
    sav = suggested_travel(
        180.3, local_max_miles=50, local_travel=75,
        anchor_miles=180.3, anchor_travel=275, round_to=5, hold_anchor=True,
    )
    assert sav == 275, sav
    print("self-test OK")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Build SP travel rate-card workbook")
    parser.add_argument("--config", help="Path to JSON config")
    parser.add_argument("--out", help="Output xlsx path")
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        return run_self_test()
    if not args.config:
        parser.error("--config is required unless --self-test")

    cfg = json.loads(Path(args.config).read_text(encoding="utf-8"))
    if "origin" not in cfg or not cfg["origin"].get("address"):
        raise SystemExit("config.origin.address is required")
    if "local" not in cfg or cfg["local"].get("travel") is None or cfg["local"].get("max_miles") is None:
        raise SystemExit("config.local.travel and config.local.max_miles are required")
    if "extended" not in cfg or not cfg["extended"].get("anchor_city"):
        raise SystemExit("config.extended.anchor_city is required")

    anchor_travel, anchor_how = resolve_anchor_travel(cfg["extended"])
    cfg["_anchor_travel"] = anchor_travel
    cfg["_anchor_how"] = anchor_how
    aliases = {str(k): str(v) for k, v in (cfg.get("aliases") or {}).items()}
    cache = load_cache()

    origin = geocode(cfg["origin"]["address"], "street", cache)
    cities_cfg = cfg.get("cities") or []
    if not cities_cfg:
        raise SystemExit("config.cities must list at least one city")

    city_hits = []
    for raw in cities_cfg:
        name = apply_alias(raw["name"] if isinstance(raw, dict) else str(raw), aliases)
        parent = (raw.get("parent_cbsa") if isinstance(raw, dict) else None) or ""
        source = (raw.get("source") if isinstance(raw, dict) else None) or "config"
        hit = geocode(name if "," in name else f"{name}, USA", "city", cache)
        route = osrm_route(origin, hit, cache)
        city_hits.append(
            {"name": name, "parent_cbsa": parent, "source": source, "display": hit.get("display"), "route": route, "hit": hit}
        )

    anchor_city = cfg["extended"]["anchor_city"]
    anchor_hit = next((c for c in city_hits if is_anchor_name(c["name"], anchor_city)), None)
    if not anchor_hit:
        hit = geocode(anchor_city, "city", cache)
        route = osrm_route(origin, hit, cache)
        anchor_hit = {"name": anchor_city, "route": route, "hit": hit, "display": hit.get("display")}
        city_hits.append({**anchor_hit, "parent_cbsa": "", "source": "anchor"})
    cfg["_anchor_miles"] = anchor_hit["route"]["miles"]

    city_rows_data = []
    for c in city_hits:
        hold = bool(cfg["extended"].get("hold_anchor_exact", True)) and is_anchor_name(
            c["name"], anchor_city
        )
        e = enrich_row(
            c["route"]["miles"], c["route"]["minutes"], cfg,
            cfg["_anchor_miles"], anchor_travel, hold_anchor=hold,
        )
        city_rows_data.append({**c, "enrich": e})

    site_rows_data = []
    sr_export = cfg.get("sr_export")
    if sr_export:
        sr_path = Path(sr_export)
        if not sr_path.exists():
            print(f"WARN: sr_export not found, skipping sites: {sr_path}", file=sys.stderr)
        else:
            for s in extract_sites(sr_path):
                try:
                    hit = geocode(s["address"], "street", cache)
                except RuntimeError as exc:
                    city_q = f"{s.get('city')}, {s.get('state') or ''}".strip().strip(",")
                    try:
                        hit = geocode(city_q, "city", cache)
                        s["note"] = f"Street geocode failed; used city centroid ({exc})"
                    except RuntimeError as exc2:
                        print(f"WARN skip site {s.get('site_name')}: {exc2}", file=sys.stderr)
                        continue
                route = osrm_route(origin, hit, cache)
                hold = bool(cfg["extended"].get("hold_anchor_exact", True)) and is_anchor_name(
                    str(s.get("city") or ""), anchor_city
                )
                e = enrich_row(
                    route["miles"], route["minutes"], cfg,
                    cfg["_anchor_miles"], anchor_travel, hold_anchor=hold,
                )
                site_rows_data.append({**s, "display": hit.get("display"), "enrich": e})

    sp = cfg.get("sp_number") or "SP"
    out = Path(args.out) if args.out else Path.home() / "Downloads" / (
        f"{sp}_Travel_Rate_Card_{date.today().isoformat()}.xlsx"
    )
    build_workbook(cfg, origin, city_rows_data, site_rows_data, out)
    print(f"Wrote {out}")
    print(f"Anchor: {anchor_city} {cfg['_anchor_miles']} miles -> ${anchor_travel} ({anchor_how})")
    print(f"Cities: {len(city_rows_data)}  Sites: {len(site_rows_data)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
