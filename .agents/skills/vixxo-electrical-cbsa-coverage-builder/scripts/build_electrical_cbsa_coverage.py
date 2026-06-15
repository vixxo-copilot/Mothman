from __future__ import annotations

import argparse
import csv
import io
import re
import unicodedata
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


TARGET_COUNTY_CBSA = {
    "adams": "Denver-Aurora-Lakewood, CO",
    "arapahoe": "Denver-Aurora-Lakewood, CO",
    "broomfield": "Denver-Aurora-Lakewood, CO",
    "clear creek": "Denver-Aurora-Lakewood, CO",
    "denver": "Denver-Aurora-Lakewood, CO",
    "douglas": "Denver-Aurora-Lakewood, CO",
    "elbert": "Denver-Aurora-Lakewood, CO",
    "gilpin": "Denver-Aurora-Lakewood, CO",
    "jefferson": "Denver-Aurora-Lakewood, CO",
    "park": "Denver-Aurora-Lakewood, CO",
    "boulder": "Boulder, CO",
    "el paso": "Colorado Springs, CO",
    "teller": "Colorado Springs, CO",
    "larimer": "Fort Collins, CO",
}


def clean_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def canonical(value) -> str:
    text = "".join(
        c
        for c in unicodedata.normalize("NFKD", str(value or ""))
        if not unicodedata.combining(c)
    )
    text = re.sub(r"&", " and ", text.lower())
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(
        r"\b(inc|incorporated|llc|ltd|corp|corporation|company|co|the)\b",
        " ",
        text,
    )
    return re.sub(r"\s+", " ", text).strip()


def token_set(value) -> set[str]:
    return {t for t in canonical(value).split() if len(t) > 1}


def norm_store(value) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"\s+", "", text)
    if not text:
        return ""
    if text.isdigit():
        return str(int(text))
    return text.lstrip("0") or text


def norm_zip(value) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[:5].zfill(5) if digits else ""


def norm_sp_num(value) -> str:
    text = str(value or "").strip()
    return re.sub(r"\.0$", "", text)


def num(value) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    return float(text) if text else 0.0


def col_index(header: list, *names: str) -> int | None:
    by_clean = {clean_header(h): i for i, h in enumerate(header) if h is not None}
    for name in names:
        if clean_header(name) in by_clean:
            return by_clean[clean_header(name)]
    return None


def headered_rows(wb, required: list[tuple[str, ...]], label: str, preferred: str | None = None):
    sheet_names = list(wb.sheetnames)
    if preferred and preferred in wb.sheetnames:
        sheet_names = [preferred] + [s for s in sheet_names if s != preferred]
    checked = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            header = list(row)
            if all(col_index(header, *aliases) is not None for aliases in required):
                return ws, header, ws.iter_rows(min_row=row_num + 1, values_only=True)
            if row_num <= 5:
                checked.append((sheet_name, row_num, header[:20]))
            if row_num >= 25:
                break
    raise SystemExit(f"{label} missing expected columns. checked={checked}")


def load_targets(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    _ws, header, rows = headered_rows(
        wb,
        [
            ("CBSA",),
            ("Service Provider",),
            ("Optimized Assigned Volume", "Optimized Annual Assigned Volume"),
        ],
        "coverage-change",
        preferred="Use for Coverage Change",
    )
    i_cbsa = col_index(header, "CBSA")
    i_name = col_index(header, "Service Provider")
    i_status = col_index(header, "Current or New SP to CBSA")
    i_hist = col_index(header, "Historical Requests")
    i_target = col_index(header, "Optimized Assigned Volume", "Optimized Annual Assigned Volume")
    i_change = col_index(header, "Change per CBSA")

    by_cbsa: dict[str, list[dict]] = defaultdict(list)
    all_rows = []
    for row in rows:
        if not row or not any(row):
            continue
        cbsa = str(row[i_cbsa] or "").strip()
        name = str(row[i_name] or "").strip()
        if not cbsa or not name:
            continue
        item = {
            "cbsa": cbsa,
            "name": name,
            "status": str(row[i_status] or "").strip() if i_status is not None else "",
            "historical": num(row[i_hist]) if i_hist is not None else 0.0,
            "target": num(row[i_target]),
            "change": num(row[i_change]) if i_change is not None else 0.0,
        }
        all_rows.append(item)
        if item["target"] > 0:
            by_cbsa[cbsa].append(item)
    for providers in by_cbsa.values():
        providers.sort(key=lambda p: (-p["target"], p["name"]))
    return {"by_cbsa": dict(by_cbsa), "all_rows": all_rows}


def load_los_sd(path: Path, area: str) -> dict[str, list[tuple[str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    _ws, header, rows = headered_rows(
        wb,
        [("NAME", "Customer #", "Customer Number"), ("LOS",), ("SHORT_DESCRIPTION", "SD")],
        "LOS/SD",
    )
    i_customer = col_index(header, "NAME", "Customer #", "Customer Number")
    i_los = col_index(header, "LOS")
    i_sd = col_index(header, "SHORT_DESCRIPTION", "SD")
    by_customer: dict[str, list[tuple[str, str]]] = defaultdict(list)
    seen: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for row in rows:
        if not row or not any(row):
            continue
        customer = str(row[i_customer] or "").strip()
        los = str(row[i_los] or "").strip()
        sd = str(row[i_sd] or "").strip() if i_sd is not None else ""
        if not customer or los.lower() != area.lower():
            continue
        key = (area, sd)
        if key not in seen[customer]:
            by_customer[customer].append(key)
            seen[customer].add(key)
    return dict(by_customer)


def decode_site_export(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-16", "utf-8-sig", "latin-1"):
        try:
            text = data.decode(encoding)
            if "\x00" not in text[:1000]:
                return text
        except UnicodeDecodeError:
            continue
    return data.decode("utf-16", errors="ignore")


def load_sites(
    path: Path,
    eligible_customers: set[str],
    allow_unlisted_customers: bool = False,
) -> tuple[list[dict], Counter]:
    reader = csv.DictReader(io.StringIO(decode_site_export(path)), delimiter="\t")
    stats = Counter()
    sites = []
    for row in reader:
        stats["site_rows"] += 1
        customer = str(row.get("Customer #") or "").strip()
        if not allow_unlisted_customers and customer not in eligible_customers:
            stats["skip_customer_not_area"] += 1
            continue
        if str(row.get("Status") or "").strip().lower() != "active":
            stats["skip_inactive"] += 1
            continue
        if str(row.get("Account Type") or "").strip().lower() not in {"", "customer site"}:
            stats["skip_not_customer_site"] += 1
            continue
        site_id = str(row.get("Site Id") or row.get("Site ID") or "").strip()
        if not site_id:
            stats["skip_no_site_id"] += 1
            continue
        site = {
            "customer": customer,
            "customer_name": str(row.get("Customer Name") or "").strip(),
            "site": norm_store(row.get("Site #")),
            "site_raw": str(row.get("Site #") or "").strip(),
            "site_id": site_id,
            "site_name": str(row.get("Site Name") or "").strip(),
            "city": str(row.get("City") or "").strip(),
            "state": str(row.get("State") or "").strip(),
            "zip": norm_zip(row.get("Zip Code")),
            "county": str(row.get("County") or "").strip(),
            "address": str(row.get("Street Address") or "").strip(),
        }
        sites.append(site)
        stats["sites_kept"] += 1
    return sites, stats


def load_work_history_geo(path: Path, area_name: str) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    _ws, header, rows = headered_rows(
        wb,
        [
            ("Parent Account Name",),
            ("Store #", "Store"),
            ("Postalcode", "Postal Code", "Zip Code"),
            ("CBSA",),
            ("Area",),
        ],
        "work-history",
        preferred="Work History",
    )
    i_parent = col_index(header, "Parent Account Name")
    i_store = col_index(header, "Store #", "Store")
    i_zip = col_index(header, "Postalcode", "Postal Code", "Zip Code")
    i_city = col_index(header, "City")
    i_state = col_index(header, "State")
    i_cbsa = col_index(header, "CBSA")
    i_area = col_index(header, "Area")
    i_category = col_index(header, "Category", "Short Description", "SHORT_DESCRIPTION", "SD")

    zip_cbsa_all: dict[str, Counter] = defaultdict(Counter)
    zip_cbsa_area: dict[str, Counter] = defaultdict(Counter)
    city_cbsa_all: dict[tuple[str, str], Counter] = defaultdict(Counter)
    site_cbsa_area: dict[tuple[str, str], Counter] = defaultdict(Counter)
    site_volume_area: Counter = Counter()
    site_sd_area: dict[tuple[str, str], Counter] = defaultdict(Counter)
    parent_sd_area: dict[str, Counter] = defaultdict(Counter)
    stats = Counter()

    for row in rows:
        if not row or not any(row):
            continue
        stats["work_rows"] += 1
        cbsa = str(row[i_cbsa] or "").strip()
        if not cbsa:
            continue
        zip_code = norm_zip(row[i_zip])
        city = canonical(row[i_city]) if i_city is not None else ""
        state = str(row[i_state] or "").strip().upper() if i_state is not None else ""
        parent = canonical(row[i_parent])
        store = norm_store(row[i_store])
        area = str(row[i_area] or "").strip().lower()
        if zip_code:
            zip_cbsa_all[zip_code][cbsa] += 1
        if city and state:
            city_cbsa_all[(city, state)][cbsa] += 1
        if area == area_name.lower():
            stats["area_work_rows"] += 1
            sd = str(row[i_category] or "").strip() if i_category is not None else ""
            sd = sd or area_name
            if parent:
                parent_sd_area[parent][sd] += 1
            if zip_code:
                zip_cbsa_area[zip_code][cbsa] += 1
            if parent and store:
                site_cbsa_area[(parent, store)][cbsa] += 1
                site_volume_area[(parent, store)] += 1
                site_sd_area[(parent, store)][sd] += 1
    return {
        "zip_cbsa_all": zip_cbsa_all,
        "zip_cbsa_area": zip_cbsa_area,
        "city_cbsa_all": city_cbsa_all,
        "site_cbsa_area": site_cbsa_area,
        "site_volume_area": site_volume_area,
        "site_sd_area": site_sd_area,
        "parent_sd_area": parent_sd_area,
        "stats": stats,
    }


def resolve_site_cbsa(site: dict, geo: dict, target_cbsas: set[str], area_name: str) -> tuple[str, str]:
    site_key = (canonical(site["customer_name"]), site["site"])
    area_label = canonical(area_name).replace(" ", "_") or "area"
    if site_key in geo["site_cbsa_area"]:
        return geo["site_cbsa_area"][site_key].most_common(1)[0][0], f"customer_store_{area_label}_history"
    if site["zip"] in geo["zip_cbsa_area"]:
        return geo["zip_cbsa_area"][site["zip"]].most_common(1)[0][0], f"zip_{area_label}_history"
    if site["zip"] in geo["zip_cbsa_all"]:
        return geo["zip_cbsa_all"][site["zip"]].most_common(1)[0][0], "zip_any_work_history"
    city_key = (canonical(site["city"]), site["state"].upper())
    if city_key in geo["city_cbsa_all"]:
        return geo["city_cbsa_all"][city_key].most_common(1)[0][0], "city_any_work_history"
    county_cbsa = TARGET_COUNTY_CBSA.get(canonical(site["county"]))
    if county_cbsa in target_cbsas:
        return county_cbsa, "target_county_fallback"
    return "", "unresolved"


def load_siebel(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    _ws, header, rows = headered_rows(
        wb,
        [("Service Contractor #",), ("Service Contractor Name",)],
        "Siebel",
    )
    i_num = col_index(header, "Service Contractor #")
    i_name = col_index(header, "Service Contractor Name")
    i_status = col_index(header, "Status")
    providers = []
    names_by_num: dict[str, Counter] = defaultdict(Counter)
    for row in rows:
        if not row or not any(row):
            continue
        sp_num = norm_sp_num(row[i_num])
        name = str(row[i_name] or "").strip()
        if not sp_num or not name:
            continue
        status = str(row[i_status] or "").strip() if i_status is not None else ""
        names_by_num[sp_num][name] += 1
        providers.append(
            {
                "num": sp_num,
                "name": name,
                "status": status,
                "canon": canonical(name),
                "tokens": token_set(name),
            }
        )
    display_by_num = {
        sp_num: sorted(names.items(), key=lambda item: (-item[1], -len(canonical(item[0])), item[0].lower()))[0][0]
        for sp_num, names in names_by_num.items()
    }
    for provider in providers:
        provider["display_name"] = display_by_num.get(provider["num"], provider["name"])
    return providers


def provider_score(query: str, provider: dict) -> float:
    q_canon = canonical(query)
    if not q_canon or not provider["canon"]:
        return 0.0
    if q_canon == provider["canon"]:
        return 1.0
    if q_canon in provider["canon"] or provider["canon"] in q_canon:
        return 0.92
    q_tokens = token_set(query)
    shared = len(q_tokens & provider["tokens"])
    token_score = shared / max(len(q_tokens | provider["tokens"]), 1)
    char_score = SequenceMatcher(None, q_canon, provider["canon"]).ratio()
    return max(token_score, char_score * 0.95)


def resolve_provider(name: str, providers: list[dict]) -> tuple[str, str, float]:
    best = ("", "", 0.0)
    for provider in providers:
        score = provider_score(name, provider)
        if score > best[2]:
            best = (provider["num"], provider.get("display_name") or provider["name"], score)
    return best if best[2] >= 0.70 else ("", "", best[2])


def provider_display(sp_num: str, sp_name: str) -> str:
    if sp_num and sp_name:
        return f"{sp_num} : {sp_name}"
    return sp_num or sp_name or ""


def rank_id(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text.split(":", 1)[0].strip() if ":" in text else text


def rank_name(value) -> str:
    text = str(value or "").strip()
    return text.split(":", 1)[1].strip() if ":" in text else ""


def compact_rank_values(values: list, limit: int) -> list[str]:
    ranks = []
    seen = set()
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        key = rank_id(text).lower()
        if key in seen:
            continue
        ranks.append(text)
        seen.add(key)
        if len(ranks) == limit:
            break
    return ranks + [""] * (limit - len(ranks))


def merge_rank_fields(
    rank_nums: list[str],
    rank_names: list[str],
    backup_rank_values: list[str],
    limit: int = 6,
) -> tuple[list[str], list[str], list[str]]:
    merged: list[tuple[str, str, str]] = []
    seen = set()

    for num, name in zip(rank_nums, rank_names):
        num = str(num or "").strip()
        name = str(name or "").strip()
        if not num or num.lower() in seen:
            continue
        merged.append((num, name, provider_display(num, name)))
        seen.add(num.lower())
        if len(merged) == limit:
            break

    for value in backup_rank_values:
        text = str(value or "").strip()
        num = rank_id(text)
        name = rank_name(text)
        if not num or num.lower() in seen:
            continue
        merged.append((num, name, provider_display(num, name) or text))
        seen.add(num.lower())
        if len(merged) == limit:
            break

    while len(merged) < limit:
        merged.append(("", "", ""))

    return [item[0] for item in merged], [item[1] for item in merged], [item[2] for item in merged]


def has_value(value) -> bool:
    return str(value or "").strip() != ""


def has_any_value(row: list) -> bool:
    return any(has_value(value) for value in row)


def coverage_review_key(customer: str, site_id: str, los: str, sd: str) -> tuple[str, str, str, str]:
    return (
        str(customer or "").strip().lower(),
        str(site_id or "").strip().lower(),
        str(los or "").strip().lower(),
        str(sd or "").strip().lower(),
    )


def coverage_review_key_label(key: tuple[str, str, str, str]) -> str:
    return " | ".join(key)


def combine_labels(*values: str) -> str:
    labels = []
    seen = set()
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        key = canonical(text)
        if key in seen:
            continue
        labels.append(text)
        seen.add(key)
    return " / ".join(labels)


def dedupe_resolved_provider_rows(rows: list[dict]) -> list[dict]:
    by_sp_num: dict[str, dict] = {}
    for row in rows:
        sp_num = norm_sp_num(row.get("sp_num"))
        if not sp_num:
            continue
        key = sp_num.lower()
        if key not in by_sp_num:
            merged = dict(row)
            merged["sp_num"] = sp_num
            by_sp_num[key] = merged
            continue
        merged = by_sp_num[key]
        merged["name"] = combine_labels(merged.get("name", ""), row.get("name", ""))
        merged["status"] = combine_labels(merged.get("status", ""), row.get("status", ""))
        merged["historical"] = num(merged.get("historical")) + num(row.get("historical"))
        merged["target"] = num(merged.get("target")) + num(row.get("target"))
        merged["change"] = num(merged.get("change")) + num(row.get("change"))
        merged["match_score"] = max(num(merged.get("match_score")), num(row.get("match_score")))
        merged["sp_name"] = merged.get("sp_name") or row.get("sp_name", "")
    return sorted(by_sp_num.values(), key=lambda p: (-num(p["target"]), p["sp_name"] or p["name"], p["sp_num"]))


def provider_summary_rows(provider_resolution: list[dict]) -> list[dict]:
    resolved_rows = []
    rows_by_cbsa: dict[str, list[dict]] = defaultdict(list)
    for row in provider_resolution:
        if row.get("sp_num"):
            rows_by_cbsa[row["cbsa"]].append(row)
    for rows in rows_by_cbsa.values():
        resolved_rows.extend(dedupe_resolved_provider_rows(rows))
    unresolved_rows = [row for row in provider_resolution if not row.get("sp_num")]
    return sorted(resolved_rows + unresolved_rows, key=lambda p: (p["cbsa"], p.get("sp_name") or p["name"]))


def resolve_target_providers(targets: dict, providers: list[dict]) -> tuple[dict, list[dict]]:
    resolved: dict[str, list[dict]] = {}
    resolution_rows = []
    for cbsa, rows in targets["by_cbsa"].items():
        cbsa_resolved_rows = []
        for item in rows:
            sp_num, sp_name, score = resolve_provider(item["name"], providers)
            row = dict(item)
            row.update({"sp_num": norm_sp_num(sp_num), "sp_name": sp_name, "match_score": score})
            resolution_rows.append(row)
            if sp_num:
                cbsa_resolved_rows.append(row)
        resolved[cbsa] = dedupe_resolved_provider_rows(cbsa_resolved_rows)
    return resolved, resolution_rows


def work_history_los_sd(site: dict, geo: dict, area_name: str) -> list[tuple[str, str]]:
    site_key = (canonical(site["customer_name"]), site["site"])
    parent_key = canonical(site["customer_name"])
    sd_counts = geo["site_sd_area"].get(site_key) or geo["parent_sd_area"].get(parent_key)
    if not sd_counts:
        return []
    return [(area_name, sd) for sd, _count in sd_counts.most_common()]


def load_existing_coverage(path: Path | None, area_name: str) -> tuple[list[dict], Counter]:
    stats = Counter()
    if not path:
        return [], stats
    wb = load_workbook(path, read_only=True, data_only=True)
    _ws, header, rows = headered_rows(
        wb,
        [("Customer Number", "Customer #"), ("Site ID", "Site Id"), ("LOS",)],
        "existing coverage",
        preferred="Customer SC Relationships",
    )
    i_customer = col_index(header, "Customer Number", "Customer #")
    i_site = col_index(header, "Site Number", "Site #")
    i_site_id = col_index(header, "Site ID", "Site Id")
    i_address = col_index(header, "Street Address", "Address")
    i_city = col_index(header, "Site City", "City")
    i_state = col_index(header, "Site State", "State")
    i_postal = col_index(header, "Site Postal Code", "Zip Code", "Postal Code")
    i_county = col_index(header, "Site County", "County")
    i_zone = col_index(header, "Zone")
    i_market = col_index(header, "Market")
    i_cbsa = col_index(header, "CBSA")
    i_los = col_index(header, "LOS")
    i_sd = col_index(header, "Short Description", "SHORT_DESCRIPTION", "SD")
    rank_indexes = []
    for idx, value in enumerate(header):
        match = re.fullmatch(r"rank(\d+)", clean_header(value))
        if match:
            rank_indexes.append((int(match.group(1)), idx))
    rank_indexes.sort()

    coverage = []
    for row in rows:
        if not row or not any(row):
            continue
        stats["existing_rows_scanned"] += 1
        los = str(row[i_los] or "").strip()
        if los.lower() != area_name.lower():
            continue
        ranks = [str(row[idx] or "").strip() for _rank, idx in rank_indexes]
        customer = str(row[i_customer] or "").strip()
        site_id = str(row[i_site_id] or "").strip()
        if not customer or not site_id:
            stats["existing_area_rows_skipped_missing_key"] += 1
            continue
        if not any(ranks):
            stats["existing_area_rows_skipped_empty_ranks"] += 1
            continue
        coverage.append(
            {
                "customer": customer,
                "site": str(row[i_site] or "").strip() if i_site is not None else "",
                "site_id": site_id,
                "address": str(row[i_address] or "").strip() if i_address is not None else "",
                "city": str(row[i_city] or "").strip() if i_city is not None else "",
                "state": str(row[i_state] or "").strip() if i_state is not None else "",
                "postal": str(row[i_postal] or "").strip() if i_postal is not None else "",
                "county": str(row[i_county] or "").strip() if i_county is not None else "",
                "zone": str(row[i_zone] or "").strip() if i_zone is not None else "",
                "market": str(row[i_market] or "").strip() if i_market is not None else "",
                "cbsa": str(row[i_cbsa] or "").strip() if i_cbsa is not None else "",
                "los": los,
                "sd": str(row[i_sd] or "").strip() if i_sd is not None else "",
                "ranks": ranks,
            }
        )
        stats["existing_area_rows"] += 1
    return coverage, stats


def assign_sites(
    sites: list[dict],
    resolved_targets: dict,
    geo: dict,
    area_name: str,
    los_sd: dict[str, list[tuple[str, str]]],
    use_work_history_los_sd: bool = False,
    default_area_sd: str = "",
) -> tuple[list[dict], list[dict], Counter]:
    target_cbsas = set(resolved_targets)
    for site in sites:
        cbsa, source = resolve_site_cbsa(site, geo, target_cbsas, area_name)
        site["cbsa"] = cbsa
        site["cbsa_source"] = source
        site_key = (canonical(site["customer_name"]), site["site"])
        site["estimated_volume"] = max(1, geo["site_volume_area"].get(site_key, 0))
        site["los_sd"] = los_sd.get(site["customer"], [])
        site["los_sd_source"] = "active_los_sd" if site["los_sd"] else ""
        if not site["los_sd"] and use_work_history_los_sd:
            site["los_sd"] = work_history_los_sd(site, geo, area_name)
            if site["los_sd"]:
                site["los_sd_source"] = "work_history_category"
        if not site["los_sd"] and default_area_sd:
            site["los_sd"] = [(area_name, default_area_sd)]
            site["los_sd_source"] = "default_sd"

    assignments = []
    exceptions = []
    stats = Counter()
    by_cbsa: dict[str, list[dict]] = defaultdict(list)
    for site in sites:
        if not site["los_sd"]:
            exceptions.append({**site, "reason": f"No active or work-history {area_name} LOS/SD"})
            stats["exception_no_los_sd"] += 1
            continue
        if not site["cbsa"]:
            exceptions.append({**site, "reason": "CBSA unresolved"})
            stats["exception_unresolved_cbsa"] += 1
            continue
        if site["cbsa"] not in resolved_targets:
            exceptions.append({**site, "reason": "No recommended target providers for CBSA"})
            stats["exception_no_target_cbsa"] += 1
            continue
        if not resolved_targets[site["cbsa"]]:
            exceptions.append({**site, "reason": "Target providers did not resolve to Siebel"})
            stats["exception_no_resolved_provider"] += 1
            continue
        by_cbsa[site["cbsa"]].append(site)

    for cbsa, cbsa_sites in sorted(by_cbsa.items()):
        providers = resolved_targets[cbsa]
        assigned_weight = Counter()
        ordered_sites = sorted(
            cbsa_sites,
            key=lambda s: (-s["estimated_volume"], s["customer"], s["city"], s["site_id"]),
        )
        for site in ordered_sites:
            r1 = min(
                providers,
                key=lambda p: (
                    assigned_weight[p["sp_num"]] / max(p["target"], 1),
                    -p["target"],
                    p["sp_num"],
                ),
            )
            assigned_weight[r1["sp_num"]] += site["estimated_volume"]
            ranks = [r1] + [p for p in providers if p["sp_num"] != r1["sp_num"]]
            ranks = ranks[:6]
            assignments.append({**site, "ranks": ranks})
            stats["covered_sites"] += 1
    return assignments, exceptions, stats


def current_placeholder_from_proposed(proposed_row: list) -> list:
    return [
        proposed_row[0],
        proposed_row[2],
        proposed_row[3],
        "",
        proposed_row[4],
        proposed_row[5],
        proposed_row[6],
        proposed_row[7],
        "",
        "",
        proposed_row[8],
        proposed_row[12],
        proposed_row[13],
        *([""] * 9),
    ]


def proposed_placeholder_from_current(current_row: list) -> list:
    return compact_rank_values(current_row[13:22], 6)


def proposed_rank_displays(proposed_row: list) -> list[str]:
    return compact_rank_values([proposed_row[idx] for idx in (16, 19, 22, 25, 28, 31)], 6)


def write_workbook(
    path: Path,
    assignments: list[dict],
    exceptions: list[dict],
    los_sd: dict[str, list[tuple[str, str]]],
    provider_resolution: list[dict],
    existing_coverage: list[dict] | None = None,
) -> Counter:
    wb = Workbook()
    ws = wb.active
    ws.title = "SP Numbers"
    ws.append(["Customer #", "Site ID", "LOS", "SD", "R1", "R2", "R3", "R4", "R5", "R6"])

    review = wb.create_sheet("account review")
    review.append(
        [
            "Customer #",
            "Customer Name",
            "Site #",
            "Site ID",
            "City",
            "State",
            "Zip",
            "County",
            "CBSA",
            "CBSA Source",
            "Estimated Volume",
            "LOS/SD Source",
            "LOS",
            "SD",
            "R1 SP #",
            "R1 SP Name",
            "R1 SP # and Name",
            "R2 SP #",
            "R2 SP Name",
            "R2 SP # and Name",
            "R3 SP #",
            "R3 SP Name",
            "R3 SP # and Name",
            "R4 SP #",
            "R4 SP Name",
            "R4 SP # and Name",
            "R5 SP #",
            "R5 SP Name",
            "R5 SP # and Name",
            "R6 SP #",
            "R6 SP Name",
            "R6 SP # and Name",
        ]
    )

    coverage_summary_ws = wb.create_sheet("new coverage summary")
    coverage_summary_ws.append(
        [
            "CBSA",
            "Recommended SP #",
            "Recommended SP Name",
            "Recommended SP # and Name",
            "Current/New",
            "Historical Requests",
            "Optimized Assigned Volume",
            "Change",
            "R1 Site Count",
            "R1 Estimated Volume",
            "R1 Coverage Rows",
            "R2-R6 Site Appearances",
            "Any-Rank Site Appearances",
            "Recommendation Used In New Coverage",
        ]
    )

    assignment_ws = wb.create_sheet("site assignment review")
    assignment_ws.append(
        [
            "Customer #",
            "Customer Name",
            "Site #",
            "Site ID",
            "City",
            "State",
            "Zip",
            "County",
            "CBSA",
            "CBSA Source",
            "Estimated Volume",
            "LOS/SD Source",
            "R1 SP #",
            "R1 SP Name",
            "R2 SP #",
            "R2 SP Name",
            "R3 SP #",
            "R3 SP Name",
            "R4 SP #",
            "R4 SP Name",
            "R5 SP #",
            "R5 SP Name",
            "R6 SP #",
            "R6 SP Name",
        ]
    )

    exceptions_ws = wb.create_sheet("site exceptions")
    exceptions_ws.append(
        [
            "Customer #",
            "Customer Name",
            "Site #",
            "Site ID",
            "City",
            "State",
            "Zip",
            "County",
            "CBSA",
            "CBSA Source",
            "LOS/SD Source",
            "Reason",
        ]
    )

    provider_ws = wb.create_sheet("provider resolution")
    provider_ws.append(
        [
            "CBSA",
            "Recommended Provider",
            "Current/New",
            "Historical Requests",
            "Optimized Assigned Volume",
            "Change",
            "Resolved SP #",
            "Resolved SP Name",
            "Match Score",
        ]
    )

    existing_ws = None
    coverage_review_ws = None
    if existing_coverage is not None:
        coverage_review_ws = wb.create_sheet("coverage review")
        coverage_review_ws.append(
            [
                "Review Status",
                "Match Key",
                "Current Customer #",
                "Current Site #",
                "Current Site ID",
                "Current Street Address",
                "Current City",
                "Current State",
                "Current Postal Code",
                "Current County",
                "Current Zone",
                "Current Market",
                "Current CBSA",
                "Current LOS",
                "Current SD",
                "Current Rank 1",
                "Current Rank 2",
                "Current Rank 3",
                "Current Rank 4",
                "Current Rank 5",
                "Current Rank 6",
                "Current Rank 7",
                "Current Rank 8",
                "Current Rank 9",
                "Proposed R1 SP # and Name",
                "Proposed R2 SP # and Name",
                "Proposed R3 SP # and Name",
                "Proposed R4 SP # and Name",
                "Proposed R5 SP # and Name",
                "Proposed R6 SP # and Name",
            ]
        )
        existing_ws = wb.create_sheet("existing coverage")
        existing_ws.append(
            [
                "Customer #",
                "Site #",
                "Site ID",
                "Street Address",
                "City",
                "State",
                "Postal Code",
                "County",
                "Zone",
                "Market",
                "CBSA",
                "LOS",
                "SD",
                "Rank 1",
                "Rank 2",
                "Rank 3",
                "Rank 4",
                "Rank 5",
                "Rank 6",
                "Rank 7",
                "Rank 8",
                "Rank 9",
            ]
        )

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    header_sheets = [ws, review, coverage_summary_ws, assignment_ws, exceptions_ws, provider_ws]
    if coverage_review_ws is not None:
        header_sheets.append(coverage_review_ws)
    if existing_ws is not None:
        header_sheets.append(existing_ws)
    for sheet in header_sheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"

    stats = Counter()
    r1_site_counts = Counter()
    r1_volume = Counter()
    r1_coverage_rows = Counter()
    rank_site_counts = Counter()
    backup_rank_site_counts = Counter()
    proposed_review_rows: dict[tuple[str, str, str, str], list[list]] = defaultdict(list)
    existing_rank_lookup: dict[tuple[str, str, str, str], list[str]] = defaultdict(list)
    sp_number_keys = set()
    for item in existing_coverage or []:
        key = coverage_review_key(item["customer"], item["site_id"], item["los"], item["sd"])
        for value in compact_rank_values(item["ranks"], 9):
            if value and rank_id(value).lower() not in {rank_id(v).lower() for v in existing_rank_lookup[key]}:
                existing_rank_lookup[key].append(value)

    for assignment in sorted(assignments, key=lambda s: (s["customer"], s["site_id"], s["cbsa"])):
        ranks = assignment["ranks"]
        rank_nums = [p["sp_num"] for p in ranks] + [""] * (6 - len(ranks))
        rank_names = [p["sp_name"] for p in ranks] + [""] * (6 - len(ranks))
        if not all(has_value(assignment.get(field)) for field in ("customer", "site_id", "state", "cbsa")):
            stats["proposed_rows_skipped_missing_context"] += 1
            continue
        if not rank_nums[0]:
            stats["proposed_rows_skipped_empty_ranks"] += 1
            continue
        rank_displays = [provider_display(num, name) for num, name in zip(rank_nums, rank_names)]
        los_sd_rows = assignment.get("los_sd", los_sd.get(assignment["customer"], []))
        r1_key = (assignment["cbsa"], rank_nums[0])
        r1_site_counts[r1_key] += 1
        r1_volume[r1_key] += assignment["estimated_volume"]
        r1_coverage_rows[r1_key] += len(los_sd_rows)
        for idx, sp_num in enumerate(rank_nums):
            if not sp_num:
                continue
            rank_key = (assignment["cbsa"], sp_num)
            rank_site_counts[rank_key] += 1
            if idx > 0:
                backup_rank_site_counts[rank_key] += 1
        assignment_row = [
            assignment["customer"],
            assignment["customer_name"],
            assignment["site_raw"],
            assignment["site_id"],
            assignment["city"],
            assignment["state"],
            assignment["zip"],
            assignment["county"],
            assignment["cbsa"],
            assignment["cbsa_source"],
            assignment["estimated_volume"],
            assignment["los_sd_source"],
        ]
        assignment_ws.append(assignment_row + [value for pair in zip(rank_nums, rank_names) for value in pair])
        stats["assignment_sites_written"] += 1
        stats[f"r1_{rank_nums[0]}"] += 1
        for los, sd in los_sd_rows:
            review_key = coverage_review_key(assignment["customer"], assignment["site_id"], los, sd)
            final_rank_nums, final_rank_names, final_rank_displays = merge_rank_fields(
                rank_nums,
                rank_names,
                existing_rank_lookup.get(review_key, []),
            )
            if not all(has_value(value) for value in (assignment["customer"], assignment["site_id"], los, sd, final_rank_nums[0])):
                stats["coverage_rows_skipped_missing_required_fields"] += 1
                continue
            ws.append([assignment["customer"], assignment["site_id"], los, sd, *final_rank_nums])
            sp_number_keys.add(review_key)
            review_ws_row = assignment_row + [los, sd]
            proposed_row = review_ws_row + [
                value for group in zip(final_rank_nums, final_rank_names, final_rank_displays) for value in group
            ]
            review.append(proposed_row)
            proposed_review_rows[review_key].append(proposed_row)
            stats["coverage_rows_written"] += 1

    for exception in sorted(exceptions, key=lambda s: (s["customer"], s["site_id"])):
        exceptions_ws.append(
            [
                exception["customer"],
                exception["customer_name"],
                exception["site_raw"],
                exception["site_id"],
                exception["city"],
                exception["state"],
                exception["zip"],
                exception["county"],
                exception.get("cbsa", ""),
                exception.get("cbsa_source", ""),
                exception.get("los_sd_source", ""),
                exception["reason"],
            ]
        )
        stats["exception_sites_written"] += 1

    for item in provider_resolution:
        provider_ws.append(
            [
                item["cbsa"],
                item["name"],
                item["status"],
                item["historical"],
                item["target"],
                item["change"],
                item["sp_num"],
                item["sp_name"],
                round(item["match_score"], 3),
            ]
        )
        if not item["sp_num"]:
            stats["unresolved_target_providers"] += 1

    for item in provider_summary_rows(provider_resolution):
        key = (item["cbsa"], item["sp_num"])
        coverage_summary_ws.append(
            [
                item["cbsa"],
                item["sp_num"],
                item["sp_name"],
                provider_display(item["sp_num"], item["sp_name"]),
                item["status"],
                item["historical"],
                item["target"],
                item["change"],
                r1_site_counts[key],
                r1_volume[key],
                r1_coverage_rows[key],
                backup_rank_site_counts[key],
                rank_site_counts[key],
                "Yes" if rank_site_counts[key] else "No",
            ]
        )

    if existing_ws is not None:
        existing_review_rows: dict[tuple[str, str, str, str], list[list]] = defaultdict(list)
        for item in sorted(existing_coverage or [], key=lambda r: (r["customer"], r["site_id"], r["los"], r["sd"])):
            ranks = item["ranks"] + [""] * (9 - len(item["ranks"]))
            existing_row = [
                item["customer"],
                item["site"],
                item["site_id"],
                item["address"],
                item["city"],
                item["state"],
                item["postal"],
                item["county"],
                item["zone"],
                item["market"],
                item["cbsa"],
                item["los"],
                item["sd"],
                *ranks[:9],
            ]
            existing_ws.append(existing_row)
            existing_review_rows[coverage_review_key(item["customer"], item["site_id"], item["los"], item["sd"])].append(existing_row)
            stats["existing_coverage_rows_written"] += 1

        if coverage_review_ws is not None:
            for key in sorted(set(existing_review_rows) | set(proposed_review_rows)):
                current_rows = existing_review_rows.get(key, [])
                proposed_rows = proposed_review_rows.get(key, [])
                row_count = max(len(current_rows), len(proposed_rows))
                for idx in range(row_count):
                    current_present = idx < len(current_rows)
                    proposed_present = idx < len(proposed_rows)
                    current_row = current_rows[idx] if current_present else current_placeholder_from_proposed(proposed_rows[idx])
                    proposed_ranks = (
                        proposed_rank_displays(proposed_rows[idx])
                        if proposed_present
                        else proposed_placeholder_from_current(current_rows[idx])
                    )
                    if not has_any_value(current_row) and not has_any_value(proposed_ranks):
                        stats["coverage_review_rows_skipped_empty"] += 1
                        continue
                    if current_present and proposed_present:
                        status = "matched_current_and_proposed"
                        current_rank_ids = [rank_id(value) for value in compact_rank_values(current_row[13:22], 6)]
                        proposed_rank_ids = [rank_id(value) for value in compact_rank_values(proposed_ranks, 6)]
                        if current_rank_ids == proposed_rank_ids:
                            stats["matched_rows_same_ranks"] += 1
                        else:
                            stats["matched_rows_changed_ranks"] += 1
                    elif current_present:
                        status = "current_only_preserved"
                        key_required = (current_row[0], current_row[2], current_row[11], current_row[12])
                        current_rank_ids = [rank_id(value) for value in compact_rank_values(current_row[13:22], 6)]
                        if all(has_value(value) for value in (*key_required, current_rank_ids[0])):
                            load_key = coverage_review_key(current_row[0], current_row[2], current_row[11], current_row[12])
                            if load_key not in sp_number_keys:
                                ws.append([current_row[0], current_row[2], current_row[11], current_row[12], *current_rank_ids])
                                sp_number_keys.add(load_key)
                                stats["current_only_rows_preserved_in_sp_numbers"] += 1
                        else:
                            stats["current_only_rows_skipped_missing_required_fields"] += 1
                    else:
                        status = "proposed_only"
                    stats[f"coverage_review_status_{status}"] += 1
                    coverage_review_ws.append([status, coverage_review_key_label(key), *current_row, *proposed_ranks])
                    stats["coverage_review_rows_written"] += 1

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 42)

    wb.save(path)
    return stats


def write_summary(
    path: Path,
    args,
    targets: dict,
    los_sd: dict,
    site_stats: Counter,
    geo_stats: Counter,
    provider_resolution: list[dict],
    assignments: list[dict],
    exceptions: list[dict],
    output_stats: Counter,
    existing_coverage_stats: Counter | None = None,
) -> None:
    area = args.area.strip()
    by_cbsa = Counter(a["cbsa"] for a in assignments)
    by_los_sd_source = Counter(a.get("los_sd_source", "") for a in assignments)
    r1_by_cbsa_sp = Counter((a["cbsa"], a["ranks"][0]["sp_num"], a["ranks"][0]["sp_name"]) for a in assignments)
    exception_reasons = Counter(e["reason"] for e in exceptions)
    lines = [
        f"# {area} CBSA Coverage Build Summary",
        "",
        f"Output workbook: `{Path(args.output).name}`",
        "",
        "## Inputs",
        f"- Coverage change workbook: `{Path(args.coverage_change).name}`",
        f"- Site export: `{Path(args.sites).name}`",
        f"- Active LOS/SD export: `{Path(args.los_sd).name}`",
        f"- Active Siebel SP export: `{Path(args.siebel).name}`",
        f"- Existing customer-SC relationship export: `{Path(args.existing_coverage).name if args.existing_coverage else 'none'}`",
        "",
        "## Build Rules",
        "- Source of recommended provider mix is the workbook sheet containing `CBSA`, `Service Provider`, and optimized assigned volume.",
        "- Only providers with optimized assigned volume greater than zero are used for coverage ranks.",
        f"- Active sites are included when they have active `{area}` LOS/SD rows, allowed work-history `{area}` evidence, or a configured default SD.",
        f"- Work-history categories are used as `{area}` SDs when active LOS/SD is missing: {args.use_work_history_los_sd}.",
        f"- Default `{area}` SD when active/work-history SD is missing: `{args.default_area_sd or 'none'}`.",
        f"- Site CBSA is resolved from {area} work-history ZIP first, then any work-history ZIP, city history, and finally a limited target-county fallback.",
        "- R1 is balanced within each CBSA by assigning sites to the provider with the lowest assigned-volume-to-target-volume ratio.",
        "- Current ranks are used to fill remaining proposed rank slots after recommended providers, with duplicates removed and rank gaps compacted left.",
        "- Current-only coverage rows are carried forward into `SP Numbers` so a load does not remove existing coverage where no new recommendation exists.",
        f"- One load row is emitted for each eligible site and each active or work-history-derived {area} SD for that customer.",
        "",
        "## Result",
        f"- Target CBSAs with recommended providers: {len(targets['by_cbsa']):,}",
        f"- Customers with active {area} LOS/SD: {len(los_sd):,}",
        f"- Site rows scanned: {site_stats['site_rows']:,}",
        f"- Site rows skipped because customer lacks active {area} LOS/SD: {site_stats['skip_customer_not_area']:,}",
        f"- Active {area} sites kept for consideration: {site_stats['sites_kept']:,}",
        f"- Work-history rows scanned: {geo_stats['work_rows']:,}",
        f"- {area} work-history rows scanned: {geo_stats['area_work_rows']:,}",
        f"- Covered sites: {len(assignments):,}",
        f"- Exception sites: {len(exceptions):,}",
        f"- Coverage rows written: {output_stats['coverage_rows_written']:,}",
        f"- Side-by-side coverage review rows written: {output_stats['coverage_review_rows_written']:,}",
        f"- Coverage review matched current/proposed rows: {output_stats['coverage_review_status_matched_current_and_proposed']:,}",
        f"- Coverage review matched rows with unchanged R1-R6 order: {output_stats['matched_rows_same_ranks']:,}",
        f"- Coverage review matched rows with changed R1-R6 order: {output_stats['matched_rows_changed_ranks']:,}",
        f"- Coverage review current-only preserved rows: {output_stats['coverage_review_status_current_only_preserved']:,}",
        f"- Coverage review proposed-only rows: {output_stats['coverage_review_status_proposed_only']:,}",
        f"- Provider target rows unresolved to Siebel: {output_stats['unresolved_target_providers']:,}",
        f"- Existing {area} coverage rows imported: {(existing_coverage_stats or Counter())['existing_area_rows']:,}",
        f"- Existing {area} rows skipped for empty ranks: {(existing_coverage_stats or Counter())['existing_area_rows_skipped_empty_ranks']:,}",
        f"- Existing {area} rows skipped for missing customer/site key: {(existing_coverage_stats or Counter())['existing_area_rows_skipped_missing_key']:,}",
        f"- Current-only rows preserved in `SP Numbers`: {output_stats['current_only_rows_preserved_in_sp_numbers']:,}",
        f"- Current-only rows skipped for missing required fields: {output_stats['current_only_rows_skipped_missing_required_fields']:,}",
        f"- Proposed rows skipped for missing required fields: {output_stats['coverage_rows_skipped_missing_required_fields']:,}",
        f"- Proposed sites skipped for missing context: {output_stats['proposed_rows_skipped_missing_context']:,}",
        f"- Proposed sites skipped for empty ranks: {output_stats['proposed_rows_skipped_empty_ranks']:,}",
        "",
        "## Covered Sites By CBSA",
    ]
    for cbsa, count in by_cbsa.most_common():
        lines.append(f"- `{cbsa}`: {count:,}")
    lines.extend(["", "## Covered Sites By LOS/SD Source"])
    for source, count in by_los_sd_source.most_common():
        lines.append(f"- `{source or 'unknown'}`: {count:,}")
    lines.extend(["", "## R1 Assignment Counts"])
    for (cbsa, sp_num, sp_name), count in sorted(r1_by_cbsa_sp.items()):
        lines.append(f"- `{cbsa}` / `{sp_num}` {sp_name}: {count:,} site(s)")
    lines.extend(["", "## Exceptions"])
    for reason, count in exception_reasons.most_common():
        lines.append(f"- {reason}: {count:,}")
    unresolved = [p for p in provider_resolution if not p["sp_num"]]
    if unresolved:
        lines.extend(["", "## Unresolved Target Providers"])
        for item in unresolved:
            lines.append(f"- `{item['cbsa']}` / `{item['name']}` (best score {item['match_score']:.3f})")
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build CBSA coverage from RFP target-volume recommendations.")
    parser.add_argument("--coverage-change", required=True, help="Workbook containing Use for Coverage Change and Work History tabs")
    parser.add_argument("--sites", required=True, help="VixxoLink site export CSV/TSV")
    parser.add_argument("--los-sd", required=True, help="Active LOS/SD export")
    parser.add_argument("--siebel", required=True, help="Active Siebel service provider export")
    parser.add_argument("--area", default="Electrical", help="LOS/Work History area to build coverage for")
    parser.add_argument(
        "--use-work-history-los-sd",
        action="store_true",
        help="Use area work-history categories as LOS/SD when active LOS/SD export lacks the area",
    )
    parser.add_argument(
        "--default-area-sd",
        default="",
        help="Default SD to use when active LOS/SD and work-history categories are unavailable",
    )
    parser.add_argument(
        "--existing-coverage",
        default="",
        help="Optional customer-SC relationship workbook to add as an existing coverage tab",
    )
    parser.add_argument("--output", required=True, help="Output coverage workbook")
    parser.add_argument("--summary", required=True, help="Markdown build summary")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    targets = load_targets(Path(args.coverage_change))
    area = args.area.strip()
    los_sd = load_los_sd(Path(args.los_sd), area)
    sites, site_stats = load_sites(
        Path(args.sites),
        set(los_sd),
        allow_unlisted_customers=args.use_work_history_los_sd,
    )
    geo = load_work_history_geo(Path(args.coverage_change), area)
    providers = load_siebel(Path(args.siebel))
    resolved_targets, provider_resolution = resolve_target_providers(targets, providers)
    existing_coverage, existing_coverage_stats = load_existing_coverage(
        Path(args.existing_coverage) if args.existing_coverage else None,
        area,
    )
    assignments, exceptions, assignment_stats = assign_sites(
        sites,
        resolved_targets,
        geo,
        area,
        los_sd,
        use_work_history_los_sd=args.use_work_history_los_sd,
        default_area_sd=args.default_area_sd.strip(),
    )
    output_stats = write_workbook(
        Path(args.output),
        assignments,
        exceptions,
        los_sd,
        provider_resolution,
        existing_coverage if args.existing_coverage else None,
    )
    output_stats.update(assignment_stats)
    write_summary(
        Path(args.summary),
        args,
        targets,
        los_sd,
        site_stats,
        geo["stats"],
        provider_resolution,
        assignments,
        exceptions,
        output_stats,
        existing_coverage_stats,
    )

    print(f"=== {area} CBSA coverage build complete ===")
    print(f"Output: {args.output}")
    print(f"Summary: {args.summary}")
    print(f"Covered sites: {len(assignments):,}")
    print(f"Exception sites: {len(exceptions):,}")
    print(f"Coverage rows written: {output_stats['coverage_rows_written']:,}")
    print(f"Side-by-side coverage review rows written: {output_stats['coverage_review_rows_written']:,}")
    print(f"Unresolved target providers: {output_stats['unresolved_target_providers']:,}")
    if args.existing_coverage:
        print(f"Existing {area} coverage rows imported: {existing_coverage_stats['existing_area_rows']:,}")
        print(f"Current-only rows preserved in SP Numbers: {output_stats['current_only_rows_preserved_in_sp_numbers']:,}")


if __name__ == "__main__":
    main()
