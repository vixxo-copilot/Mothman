#!/usr/bin/env python3
"""Export Crystal's open Salesforce Case queue to Excel.

Breaks down owned open Cases by RecordType (case type) and Status, with
oldest CreatedDate at the top of each status group.

Intended daily refresh from mothman-good-morning:

  python .agents/skills/mothman-good-morning/scripts/export_sf_queue_workbook.py

Default outputs (overwrite stable + dated copy):
  .tmp/mothman-good-morning/Crystal-SF-Queue.xlsx
  .tmp/mothman-good-morning/Crystal-SF-Queue-YYYY-MM-DD.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_ROOT = SCRIPT_DIR.parent
# .../mothman/.agents/skills/mothman-good-morning → repo = parents[2]
REPO_ROOT = SKILL_ROOT.parents[2]
HELPERS = SKILL_ROOT.parent / "sf-case-email-sync" / "scripts"
sys.path.insert(0, str(HELPERS))

from sf_helpers import resolve_user_id, sf_query  # noqa: E402


def _chicago_tz():
    try:
        from zoneinfo import ZoneInfo

        return ZoneInfo("America/Chicago")
    except Exception:
        # Windows without tzdata: fixed Central offset (no DST flip — fine for as-of dating)
        return timezone(timedelta(hours=-5), name="America/Chicago")


TZ = _chicago_tz()
LIGHTNING_BASE = "https://vixxo.lightning.force.com/lightning/r/Case"
DEFAULT_OWNER_EMAIL = "Crystal.Gagner@vixxo.com"
OUTPUT_DIR = REPO_ROOT / ".tmp" / "mothman-good-morning"
STABLE_NAME = "Crystal-SF-Queue.xlsx"

# Display labels (RecordType.Name → sheet / summary label)
TYPE_LABELS = {
    "Rate Negotiation": "Rate Changes",
}
# Sheet / group order — Rate Changes first, then by typical volume
TYPE_ORDER = [
    "Rate Negotiation",
    "Service Provider Support",
    "Provider Onboarding",
    "Coverage Change",
]

# Preferred status order within a type (unknown statuses sort after, A–Z)
STATUS_ORDER = {
    "Rate Negotiation": [
        "New",
        "In Negotiation",
        "Approved",
        "Agreement Sent",
        "Declined",
        "Closed",
    ],
    "Service Provider Support": ["New", "Working", "Pending", "On Hold"],
    "Provider Onboarding": [
        "New",
        "Compliance In Progress",
        "Enablement",
        "Setup/Training",
        "Working",
    ],
    "Coverage Change": ["New", "Working", "Pending"],
}

HEADERS = [
    "Case Number",
    "Subject",
    "Status",
    "Priority",
    "SP Number",
    "Account",
    "Origin",
    "Created Date",
    "Age (Days)",
    "Last Modified",
    "Case URL",
]

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1A1228")
HEADER_FONT = Font(name="Calibri", bold=True, color="F0E6F6", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2A1F3D")
SECTION_FONT = Font(name="Calibri", bold=True, color="C4B5D8", size=11)
PRIORITY_FILL = PatternFill("solid", fgColor="E63E3E")
PRIORITY_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F7F4FA")
NEW_FILL = PatternFill("solid", fgColor="FFF3E0")
THIN = Border(
    left=Side(style="thin", color="D0C8D8"),
    right=Side(style="thin", color="D0C8D8"),
    top=Side(style="thin", color="D0C8D8"),
    bottom=Side(style="thin", color="D0C8D8"),
)
LINK_FONT = Font(name="Calibri", color="0563C1", underline="single")
BODY_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1A1228")
MUTED_FONT = Font(name="Calibri", size=10, color="666666")


def _today_central() -> date:
    return datetime.now(TZ).date()


def _parse_sf_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    raw = value.replace("+0000", "+00:00")
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ)


def _fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def _fmt_date(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d")


def _age_days(created: datetime | None, as_of: date) -> int | None:
    if not created:
        return None
    return (as_of - created.date()).days


def _type_label(record_type: str) -> str:
    return TYPE_LABELS.get(record_type, record_type or "(No Record Type)")


def _type_sort_key(record_type: str) -> tuple[int, str]:
    if record_type in TYPE_ORDER:
        return (TYPE_ORDER.index(record_type), record_type)
    return (len(TYPE_ORDER), record_type or "")


def _status_sort_key(record_type: str, status: str) -> tuple[int, str]:
    order = STATUS_ORDER.get(record_type, ["New", "Working"])
    if status in order:
        return (order.index(status), status)
    return (len(order), status or "")


def _sheet_name(record_type: str) -> str:
    label = _type_label(record_type)
    # Excel sheet name max 31 chars
    safe = label.replace("/", "-").replace("\\", "-")[:31]
    return safe or "Cases"


def fetch_open_cases(owner_id: str, org: str) -> list[dict[str, Any]]:
    soql = (
        "SELECT Id, CaseNumber, Subject, Status, Priority, Origin, Type, "
        "CreatedDate, LastModifiedDate, RecordType.Name, "
        "Account.Name, Account.Service_Provider_Number__c "
        "FROM Case "
        f"WHERE OwnerId = '{owner_id}' AND IsClosed = false "
        "ORDER BY CreatedDate ASC"
    )
    return sf_query(soql, org=org)


def normalize_case(row: dict[str, Any], as_of: date) -> dict[str, Any]:
    rt = ((row.get("RecordType") or {}).get("Name")) or "(No Record Type)"
    acct = row.get("Account") or {}
    created = _parse_sf_dt(row.get("CreatedDate"))
    modified = _parse_sf_dt(row.get("LastModifiedDate"))
    case_id = row.get("Id") or ""
    return {
        "id": case_id,
        "case_number": (row.get("CaseNumber") or "").lstrip("0") or row.get("CaseNumber") or "",
        "case_number_raw": row.get("CaseNumber") or "",
        "subject": row.get("Subject") or "",
        "status": row.get("Status") or "",
        "priority": row.get("Priority") or "",
        "origin": row.get("Origin") or "",
        "type_picklist": row.get("Type") or "",
        "record_type": rt,
        "label": _type_label(rt),
        "account": acct.get("Name") or "",
        "sp_number": acct.get("Service_Provider_Number__c") or "",
        "created": created,
        "created_str": _fmt_date(created),
        "modified_str": _fmt_dt(modified),
        "age_days": _age_days(created, as_of),
        "url": f"{LIGHTNING_BASE}/{case_id}/view" if case_id else "",
    }


def sort_cases(cases: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        cases,
        key=lambda c: (
            _type_sort_key(c["record_type"]),
            _status_sort_key(c["record_type"], c["status"]),
            c["created"] or datetime.max.replace(tzinfo=TZ),
            c["case_number_raw"],
        ),
    )


def _autosize(ws, min_width: int = 8, max_width: int = 56) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        widest = min_width
        for cell in ws[letter]:
            if cell.value is None:
                continue
            widest = max(widest, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = widest


def _write_headers(ws, row: int = 1) -> None:
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(HEADERS))}{row}"


def _write_case_row(ws, row_idx: int, case: dict[str, Any], alt: bool = False) -> None:
    values = [
        case["case_number"],
        case["subject"],
        case["status"],
        case["priority"],
        case["sp_number"],
        case["account"],
        case["origin"],
        case["created_str"],
        case["age_days"],
        case["modified_str"],
        case["url"],
    ]
    fill = NEW_FILL if case["status"] == "New" else (ALT_FILL if alt else None)
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.font = BODY_FONT
        cell.border = THIN
        cell.alignment = Alignment(vertical="center", wrap_text=(col == 2))
        if fill:
            cell.fill = fill
    # Hyperlink on Case Number + URL column
    if case["url"]:
        num_cell = ws.cell(row=row_idx, column=1)
        num_cell.hyperlink = case["url"]
        num_cell.font = LINK_FONT
        url_cell = ws.cell(row=row_idx, column=len(HEADERS))
        url_cell.hyperlink = case["url"]
        url_cell.font = LINK_FONT


def _write_section_banner(ws, row_idx: int, text: str, priority: bool = False) -> None:
    ws.merge_cells(
        start_row=row_idx,
        start_column=1,
        end_row=row_idx,
        end_column=len(HEADERS),
    )
    cell = ws.cell(row=row_idx, column=1, value=text)
    cell.fill = PRIORITY_FILL if priority else SECTION_FILL
    cell.font = PRIORITY_FONT if priority else SECTION_FONT
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row_idx].height = 18


def build_summary_sheet(
    wb: Workbook,
    cases: list[dict[str, Any]],
    as_of: date,
    owner_name: str,
) -> None:
    ws = wb.active
    ws.title = "Summary"

    by_type_status: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for c in cases:
        by_type_status[c["record_type"]][c["status"]].append(c)

    ws["A1"] = f"{owner_name} — Open SF Case Queue"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"As of {as_of.isoformat()} (America/Chicago) · "
        f"{len(cases)} open cases · sorted oldest CreatedDate within each Status"
    )
    ws["A2"].font = MUTED_FONT
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    # Totals by type
    ws["A4"] = "By Case Type"
    ws["A4"].font = Font(name="Calibri", bold=True, size=12)
    headers = ["Case Type", "Total Open", "New", "Oldest Created", "Newest Created", "Oldest New"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN

    type_keys = sorted(by_type_status.keys(), key=_type_sort_key)
    row = 6
    for rt in type_keys:
        statuses = by_type_status[rt]
        all_in_type = [c for lst in statuses.values() for c in lst]
        news = statuses.get("New", [])
        created_dates = [c["created"] for c in all_in_type if c["created"]]
        new_created = [c["created"] for c in news if c["created"]]
        label = _type_label(rt)
        vals = [
            label,
            len(all_in_type),
            len(news),
            _fmt_date(min(created_dates)) if created_dates else "",
            _fmt_date(max(created_dates)) if created_dates else "",
            _fmt_date(min(new_created)) if new_created else "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = THIN
            cell.font = BODY_FONT
            if rt == "Rate Negotiation":
                cell.fill = PatternFill("solid", fgColor="FDECEA")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="By Case Type × Status").font = Font(
        name="Calibri", bold=True, size=12
    )
    row += 1
    for col, h in enumerate(["Case Type", "Status", "Count", "Oldest", "Newest", "Max Age (Days)"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    row += 1

    for rt in type_keys:
        status_keys = sorted(
            by_type_status[rt].keys(),
            key=lambda s: _status_sort_key(rt, s),
        )
        for status in status_keys:
            group = by_type_status[rt][status]
            created_dates = [c["created"] for c in group if c["created"]]
            ages = [c["age_days"] for c in group if c["age_days"] is not None]
            vals = [
                _type_label(rt),
                status,
                len(group),
                _fmt_date(min(created_dates)) if created_dates else "",
                _fmt_date(max(created_dates)) if created_dates else "",
                max(ages) if ages else "",
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = THIN
                cell.font = BODY_FONT
                if status == "New":
                    cell.fill = NEW_FILL
            row += 1

    row += 2
    ws.cell(row=row, column=1, value="Sheets").font = Font(name="Calibri", bold=True, size=12)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=(
            "All Cases = full queue (type → status → oldest first). "
            "One sheet per case type with status section banners."
        ),
    ).font = MUTED_FONT

    _autosize(ws)
    ws.column_dimensions["B"].width = 48


def build_all_cases_sheet(wb: Workbook, cases: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("All Cases")
    # Extra Type column for the combined sheet
    all_headers = ["Case Type"] + HEADERS
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}1"

    for i, case in enumerate(cases):
        row_idx = i + 2
        ws.cell(row=row_idx, column=1, value=case["label"]).font = BODY_FONT
        ws.cell(row=row_idx, column=1).border = THIN
        # Shift case columns by 1
        values = [
            case["case_number"],
            case["subject"],
            case["status"],
            case["priority"],
            case["sp_number"],
            case["account"],
            case["origin"],
            case["created_str"],
            case["age_days"],
            case["modified_str"],
            case["url"],
        ]
        fill = NEW_FILL if case["status"] == "New" else (ALT_FILL if i % 2 else None)
        for col, value in enumerate(values, 2):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = BODY_FONT
            cell.border = THIN
            if fill:
                cell.fill = fill
        if case["url"]:
            num_cell = ws.cell(row=row_idx, column=2)
            num_cell.hyperlink = case["url"]
            num_cell.font = LINK_FONT
            url_cell = ws.cell(row=row_idx, column=len(all_headers))
            url_cell.hyperlink = case["url"]
            url_cell.font = LINK_FONT

    _autosize(ws)
    ws.column_dimensions["C"].width = 48


def build_type_sheet(
    wb: Workbook,
    record_type: str,
    cases: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(_sheet_name(record_type))
    priority = record_type == "Rate Negotiation"
    label = _type_label(record_type)

    ws["A1"] = f"{label} — {len(cases)} open"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    _write_headers(ws, row=2)
    # Freeze below title + header
    ws.freeze_panes = "A3"

    # Group by status (already globally sorted; re-group preserving order)
    by_status: dict[str, list] = defaultdict(list)
    for c in cases:
        by_status[c["status"]].append(c)
    status_keys = sorted(by_status.keys(), key=lambda s: _status_sort_key(record_type, s))

    row = 3
    for status in status_keys:
        group = by_status[status]
        # Within status: oldest CreatedDate first
        group_sorted = sorted(
            group,
            key=lambda c: (
                c["created"] or datetime.max.replace(tzinfo=TZ),
                c["case_number_raw"],
            ),
        )
        oldest = group_sorted[0]["created_str"] if group_sorted else ""
        banner = f"{status}  ·  {len(group_sorted)} cases  ·  oldest {oldest}"
        _write_section_banner(ws, row, banner, priority=(priority and status == "New"))
        row += 1
        for i, case in enumerate(group_sorted):
            _write_case_row(ws, row, case, alt=(i % 2 == 1))
            row += 1

    # Reset autofilter to cover data (skip title row)
    if row > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{row - 1}"
    _autosize(ws)
    ws.column_dimensions["B"].width = 52


def build_workbook(
    cases: list[dict[str, Any]],
    as_of: date,
    owner_name: str,
) -> Workbook:
    ordered = sort_cases(cases)
    wb = Workbook()
    build_summary_sheet(wb, ordered, as_of, owner_name)
    build_all_cases_sheet(wb, ordered)

    by_type: dict[str, list] = defaultdict(list)
    for c in ordered:
        by_type[c["record_type"]].append(c)

    for rt in sorted(by_type.keys(), key=_type_sort_key):
        build_type_sheet(wb, rt, by_type[rt])

    return wb


def export_queue(
    org: str = "vixxo",
    owner_email: str = DEFAULT_OWNER_EMAIL,
    output: Path | None = None,
    also_dated: bool = True,
) -> dict[str, Any]:
    as_of = _today_central()
    owner_id = resolve_user_id(org, owner_email)
    raw = fetch_open_cases(owner_id, org)
    cases = [normalize_case(r, as_of) for r in raw]
    wb = build_workbook(cases, as_of, owner_name=owner_email.split("@")[0].replace(".", " ").title())

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stable_path = output or (OUTPUT_DIR / STABLE_NAME)
    stable_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(stable_path)

    dated_path = None
    if also_dated:
        dated_path = OUTPUT_DIR / f"Crystal-SF-Queue-{as_of.isoformat()}.xlsx"
        wb.save(dated_path)

    # Breakdown for JSON / chat summary
    breakdown: dict[str, dict[str, int]] = defaultdict(dict)
    for c in cases:
        breakdown[c["label"]][c["status"]] = breakdown[c["label"]].get(c["status"], 0) + 1

    result = {
        "ok": True,
        "as_of": as_of.isoformat(),
        "owner_email": owner_email,
        "owner_id": owner_id,
        "total_open": len(cases),
        "breakdown": {k: dict(v) for k, v in breakdown.items()},
        "path": str(stable_path),
        "dated_path": str(dated_path) if dated_path else None,
    }
    return result


def main() -> int:
    p = argparse.ArgumentParser(description="Export open SF Case queue workbook")
    p.add_argument("--org", default="vixxo")
    p.add_argument("--owner-email", default=DEFAULT_OWNER_EMAIL)
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help=f"Stable workbook path (default: {OUTPUT_DIR / STABLE_NAME})",
    )
    p.add_argument(
        "--no-dated",
        action="store_true",
        help="Skip dated copy Crystal-SF-Queue-YYYY-MM-DD.xlsx",
    )
    p.add_argument("--json", action="store_true", help="Print result JSON to stdout")
    args = p.parse_args()

    try:
        result = export_queue(
            org=args.org,
            owner_email=args.owner_email,
            output=args.output,
            also_dated=not args.no_dated,
        )
    except Exception as exc:
        err = {"ok": False, "error": str(exc)}
        print(json.dumps(err, indent=2) if args.json else f"ERROR: {exc}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print(f"Wrote {result['total_open']} open cases → {result['path']}")
        if result.get("dated_path"):
            print(f"Dated copy → {result['dated_path']}")
        for label, statuses in result["breakdown"].items():
            parts = ", ".join(f"{s}: {n}" for s, n in statuses.items())
            print(f"  {label}: {parts}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
